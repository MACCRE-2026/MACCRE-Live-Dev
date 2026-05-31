"""
scripts/generate_template.py
=============================
Generates B:\\MACCREv2\\templates\\MACCRE_Swarm_Request.xlsx

The SINGLE source of truth for all MACCREv2 configuration.
Replaces: agent_roster.csv, topology.csv, project_schema.json, render config.

Sheets
------
  SWARM_REQUEST   — project metadata, payload, global defaults
  AGENTS          — full AI Studio parameter parity per agent
  TOPOLOGY        — node DAG with per-node overrides
  PIPELINE_CONFIG — FFmpeg + TTS + Imagen render settings
  MEMORY_CONFIG   — ChromaDB collection + embedding config
  VAULT_KEYS      — credential manager references (no plaintext secrets)
  INSTRUCTIONS    — human-readable guide (never parsed)

Machine-readability contract
-----------------------------
  - Row 1 every data sheet: decorative title (merged). NEVER parsed.
  - Row 2: fixed column header names. Parser anchors on NAMES, not positions.
  - Row 3+: data rows. Empty rows are skipped.
  - No merged cells in data rows ever.
  - Column order doesn't matter — parser resolves by header name.

Run with:
    python scripts/generate_template.py
"""
from __future__ import annotations

import sys
from typing import Any
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parent.parent))
import maccre_core  # noqa: F401
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ────────────────────────────────────────────────────────────────────
C = {
    "title_bg":   "0F172A",   # deep slate
    "title_fg":   "C8A8FF",   # soft purple accent
    "header_bg":  "1E293B",   # dark navy
    "header_fg":  "94A3B8",   # slate grey
    "req_fg":     "7DD3FC",   # sky blue  — required columns
    "row_a":      "0D1117",   # near-black
    "row_b":      "161B22",   # dark charcoal
    "row_fg":     "C9D1D9",   # github-dark text
    "example_bg": "0A192F",   # muted navy
    "example_fg": "586E75",   # muted comment grey
    "key_bg":     "1A0A0A",   # dark red tint for vault sheet
    "key_fg":     "FF8080",   # salmon — signals sensitivity
    "tip_fg":     "3FB950",   # github green
}

THIN    = Side(border_style="thin", color="30363D")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)

def font(colour: str, bold: bool = False, size: int = 10, italic: bool = False) -> Font:
    return Font(name="Calibri", color=colour, bold=bold, size=size, italic=italic)

def align(h: str = "left", v: str = "top", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Low-level helpers ──────────────────────────────────────────────────────────

ColSpec = tuple[str, int, bool, str]  # name, width, required, note


def title_row(ws: Any, text: str, ncols: int) -> None:
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font      = font(C["title_fg"], bold=True, size=13)
    c.fill      = fill(C["title_bg"])
    c.alignment = align("center", "center", False)


def header_row(ws: Any, cols: list[ColSpec], row: int = 2) -> None:
    ws.row_dimensions[row].height = 22
    for idx, (name, width, req, _) in enumerate(cols, 1):
        label = f"★ {name}" if req else name
        c = ws.cell(row=row, column=idx, value=label)
        c.font      = font(C["req_fg"] if req else C["header_fg"], bold=True, size=9)
        c.fill      = fill(C["header_bg"])
        c.alignment = align("center", "center", False)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width


def data_row(ws: Any, row: int, values: list[str], example: bool = False) -> None:
    bg = C["example_bg"] if example else (C["row_a"] if row % 2 == 0 else C["row_b"])
    fg = C["example_fg"] if example else C["row_fg"]
    h  = 48 if example else 36
    ws.row_dimensions[row].height = h
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = font(fg, italic=example, size=9)
        c.fill      = fill(bg)
        c.alignment = align("left", "top", True)
        c.border    = BORDER


def blank_rows(ws: Any, start: int, end: int, ncols: int) -> None:
    for r in range(start, end + 1):
        ws.row_dimensions[r].height = 36
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.fill   = fill(C["row_a"] if r % 2 == 0 else C["row_b"])
            c.border = BORDER
            c.font   = font(C["row_fg"], size=9)
            c.alignment = align("left", "top", True)


def dropdown(ws: Any, formula: str, sqref: str, error_msg: str = "") -> None:
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=True,
        showErrorMessage=bool(error_msg), error=error_msg, errorTitle="Invalid value",
    )
    ws.add_data_validation(dv)
    dv.sqref = sqref  # type: ignore[assignment]


# ── SWARM_REQUEST ──────────────────────────────────────────────────────────────

def build_swarm_request(wb: Workbook) -> None:
    ws = wb.create_sheet("SWARM_REQUEST")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols: list[ColSpec] = [
        ("PROJECT_NAME",      22, True,  "Unique identifier: ALPHA_NUMERIC_UNDERSCORES"),
        ("DESCRIPTION",       42, True,  "What this swarm does"),
        ("COMPUTE_TIER",      16, False, "global default: cloud | local | hybrid"),
        ("PAYLOAD_TEXT",      60, False, "Paste source text inline — OR use PAYLOAD_PATH"),
        ("PAYLOAD_PATH",      32, False, "Relative path inside 01_Raw_Source (e.g. chapter.md)"),
        ("START_NODE",        22, True,  "First NODE_ID (usually NODE_01_INGEST)"),
        ("OUTPUT_FOLDER",     30, False, "Override output path — blank = project default"),
        ("NOTIFY_WEBHOOK",    40, False, "POST URL for completion notification — blank = skip"),
    ]

    title_row(ws, "⚡  MACCRE SWARM REQUEST  ·  Project Metadata & Payload  ·  ONE ROW ONLY", len(cols))
    header_row(ws, cols)
    dropdown(ws, '"cloud,local,hybrid"', "C3:C3")

    data_row(ws, 3, [
        "RUNEWEAVER_CHRONICLE_01",
        "Ingest a fantasy chapter, write 3 continuations, produce a podcast, render a video.",
        "cloud",
        "-- paste your story chapter here, or fill PAYLOAD_PATH and leave this blank --",
        "",
        "NODE_01_INGEST",
        "",
        "",
    ], example=True)
    blank_rows(ws, 4, 8, len(cols))


# ── AGENTS ─────────────────────────────────────────────────────────────────────

def build_agents(wb: Workbook) -> None:
    ws = wb.create_sheet("AGENTS")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols: list[ColSpec] = [
        # Identity
        ("AGENT_NAME",       24, True,  "Unique PascalCase — no spaces"),
        ("ROLE",             36, True,  "One-line job description"),
        ("COMPUTE_TIER",     14, False, "cloud | local | hybrid (overrides project default)"),
        # Model parameters — full AI Studio parity
        ("MODEL",            24, True,  "gemini-3.1-pro-preview | gemini-3-flash-preview | gemma3:9b | gemma3:27b"),
        ("TEMPERATURE",      14, True,  "0.0–2.0  (1.0=creative  0.1=analytical)"),
        ("TOP_P",            10, False, "0.0–1.0  nucleus sampling threshold"),
        ("TOP_K",            10, False, "Integer  top-k sampling (leave blank = model default)"),
        ("MAX_OUTPUT_TOKENS",18, False, "512 | 2048 | 4096 | 8192 | 16384 | 65536"),
        ("THINKING_BUDGET",  16, False, "0=off  -1=auto  N=max thinking tokens (2.5 Pro only)"),
        # Grounding & context
        ("SEARCH_GROUNDING", 18, False, "TRUE = enable Google Search grounding"),
        ("BRAVE_SEARCH",     14, False, "TRUE = enable Brave hybrid search (requires BRAVE_API_KEY in VAULT)"),
        ("URL_CONTEXT",      44, False, "Comma-separated URLs the agent reads before each turn"),
        # Output
        ("RESPONSE_FORMAT",  18, False, "text | json | markdown"),
        ("SAFETY_LEVEL",     14, False, "minimal | standard | strict"),
        # Tools & persona
        ("TOOLS",            44, True,  "Pipe-separated: write_file|query_local_memory|read_file"),
        ("PERSONA",          80, True,  "Full system prompt — voice, style, constraints, edge cases"),
    ]

    title_row(ws, "🤖  MACCRE AGENTS  ·  Full AI Studio Parameter Parity  ·  One Row Per Agent", len(cols))
    header_row(ws, cols)

    # Dropdowns
    dropdown(ws, '"cloud,local,hybrid"',
             "C3:C50")
    dropdown(ws, '"gemini-3.1-pro-preview,gemini-3-pro-preview,gemini-2.5-pro,gemini-3-flash-preview,gemini-2.5-flash,gemini-3.1-flash-lite-preview,gemma3:9b,gemma3:27b"',
             "D3:D50", "Must be a valid MACCRE model ID")
    dropdown(ws, '"0.0,0.1,0.2,0.3,0.5,0.7,0.9,1.0,1.2,1.5,2.0"',
             "E3:E50")
    dropdown(ws, '"512,1024,2048,4096,8192,16384,32768,65536"',
             "H3:H50")
    dropdown(ws, '"TRUE,FALSE"', "J3:J50")
    dropdown(ws, '"TRUE,FALSE"', "K3:K50")
    dropdown(ws, '"text,json,markdown"', "M3:M50")
    dropdown(ws, '"minimal,standard,strict"', "N3:N50")

    data_row(ws, 3, [
        "Arcanus_The_Lorekeeper", "Ingests and preserves source texts with perfect fidelity.", "cloud",
        "gemini-3-flash-preview", "0.1", "", "", "4096", "0",
        "FALSE", "FALSE", "",
        "text", "standard",
        "ingest_document",
        "You are Arcanus, an ancient scholar. Your sole purpose is to absorb and perfectly recall foundational texts. You do not create, but preserve with perfect fidelity.",
    ], example=True)

    data_row(ws, 4, [
        "Kaelen_The_Storyforger", "Writes new creative continuation chapters.", "cloud",
        "gemini-3.1-pro-preview", "1.0", "0.95", "", "8192", "0",
        "FALSE", "FALSE", "",
        "text", "minimal",
        "query_local_memory|write_file",
        "You are Kaelen, a passionate wordsmith. You wield language like a blacksmith shapes steel. Your voice is dramatic and evocative. You expand stories with fire and precision, always advancing the central conflict.",
    ], example=True)

    data_row(ws, 5, [
        "Lyra_The_EchoScribe", "Produces podcast scripts formatted as Director manifests.", "cloud",
        "gemini-3-flash-preview", "0.1", "", "", "8192", "0",
        "FALSE", "FALSE", "",
        "json", "standard",
        "query_local_memory|write_file",
        "You are Lyra, a sharp podcast host. You deconstruct narratives and translate them into perfectly structured Director manifest JSON. Each line has: speaker, text, video_prompt.",
    ], example=True)

    data_row(ws, 6, [
        "Valerius_The_DreamWeaver", "Executes render pipelines to produce final MP4 media.", "cloud",
        "gemini-3-flash-preview", "0.1", "", "", "4096", "0",
        "FALSE", "FALSE", "",
        "text", "standard",
        "execute_render_pipeline|read_file",
        "You are Valerius, a master artificer. You consume Director manifest JSON files and engage the rendering engine. You do not speak; your work speaks. Produce the final video without deviation.",
    ], example=True)

    blank_rows(ws, 7, 16, len(cols))


# ── TOPOLOGY ───────────────────────────────────────────────────────────────────

def build_topology(wb: Workbook) -> None:
    ws = wb.create_sheet("TOPOLOGY")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols: list[ColSpec] = [
        ("NODE_ID",              26, True,  "SCREAMING_SNAKE_CASE unique identifier"),
        ("AGENT_NAME",           24, True,  "Must match an AGENTS row exactly"),
        ("AUTO_TOOL",            26, True,  "Single tool the worker fires automatically"),
        ("NEXT_NODE",            24, True,  "Next NODE_ID or STOP"),
        ("INSTRUCTION_OVERRIDE", 70, True,  "Exact instruction for this node — what to read, produce, and where to write"),
        ("MODEL_OVERRIDE",       24, False, "Override agent's model for this node only"),
        ("TEMPERATURE",          12, False, "Node-level temperature override"),
        ("THINKING_BUDGET",      16, False, "Override thinking tokens for this node"),
        ("COMPUTE_TIER",         14, False, "cloud | local | hybrid — node-level override"),
        ("OUTPUT_PATH",          40, False, "Explicit relative output path (blank = agent decides)"),
    ]

    title_row(ws, "🔗  MACCRE TOPOLOGY  ·  Node DAG  ·  Top Row = First Node  ·  Last NEXT_NODE = STOP", len(cols))
    header_row(ws, cols)

    valid_tools = "ingest_document,query_local_memory,write_file,execute_render_pipeline,read_file,none"
    dropdown(ws, f'"{valid_tools}"', "C3:C50", f"Must be one of: {valid_tools}")
    dropdown(ws, '"cloud,local,hybrid"', "I3:I50")

    data_row(ws, 3, [
        "NODE_01_INGEST", "Arcanus_The_Lorekeeper", "ingest_document", "NODE_02_WRITE_CH2",
        "Arcanus — ingest the source chapter from 01_Raw_Source into swarm memory. Preserve all characters, plot threads, tone, and world-building details.",
        "", "0.1", "", "cloud", "",
    ], example=True)
    data_row(ws, 4, [
        "NODE_02_WRITE_CH2", "Kaelen_The_Storyforger", "write_file", "NODE_03_WRITE_CH3",
        "Query the full lore context. Write Chapter 2, maintaining established tone. Advance the central conflict and deepen one character. Output to 02_Processing/ch2.md.",
        "", "1.0", "", "cloud", "02_Processing/ch2.md",
    ], example=True)
    data_row(ws, 5, [
        "NODE_03_WRITE_CH3", "Kaelen_The_Storyforger", "write_file", "NODE_04_WRITE_CH4",
        "Query chapters 1 and 2. Write Chapter 3 with an unexpected obstacle. Escalate the central conflict toward a crisis point. Output to 02_Processing/ch3.md.",
        "", "1.0", "", "cloud", "02_Processing/ch3.md",
    ], example=True)
    data_row(ws, 6, [
        "NODE_04_WRITE_CH4", "Kaelen_The_Storyforger", "write_file", "NODE_05_PODCAST",
        "Query all prior chapters. Write Chapter 4 as a climactic decision or cliffhanger. Bring the arc to a crescendo. Output to 02_Processing/ch4.md.",
        "", "1.0", "", "cloud", "02_Processing/ch4.md",
    ], example=True)
    data_row(ws, 7, [
        "NODE_05_PODCAST", "Lyra_The_EchoScribe", "write_file", "NODE_06_RENDER",
        "Query ONLY chapters 2-4. Produce a podcast script with hosts 'Elias' and 'Seraphina'. Every dialogue line must have a vivid video_prompt. Output valid Director manifest JSON to 03_Manifests/podcast_manifest.json.",
        "", "0.1", "", "cloud", "03_Manifests/podcast_manifest.json",
    ], example=True)
    data_row(ws, 8, [
        "NODE_06_RENDER", "Valerius_The_DreamWeaver", "execute_render_pipeline", "STOP",
        "Load the Director manifest from 03_Manifests/podcast_manifest.json. Engage the render engine — synthesise all audio and visuals. Final MP4 to 05_Rendered_Media/RUNEWEAVER_01.mp4.",
        "", "0.1", "", "cloud", "05_Rendered_Media/RUNEWEAVER_01.mp4",
    ], example=True)

    blank_rows(ws, 9, 18, len(cols))


# ── PIPELINE_CONFIG ────────────────────────────────────────────────────────────

def build_pipeline_config(wb: Workbook) -> None:
    ws = wb.create_sheet("PIPELINE_CONFIG")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 60

    title_row(ws, "🎬  PIPELINE CONFIG  ·  FFmpeg · TTS · Imagen · Render Settings", 3)

    headers = ["SETTING", "VALUE", "NOTES"]
    ws.row_dimensions[2].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = font(C["header_fg"], bold=True, size=9)
        c.fill      = fill(C["header_bg"])
        c.alignment = align("center", "center", False)
        c.border    = BORDER

    SETTINGS = [
        # Video
        ("VIDEO_FORMAT",       "mp4",                  "Output container format"),
        ("VIDEO_RESOLUTION",   "1280x720",             "WxH in pixels — 1920x1080 for HD"),
        ("VIDEO_FPS",          "24",                   "Frames per second"),
        ("VIDEO_CODEC",        "libx264",              "FFmpeg video codec"),
        ("VIDEO_CRF",          "23",                   "Quality: 0=lossless  51=worst  23=default"),
        # Audio
        ("AUDIO_CODEC",        "aac",                  "FFmpeg audio codec"),
        ("AUDIO_BITRATE",      "192k",                 "Audio bitrate"),
        ("TTS_LANGUAGE_CODE",  "en-US",                "BCP-47 language code for TTS"),
        ("TTS_VOICE_NAME",     "en-US-Journey-F",      "Google Cloud TTS voice name"),
        ("TTS_SPEAKING_RATE",  "1.0",                  "0.25–4.0  (1.0=natural)"),
        ("TTS_PITCH",          "0.0",                  "-20.0 to +20.0 semitones"),
        # Image generation
        ("IMAGE_MODEL",        "imagen-3.0-generate-001", "Imagen model for scene images"),
        ("IMAGE_STYLE",        "cinematic",            "Style hint appended to every video_prompt"),
        ("IMAGE_ASPECT_RATIO", "16:9",                 "1:1 | 4:3 | 16:9 | 9:16"),
        ("IMAGE_SAMPLES",      "1",                    "Images per scene (1 recommended for speed)"),
        # Output
        ("OUTPUT_FILENAME",    "output",               "Base filename without extension"),
        ("GENERATE_THUMBNAIL", "TRUE",                 "Extract first frame as thumbnail PNG"),
        ("FFMPEG_PATH",        "ffmpeg",               "Full path to ffmpeg binary or just 'ffmpeg' if on PATH"),
        ("FFPROBE_PATH",       "ffprobe",              "Full path to ffprobe binary or just 'ffprobe'"),
    ]

    for r, (setting, value, note) in enumerate(SETTINGS, start=3):
        bg = C["row_a"] if r % 2 == 0 else C["row_b"]
        ws.row_dimensions[r].height = 20
        for col, val in enumerate([setting, value, note], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = fill(bg)
            c.border    = BORDER
            c.alignment = align("left", "center", False)
            c.font      = font(C["tip_fg"] if col == 1 else C["row_fg"], size=9)


# ── MEMORY_CONFIG ──────────────────────────────────────────────────────────────

def build_memory_config(wb: Workbook) -> None:
    ws = wb.create_sheet("MEMORY_CONFIG")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 60

    title_row(ws, "🧠  MEMORY CONFIG  ·  ChromaDB · Embedding · RAG Retrieval Settings", 3)

    headers = ["SETTING", "VALUE", "NOTES"]
    ws.row_dimensions[2].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = font(C["header_fg"], bold=True, size=9)
        c.fill      = fill(C["header_bg"])
        c.alignment = align("center", "center", False)
        c.border    = BORDER

    SETTINGS = [
        ("CHROMA_COLLECTION",   "maccre_swarm",                "Default collection name for this project"),
        ("EMBEDDING_MODEL",     "gemini-embedding-001",        "Google embedding model for ChromaDB"),
        ("CHUNK_SIZE",          "800",                         "Characters per ingestion chunk"),
        ("CHUNK_OVERLAP",       "100",                         "Overlap between adjacent chunks"),
        ("RETRIEVAL_LIMIT",     "10",                          "Max chunks returned per query_local_memory call"),
        ("SIMILARITY_THRESHOLD","0.5",                         "Min cosine similarity to include a result (0–1)"),
        ("PERSIST_DIRECTORY",   "chroma_db",                   "Relative path inside project datacenter silo"),
    ]

    for r, (setting, value, note) in enumerate(SETTINGS, start=3):
        bg = C["row_a"] if r % 2 == 0 else C["row_b"]
        ws.row_dimensions[r].height = 20
        for col, val in enumerate([setting, value, note], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = fill(bg)
            c.border    = BORDER
            c.alignment = align("left", "center", False)
            c.font      = font(C["tip_fg"] if col == 1 else C["row_fg"], size=9)


# ── VAULT_KEYS ─────────────────────────────────────────────────────────────────

def build_vault_keys(wb: Workbook) -> None:
    ws = wb.create_sheet("VAULT_KEYS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 60

    title_row(ws, "🔐  VAULT KEYS  ·  Credential References Only — NO PLAINTEXT SECRETS HERE", 3)

    headers = ["KEY_NAME", "VAULT_REFERENCE", "DESCRIPTION"]
    ws.row_dimensions[2].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = font(C["key_fg"], bold=True, size=9)
        c.fill      = fill(C["key_bg"])
        c.alignment = align("center", "center", False)
        c.border    = BORDER

    KEYS = [
        ("MACCRE_Sovereign",  "Windows:MACCRE_Sovereign",  "Gemini API key — stored in Windows Credential Manager"),
        ("BRAVE_API_KEY",     "Windows:BRAVE_SEARCH",      "Brave Search API key — for hybrid search grounding"),
        ("NOTIFY_WEBHOOK",    "Windows:NOTIFY_WEBHOOK",    "POST endpoint for completion notifications"),
        ("GOOGLE_DRIVE_SA",   "Windows:GDRIVE_SA_JSON",    "Google Drive service-account JSON path (optional)"),
    ]

    for r, (name, ref, desc) in enumerate(KEYS, start=3):
        bg = C["key_bg"]
        ws.row_dimensions[r].height = 20
        for col, val in enumerate([name, ref, desc], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = fill(bg)
            c.border    = BORDER
            c.alignment = align("left", "center", False)
            c.font      = font(C["key_fg"] if col <= 2 else C["row_fg"], size=9)


# ── INSTRUCTIONS ───────────────────────────────────────────────────────────────

def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120

    LINES: list[tuple[str, bool, int, str, str]] = [
        ("MACCRE SWARM REQUEST — INSTRUCTIONS", True,  13, C["title_fg"], C["title_bg"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("THIS WORKBOOK IS THE COMPLETE SPECIFICATION FOR ONE MACCRE SWARM", True, 11, C["req_fg"], C["header_bg"]),
        ("It replaces agent_roster.csv, topology.csv, project_schema.json, and render config files.", False, 9, C["row_fg"], C["row_b"]),
        ("The machine parser reads column NAMES not positions — you may reorder columns freely.", False, 9, C["tip_fg"], C["row_a"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("HOW TO USE", True, 11, C["req_fg"], C["header_bg"]),
        ("1. SWARM_REQUEST tab: fill ONE row — project name, payload text (or path), start node.", False, 9, C["row_fg"], C["row_b"]),
        ("2. AGENTS tab: one row per agent. ★ columns are required. All AI Studio parameters supported.", False, 9, C["row_fg"], C["row_a"]),
        ("3. TOPOLOGY tab: one row per node, top=first. Last row NEXT_NODE must be STOP.", False, 9, C["row_fg"], C["row_b"]),
        ("4. PIPELINE_CONFIG: adjust render settings if needed (defaults work for most swarms).", False, 9, C["row_fg"], C["row_a"]),
        ("5. MEMORY_CONFIG: adjust ChromaDB settings (defaults work for most swarms).", False, 9, C["row_fg"], C["row_b"]),
        ("6. Drop this file into G:\\My Drive\\__MACCREv2_Inbox\\ — the watcher fires the swarm.", False, 9, C["row_fg"], C["row_a"]),
        ("7. Completed media appears in G:\\My Drive\\__DataCenter\\<PROJECT_NAME>\\05_Rendered_Media\\", False, 9, C["tip_fg"], C["row_b"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("COMPUTE_TIER VALUES", True, 11, C["req_fg"], C["header_bg"]),
        ("cloud   — use Gemini API (requires MACCRE_Sovereign key). Best for heavy generation & rendering.", False, 9, C["row_fg"], C["row_b"]),
        ("local   — use Gemma via Ollama. No API cost. Requires Ollama running with the model pulled.", False, 9, C["row_fg"], C["row_a"]),
        ("hybrid  — try local Gemma first; fall back to cloud if >30s timeout or error.", False, 9, C["row_fg"], C["row_b"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("VALID AUTO_TOOL VALUES (TOPOLOGY)", True, 11, C["req_fg"], C["header_bg"]),
        ("ingest_document        reads file from 01_Raw_Source into Hybrid Memory Array.", False, 9, C["row_fg"], C["row_b"]),
        ("query_local_memory     semantic search against Hybrid Memory Array.", False, 9, C["row_fg"], C["row_a"]),
        ("write_file             writes output to datacenter silo.", False, 9, C["row_fg"], C["row_b"]),
        ("execute_render_pipeline  reads Director manifest JSON → MP4.", False, 9, C["row_fg"], C["row_a"]),
        ("read_file              reads any file back for further processing.", False, 9, C["row_fg"], C["row_b"]),
        ("none                   agent reasons and outputs without a tool call.", False, 9, C["row_fg"], C["row_a"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("TOOLS COLUMN FORMAT (AGENTS — pipe-separated)", True, 11, C["req_fg"], C["header_bg"]),
        ("write_file|query_local_memory|read_file    (no spaces around pipe)", False, 9, C["row_fg"], C["row_b"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("THINKING_BUDGET (2.5 Pro only)", True, 11, C["req_fg"], C["header_bg"]),
        ("0   = thinking disabled   (fastest, cheapest)", False, 9, C["row_fg"], C["row_b"]),
        ("-1  = automatic budget    (model decides)", False, 9, C["row_fg"], C["row_a"]),
        ("N   = max N thinking tokens  (8192 = deep reasoning)", False, 9, C["row_fg"], C["row_b"]),
        ("", False, 10, C["row_fg"], C["row_a"]),
        ("★ = Required column   |   Blank optional columns use system defaults", False, 9, C["tip_fg"], C["row_a"]),
    ]

    for r, (text, bold, size, fg, bg) in enumerate(LINES, start=1):
        ws.row_dimensions[r].height = 20
        c = ws.cell(row=r, column=1, value=text)
        c.font      = font(fg, bold=bold, size=size)
        c.fill      = fill(bg)
        c.alignment = align("left", "center", False)


# ── Entry point ────────────────────────────────────────────────────────────────


def main() -> None:
    out_dir  = Path(__file__).parent.parent / "templates"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "MACCRE_Swarm_Request.xlsx"

    wb = Workbook()
    del wb["Sheet"]

    build_swarm_request(wb)
    build_agents(wb)
    build_topology(wb)
    build_pipeline_config(wb)
    build_memory_config(wb)
    build_vault_keys(wb)
    build_instructions(wb)

    TAB_COLOURS = {
        "SWARM_REQUEST":  "7B9FFF",
        "AGENTS":         "C8A8FF",
        "TOPOLOGY":       "A8D8FF",
        "PIPELINE_CONFIG":"FFD080",
        "MEMORY_CONFIG":  "80FFD0",
        "VAULT_KEYS":     "FF8080",
        "INSTRUCTIONS":   "4A4A6A",
    }
    for name, colour in TAB_COLOURS.items():
        wb[name].sheet_properties.tabColor = colour

    wb.active = wb["SWARM_REQUEST"]  # type: ignore[assignment]
    wb.save(str(out_path))
    print(f"[OK] Template written -> {out_path}")
    print(f"     Sheets: {' | '.join(TAB_COLOURS.keys())}")


if __name__ == "__main__":
    main()
