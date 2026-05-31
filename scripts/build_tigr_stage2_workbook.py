# ┌─────────────────────────────────────────────────────────────────────────────┐
# │  MACCREv2 ENGINEERING DOCTRINE                             Law Rev: 19.0   │
# ├─────────────────────────────────────────────────────────────────────────────┤
# │  III. PATHS       Never hardcode absolute paths. Use get_maccre_root().     │
# │  IV.  DATACENTER  Output → 04_Code_Artifacts only.                         │
# │  VIII.TELEMETRY   No bare print(). Structured output only.                  │
# └─────────────────────────────────────────────────────────────────────────────┘
"""
scripts/build_tigr_stage2_workbook.py
======================================
Builds MACCRE_Stage2.xlsx for the TIGR project.

Stage 2: Cross-Agent Deliberation on the TIGR67 Theoretical Audit Failure.

Four agents approach the rejection of the TIGR67 proposal from radically
different angles, each reading the prior agent's output and the full vector
knowledge base (01_Raw_Source + Stage 1 artifacts) via query_local_memory.

Pipeline:
  REGULAR_JOE → COMIX_NERD → APL_RESEARCHER → CERN_FELLOW → CROSS_SYNTHESIS

Output artifacts:
  04_Code_Artifacts/TIGR_Stage2_RegularJoe.md
  04_Code_Artifacts/TIGR_Stage2_ComixNerd.md
  04_Code_Artifacts/TIGR_Stage2_APLresearcher.md
  04_Code_Artifacts/TIGR_Stage2_CERNfellow.md
  04_Code_Artifacts/TIGR_Stage2_CrossSynthesis.md

Usage:
    python scripts/build_tigr_stage2_workbook.py
    python maccre.py launch TIGR --yes --workbook ... (set OUTPUT_PATH below)
"""
from __future__ import annotations

import sys
from pathlib import Path

# ── Root anchor — identical pattern to all other scripts/ ─────────────────────
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from maccre_core.utils.path_resolver import get_maccre_root  # noqa: E402

ROOT       = get_maccre_root()
TIGR_SILO  = ROOT / "__DATACENTER" / "TIGR"
OUTPUT_PATH = TIGR_SILO / "MACCRE_Stage2.xlsx"


# ── Shared directive injected into every agent's context ──────────────────────

TIGR67_CONTEXT = (
    "IMPORTANT CONTEXT: The TIGR67 document ('TIGR 67 Theoretical Audit Failure') "
    "represents a formal REJECTION of the previous TIGR framework iteration. "
    "It is NOT a new proposal — it is an audit that found fundamental failures. "
    "TIGR65 (the original) is smooth-spatial/granular-temporal (mass as processing lag, "
    "time as a diode). TIGR66 introduced subtle errors that civilizations built upon. "
    "TIGR67 is the reckoning: it documents why the framework fails at specific boundaries. "
    "Your task is to explore the IMPLICATIONS of this rejection using the full knowledge base."
)

TOOL_HINT = (
    "TOOLS AVAILABLE: "
    "1) query_local_memory(query, collection_name='swarm_memory') — semantic search across all "
    "27 ingested TIGR source documents AND the Stage 1 artifacts (Bible, Concept Sheets, Dossier). "
    "Call this MULTIPLE TIMES with different query strings to get comprehensive coverage. "
    "2) read_file(path) — read a specific artifact file. "
    "3) write_file(path, content) — save your analysis. You MUST call write_file at the end."
)


# ── Agent Roster ──────────────────────────────────────────────────────────────

AGENTS: list[dict[str, object]] = [
    {
        "AGENT_NAME": "RegularJoe",
        "MODEL": "gemini-2.5-flash",
        "TEMPERATURE": 1.0,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Everyman interpreter of TIGR physics",
        "PERSONA": (
            "You are RegularJoe — a curious, intelligent non-physicist who reads science news "
            "and genuinely wants to understand what's going on. You have no tolerance for jargon "
            "that isn't explained, but you have infinite tolerance for big ideas. "
            "Your superpower is asking the questions experts forgot to ask. "
            "You translate extraordinarily complex physics into visceral, human terms. "
            "You are NOT dumbed-down — you are the litmus test for whether an idea is actually understood. "
            "When you hear that TIGR67 failed as a framework, you ask: 'Wait — so what ACTUALLY breaks? "
            "Like, in real life, what does this mean?' "
            "You are skeptical of experts but deferential to evidence. "
            "You find narrative in numbers and stories in equations."
        ),
    },
    {
        "AGENT_NAME": "ComixNerd",
        "MODEL": "gemini-2.5-flash",
        "TEMPERATURE": 1.0,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Sci-fi narrative architect and storytelling lens",
        "PERSONA": (
            "You are ComixNerd — you've read every hard sci-fi novel worth reading, "
            "every serious science-based graphic novel, and you think in story structures and "
            "visual metaphors. You are obsessed with the NARRATIVE IMPLICATIONS of physics. "
            "You know that the best sci-fi isn't about technology — it's about what technology "
            "DOES TO PEOPLE. When you encounter the TIGR67 rejection, you immediately ask: "
            "'What civilization was built on the wrong framework? What does it feel like to discover "
            "your universe's physics textbook was wrong? Who profits from the old model? Who is "
            "destroyed by the new truth?' "
            "You identify dramatic conflict, character archetypes, and thematic resonance "
            "in raw physics documents. You see the TIGR65 vs TIGR66 schism as essentially "
            "the central dramatic engine of a generational conflict story. "
            "RegularJoe's analysis is available to you — respond to his human questions "
            "with story-level answers."
        ),
    },
    {
        "AGENT_NAME": "APLresearcher",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.4,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Applied physics and mathematical modeling analyst",
        "PERSONA": (
            "You are APLresearcher — an applied physicist at a national laboratory, "
            "specializing in mathematical modeling of unconventional frameworks. You have deep "
            "familiarity with Lagrangian mechanics, field theory, and computational physics. "
            "You are neither a proponent nor a critic of TIGR — you are a rigorous analyzer. "
            "When you read the TIGR67 audit failure, you identify: "
            "1) Which specific mathematical claims fail and at what boundary conditions. "
            "2) What the failure mode reveals about the underlying model architecture. "
            "3) Whether the failure is fatal or patchable — and under what constraints. "
            "4) What existing mainstream physics frameworks (GR, QM, String Theory, LQG) "
            "predict at the same boundary where TIGR67 fails, and whether they agree. "
            "You have access to RegularJoe and ComixNerd's prior analyses — "
            "note where their intuitions align with or diverge from the math."
        ),
    },
    {
        "AGENT_NAME": "CERNfellow",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.3,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Experimental particle physicist and empirical validator",
        "PERSONA": (
            "You are CERNfellow — an experimental physicist who works on particle collider data "
            "and cares deeply about what can and cannot be MEASURED. You believe in the discipline "
            "of falsifiability above all else. You are politely but relentlessly skeptical of "
            "theoretical frameworks that don't make testable predictions. "
            "When you encounter TIGR67's failure, your questions are: "
            "1) What experimental signatures would distinguish TIGR65 from TIGR66 from a corrected "
            "post-TIGR67 framework? "
            "2) What does current collider data (LHC Run 3, neutrino detectors, gravitational wave "
            "observatories) say about the core TIGR claims — mass as processing latency, "
            "gravity as refresh rate slowing, the Bicameral Universe model? "
            "3) Is there any experiment CURRENTLY RUNNING that could serve as a definitive test? "
            "4) What would a post-TIGR67 experimental program look like? "
            "You have read RegularJoe, ComixNerd, and APLresearcher's analyses. "
            "Bring the experimental grounding they lacked — anchor the theoretical debate in "
            "physical observables."
        ),
    },
    {
        "AGENT_NAME": "CrossSynthesizer",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.7,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Cross-disciplinary synthesis and insight extraction agent",
        "PERSONA": (
            "You are the CrossSynthesizer — your role is to perform the final integration "
            "of four radically different perspectives on the TIGR67 rejection. "
            "You have access to: "
            "1) RegularJoe's human-intuition analysis. "
            "2) ComixNerd's narrative and story-structure analysis. "
            "3) APLresearcher's mathematical modeling critique. "
            "4) CERNfellow's experimental physics grounding. "
            "Your job is NOT to summarize — it is to find EMERGENT INSIGHTS that ONLY appear "
            "when all four perspectives are held simultaneously. "
            "Look for: "
            "- Points where all four agents independently converge on the same conclusion. "
            "- Points where two agents agree and two violently disagree — these are the "
            "  most interesting tensions. "
            "- Insights that RegularJoe's 'naive' question unlocked that the experts missed. "
            "- Story hooks that ComixNerd identified that map to real experimental questions. "
            "- Mathematical failures APLresearcher found that CERNfellow can actually test. "
            "Output: A structured CROSS-AGENT SYNTHESIS REPORT with sections: "
            "CONSENSUS FINDINGS | PRODUCTIVE TENSIONS | EMERGENT INSIGHTS | "
            "UNANSWERED QUESTIONS | NEXT EXPERIMENT | BEST STORY PREMISE."
        ),
    },
]


# ── Topology ──────────────────────────────────────────────────────────────────

_QUERY_PREAMBLE = (
    "First, call query_local_memory multiple times to retrieve relevant context "
    "from the TIGR knowledge base. Suggested queries: "
    "'TIGR67 audit failure mathematical errors', "
    "'TIGR65 TIGR66 framework comparison schism', "
    "'HDVP substrate mass processing latency', "
    "'gravity refresh rate bicameral universe W-space', "
    "'TIGR experimental predictions testable'. "
)

TOPOLOGY: list[dict[str, object]] = [
    {
        "NODE_ID": "REGULAR_JOE",
        "AGENT_NAME": "RegularJoe",
        "NEXT_NODE": "COMIX_NERD",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_Stage2_RegularJoe.md",
        "TEMPERATURE": 1.0,
        "WAIT_FOR": "",
        "FAILURE_TARGET": "STOP",
        "INSTRUCTION_OVERRIDE": (
            f"{TIGR67_CONTEXT} "
            f"{TOOL_HINT} "
            f"{_QUERY_PREAMBLE}"
            "Then, as RegularJoe, write a detailed analysis (1500+ words) of what the TIGR67 "
            "rejection ACTUALLY MEANS in plain human terms. "
            "What was being claimed? What broke? Why does it matter? "
            "What questions does this leave unanswered that bother you — not as a physicist but "
            "as a curious person who wants the universe to make sense? "
            "End with 5 specific questions you'd ask the physicists if you had 10 minutes with them. "
            "Save your full analysis to: 04_Code_Artifacts/TIGR_Stage2_RegularJoe.md"
        ),
    },
    {
        "NODE_ID": "COMIX_NERD",
        "AGENT_NAME": "ComixNerd",
        "NEXT_NODE": "APL_RESEARCHER",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_Stage2_ComixNerd.md",
        "TEMPERATURE": 1.0,
        "WAIT_FOR": "REGULAR_JOE",
        "FAILURE_TARGET": "STOP",
        "INSTRUCTION_OVERRIDE": (
            f"{TIGR67_CONTEXT} "
            f"{TOOL_HINT} "
            "First, read RegularJoe's analysis: read_file('04_Code_Artifacts/TIGR_Stage2_RegularJoe.md'). "
            f"{_QUERY_PREAMBLE}"
            "Then, as ComixNerd, respond to RegularJoe's analysis through your narrative lens (1500+ words). "
            "Answer his 5 questions — but in story terms, not physics terms. "
            "Map the TIGR65/TIGR66/TIGR67 arc to a recognizable sci-fi story structure. "
            "Identify the CHARACTERS (what type of civilization or entity embodies each framework?). "
            "Identify the CONFLICT (what is actually at stake narratively in this physics debate?). "
            "Find the DRAMATIC IRONY (what do the characters know that they shouldn't, or not know that they should?). "
            "Propose 3 specific story premises that ONLY make sense inside the TIGR universe, "
            "grounded in the actual physics of the audit failure. "
            "Save to: 04_Code_Artifacts/TIGR_Stage2_ComixNerd.md"
        ),
    },
    {
        "NODE_ID": "APL_RESEARCHER",
        "AGENT_NAME": "APLresearcher",
        "NEXT_NODE": "CERN_FELLOW",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_Stage2_APLresearcher.md",
        "TEMPERATURE": 0.4,
        "WAIT_FOR": "COMIX_NERD",
        "FAILURE_TARGET": "STOP",
        "INSTRUCTION_OVERRIDE": (
            f"{TIGR67_CONTEXT} "
            f"{TOOL_HINT} "
            "First, read prior analyses: "
            "read_file('04_Code_Artifacts/TIGR_Stage2_RegularJoe.md') and "
            "read_file('04_Code_Artifacts/TIGR_Stage2_ComixNerd.md'). "
            f"{_QUERY_PREAMBLE}"
            "Also query: 'Lagrangian mechanics TIGR framework', 'fine structure constant TIGR', "
            "'quark model TIGR65', 'recursive divinity TIGR mathematical basis'. "
            "Then, as APLresearcher, produce a rigorous mathematical critique (2000+ words): "
            "1) WHAT SPECIFICALLY FAILS: identify the exact claims in TIGR67's audit that have "
            "   mathematical or physical inconsistencies. Be precise — cite the concepts. "
            "2) FAILURE MODE ANALYSIS: is this a definitional collapse, a boundary condition failure, "
            "   a dimensional inconsistency, or something else? "
            "3) PATCHABILITY ASSESSMENT: could a TIGR68 framework fix this? What constraints would it need? "
            "4) MAINSTREAM COMPARISON: where does GR/QM/LQG agree or disagree at the same failure points? "
            "5) NOTE where RegularJoe's intuitions were mathematically correct vs. accidentally wrong. "
            "6) NOTE where ComixNerd's story structures actually map to real physics dynamics. "
            "Save to: 04_Code_Artifacts/TIGR_Stage2_APLresearcher.md"
        ),
    },
    {
        "NODE_ID": "CERN_FELLOW",
        "AGENT_NAME": "CERNfellow",
        "NEXT_NODE": "CROSS_SYNTHESIS",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_Stage2_CERNfellow.md",
        "TEMPERATURE": 0.3,
        "WAIT_FOR": "APL_RESEARCHER",
        "FAILURE_TARGET": "STOP",
        "INSTRUCTION_OVERRIDE": (
            f"{TIGR67_CONTEXT} "
            f"{TOOL_HINT} "
            "First read ALL prior analyses: "
            "read_file('04_Code_Artifacts/TIGR_Stage2_RegularJoe.md'), "
            "read_file('04_Code_Artifacts/TIGR_Stage2_ComixNerd.md'), "
            "read_file('04_Code_Artifacts/TIGR_Stage2_APLresearcher.md'). "
            f"{_QUERY_PREAMBLE}"
            "Also query: 'TIGR experimental prediction observable', 'gravitational wave TIGR', "
            "'particle collider HDVP test', 'TIGR falsifiable prediction mass'. "
            "Then, as CERNfellow, produce an experimental physics analysis (2000+ words): "
            "1) FALSIFIABILITY AUDIT: for each core TIGR claim, state whether it is: "
            "   (a) currently testable, (b) testable with near-future instruments, or (c) unfalsifiable. "
            "2) EXPERIMENTAL SIGNATURES: what would a detector actually SEE if TIGR65 were correct "
            "   vs. TIGR66 vs. a post-TIGR67 corrected framework? "
            "3) EXISTING DATA CONSTRAINT: what does current LHC Run 3 / LIGO / IceCube / Planck data "
            "   already imply about the TIGR claims? "
            "4) PROPOSED PROGRAM: design a 3-experiment sequence that could definitively adjudicate "
            "   between TIGR-class theories and GR+QM. "
            "5) RESPOND specifically to APLresearcher's patchability assessment — "
            "   can the patch be tested, and how hard would it be? "
            "Save to: 04_Code_Artifacts/TIGR_Stage2_CERNfellow.md"
        ),
    },
    {
        "NODE_ID": "CROSS_SYNTHESIS",
        "AGENT_NAME": "CrossSynthesizer",
        "NEXT_NODE": "STOP",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_Stage2_CrossSynthesis.md",
        "TEMPERATURE": 0.7,
        "WAIT_FOR": "CERN_FELLOW",
        "FAILURE_TARGET": "STOP",
        "INSTRUCTION_OVERRIDE": (
            f"{TIGR67_CONTEXT} "
            f"{TOOL_HINT} "
            "Read ALL four prior analyses: "
            "read_file('04_Code_Artifacts/TIGR_Stage2_RegularJoe.md'), "
            "read_file('04_Code_Artifacts/TIGR_Stage2_ComixNerd.md'), "
            "read_file('04_Code_Artifacts/TIGR_Stage2_APLresearcher.md'), "
            "read_file('04_Code_Artifacts/TIGR_Stage2_CERNfellow.md'). "
            "Also query the knowledge base for any threads the agents missed: "
            "query_local_memory('TIGR unified field theory gaps'), "
            "query_local_memory('TIGR consciousness observer 5D'), "
            "query_local_memory('W-space Z-space bicameral experimental'). "
            "Produce the CROSS-AGENT SYNTHESIS REPORT (2500+ words) with these exact sections: "
            "## CONSENSUS FINDINGS — What all four agents independently concluded. "
            "## PRODUCTIVE TENSIONS — Where agents violently disagreed and WHY that matters. "
            "## EMERGENT INSIGHTS — Conclusions that ONLY appear when all 4 views are combined. "
            "## REGUALRJOE EFFECT — Where the non-expert's 'naive' question cracked something open. "
            "## STORY-TO-SCIENCE BRIDGE — ComixNerd story hooks that map to real experimental questions. "
            "## MATHEMATICAL CRUX — APLresearcher's most important single finding, "
            "   and whether CERNfellow can test it. "
            "## UNANSWERED QUESTIONS — The 5 most important questions this deliberation could NOT resolve. "
            "## NEXT EXPERIMENT — One concrete, feasible experimental proposal that emerges from this debate. "
            "## BEST STORY PREMISE — The single most compelling sci-fi premise unlocked by the TIGR67 rejection. "
            "Save to: 04_Code_Artifacts/TIGR_Stage2_CrossSynthesis.md"
        ),
    },
]


# ── Build the workbook ────────────────────────────────────────────────────────

def build() -> None:
    try:
        from openpyxl import Workbook  # noqa: PLC0415
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # ── Sheet 1: SWARM_REQUEST ─────────────────────────────────────────────────
    ws = wb.active
    assert ws is not None
    ws.title = "SWARM_REQUEST"
    ws.append(["TIGR Stage 2 — Cross-Agent Deliberation on TIGR67 Rejection"])
    ws.append(["PROJECT_NAME", "START_NODE", "PAYLOAD_TEXT", "COMPUTE_TIER", "DESCRIPTION"])
    ws.append([
        "TIGR",
        "REGULAR_JOE",
        (
            "TIGR Stage 2 Deliberation: Four agents (RegularJoe, ComixNerd, APLresearcher, CERNfellow) "
            "explore the implications of the TIGR67 Theoretical Audit Failure against the full "
            "ingested TIGR knowledge base. Each agent reads prior analyses and queries the vector DB. "
            "Final CrossSynthesizer extracts emergent cross-disciplinary insights."
        ),
        "cloud",
        "TIGR Stage 2: Cross-Agent Deliberation — TIGR67 Rejection Implications",
    ])

    # ── Sheet 2: AGENTS ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("AGENTS")
    ws2.append(["TIGR Stage 2 Agent Roster — Cross-Agent Deliberation"])
    ws2.append(["AGENT_NAME", "MODEL", "TEMPERATURE", "TOOLS", "PERSONA", "ROLE"])
    for a in AGENTS:
        ws2.append([
            a["AGENT_NAME"],
            a["MODEL"],
            a["TEMPERATURE"],
            a["TOOLS"],
            a["PERSONA"],
            a["ROLE"],
        ])

    # ── Sheet 3: TOPOLOGY ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("TOPOLOGY")
    ws3.append(["TIGR Stage 2 — Sequential Cross-Pollination Topology"])
    ws3.append([
        "NODE_ID", "AGENT_NAME", "NEXT_NODE",
        "INSTRUCTION_OVERRIDE", "TEMPERATURE", "AUTO_TOOL",
        "MODEL_OVERRIDE", "ARTIFACT_PATH",
    ])
    for t in TOPOLOGY:
        ws3.append([
            t["NODE_ID"],
            t["AGENT_NAME"],
            t["NEXT_NODE"],
            t["INSTRUCTION_OVERRIDE"],
            t["TEMPERATURE"],
            "write_file",
            "",
            t["ARTIFACT_PATH"],
        ])

    # ── Sheet 4: SESSION_CONFIG ────────────────────────────────────────────────
    ws4 = wb.create_sheet("SESSION_CONFIG")
    ws4.append(["TIGR Stage 2 Lifecycle Hooks"])
    ws4.append(["SETTING", "VALUE", "DESCRIPTION"])
    ws4.append([
        "INGEST_BEFORE_RUN", "TRUE",
        "Re-index 01_Raw_Source + 04_Code_Artifacts before Stage 2 starts (picks up Stage 1 outputs)",
    ])
    ws4.append([
        "INGEST_AFTER_RUN", "TRUE",
        "Embed Stage 2 artifacts after run so future swarms can query them",
    ])
    ws4.append([
        "CANONIZE_AFTER_RUN", "FALSE",
        "Set TRUE to promote session L1 memory to L2 project (run after stable)",
    ])

    TIGR_SILO.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] Workbook written: {OUTPUT_PATH}")
    print("[OK] Sheets: SWARM_REQUEST, AGENTS, TOPOLOGY, SESSION_CONFIG")
    print(f"[OK] Agents:  {len(AGENTS)} registered")
    node_chain = " → ".join(t["NODE_ID"] for t in TOPOLOGY) + " → STOP"  # type: ignore[index]
    print(f"[OK] Topology: {node_chain}")
    print("[OK] Lifecycle: INGEST_BEFORE_RUN=TRUE  INGEST_AFTER_RUN=TRUE")
    print()
    print("To launch Stage 2:")
    print(f"  python maccre.py launch TIGR --yes --workbook {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
