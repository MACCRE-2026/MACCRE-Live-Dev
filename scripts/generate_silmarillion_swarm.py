"""
scripts/generate_silmarillion_swarm.py
========================================
Generates B:\\MACCREv2\\MACCRE_Silmarillion.xlsx

Two-part production swarm for:
  Part A — Scholarly analysis of Silmarillion → LOTR narrative connections
  Part B — 10-minute video podcast with contextual image changes every 30s

Key lessons applied from NewsCast forensic autopsy:
  1. Director token starvation:
       The newscast Director produced only 13 / 40 scenes because its input
       context (all prior agent outputs) consumed most of the output budget.
       FIX: ScriptWriter writes a full ~1,400-word raw screenplay first.
            ManifestBuilder reads ONLY the screenplay file and converts it
            to JSON — its input context is therefore minimal and its full
            output budget is available for the manifest.

  2. Word count enforcement via recursion:
       10 min @ 130 wpm = 1,300 words minimum.
       A dedicated WordCountReviewer node reads the screenplay and either:
         - Approves it (routes to ManifestBuilder), or
         - Returns it to ScriptWriter with a specific expansion brief
           (max 3 recursions before forcing through).

  3. Scene math is hardcoded into the instruction:
       10 min = 600s / 30s per scene = 20 scene changes.
       Each scene ~65 words of dialogue.
       The manifest schema is given as a counting example in the instruction.

  4. Image prompts are decoupled from dialogue:
       Separate ImageDescriber node reads the screenplay chapter by chapter
       and writes 20 image prompts to a sidecar file.
       ManifestBuilder reads both files to assemble the final manifest,
       preventing the Director from trying to write 1,300 words of dialogue
       AND compose 20 image prompts in one generation pass.

Run with:
    python scripts/generate_silmarillion_swarm.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

sys.path.append(str(Path(__file__).resolve().parent.parent))
import maccre_core  # noqa: F401

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette (consistent with newscast template) ───────────────────────────────
C: dict[str, str] = {
    "title_bg":   "0F172A",
    "title_fg":   "C8A8FF",
    "header_bg":  "1E293B",
    "header_fg":  "94A3B8",
    "req_fg":     "7DD3FC",
    "row_a":      "0D1117",
    "row_b":      "161B22",
    "row_fg":     "C9D1D9",
    "example_bg": "0A192F",
    "example_fg": "586E75",
}

THIN   = Side(border_style="thin", color="30363D")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ColSpec = tuple[str, int, bool, str]


def fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def font(colour: str, bold: bool = False, size: int = 10, italic: bool = False) -> Font:
    return Font(name="Calibri", color=colour, bold=bold, size=size, italic=italic)


def align(h: str = "left", v: str = "top", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def title_row(ws: Any, text: str, ncols: int) -> None:
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font      = font(C["title_fg"], bold=True, size=13)
    c.fill      = fill(C["title_bg"])
    c.alignment = align("center", "center", False)


def header_row(ws: Any, cols: list[ColSpec], row: int = 2) -> None:
    ws.row_dimensions[row].height = 22
    for idx, (name, width, req, _) in enumerate(cols, 1):
        label = f"* {name}" if req else name
        c = ws.cell(row=row, column=idx, value=label)
        c.font      = font(C["req_fg"] if req else C["header_fg"], bold=True, size=9)
        c.fill      = fill(C["header_bg"])
        c.alignment = align("center", "center", False)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width


def data_row(ws: Any, row: int, values: list[str], example: bool = False) -> None:
    bg = C["example_bg"] if example else (C["row_a"] if row % 2 == 0 else C["row_b"])
    fg = C["example_fg"] if example else C["row_fg"]
    ws.row_dimensions[row].height = 60
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = font(fg, italic=example, size=9)
        c.fill      = fill(bg)
        c.alignment = align("left", "top", True)
        c.border    = BORDER


def blank_rows(ws: Any, start: int, end: int, ncols: int) -> None:
    for r in range(start, end + 1):
        ws.row_dimensions[r].height = 36
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.fill   = fill(C["row_a"] if r % 2 == 0 else C["row_b"])
            c.border = BORDER
            c.font   = font(C["row_fg"], size=9)
            c.alignment = align("left", "top", True)


def dropdown(ws: Any, formula: str, sqref: str, error_msg: str = "") -> None:
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=True,
        showErrorMessage=bool(error_msg), error=error_msg, errorTitle="Invalid value",
    )
    ws.add_data_validation(dv)
    dv.sqref = sqref  # type: ignore[assignment]


# ── SWARM_REQUEST ─────────────────────────────────────────────────────────────

def build_swarm_request(wb: Workbook) -> None:
    ws = wb.create_sheet("SWARM_REQUEST")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols: list[ColSpec] = [
        ("PROJECT_NAME",    22, True,  "Unique identifier"),
        ("DESCRIPTION",     55, True,  "What this swarm does"),
        ("COMPUTE_TIER",    14, False, "cloud | local | hybrid"),
        ("PAYLOAD_TEXT",    60, False, "Inline source text"),
        ("PAYLOAD_PATH",    32, False, "Relative path inside 01_Raw_Source"),
        ("START_NODE",      28, True,  "First NODE_ID"),
    ]

    title_row(ws, "MACCRE SWARM REQUEST  --  Silmarillion vs LOTR  --  Analysis + Podcast", len(cols))
    header_row(ws, cols)
    dropdown(ws, '"cloud,local,hybrid"', "C3:C3")

    data_row(ws, 3, [
        "SilmLOTR",
        "Part A: Scholarly analysis of how Silmarillion events tie into LOTR. "
        "Part B: 10-minute video podcast with contextual Imagen scene images every 30 seconds.",
        "cloud",
        "Begin the Silmarillion-to-LOTR analysis and podcast production pipeline.",
        "",
        "ANCHOR",
    ])
    blank_rows(ws, 4, 8, len(cols))


# ── AGENTS ────────────────────────────────────────────────────────────────────

def build_agents(wb: Workbook) -> None:
    ws = wb.create_sheet("AGENTS")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols: list[ColSpec] = [
        ("AGENT_NAME",        22, True,  "Unique PascalCase"),
        ("ROLE",              36, True,  "One-line job"),
        ("COMPUTE_TIER",      12, False, "cloud | local"),
        ("MODEL",             26, True,  "Model ID"),
        ("TEMPERATURE",       12, True,  "0.0-2.0"),
        ("MAX_OUTPUT_TOKENS", 18, False, "Token budget"),
        ("SEARCH_GROUNDING",  16, False, "TRUE/FALSE"),
        ("TOOLS",             44, True,  "Pipe-separated tools"),
        ("PERSONA",           100, True, "Full system prompt"),
    ]

    title_row(ws, "MACCRE AGENTS  --  Silmarillion/LOTR Podcast Swarm", len(cols))
    header_row(ws, cols)
    dropdown(ws, '"TRUE,FALSE"', "G3:G20")

    # ── Agent definitions ──────────────────────────────────────────────────────

    agents: list[list[str]] = [

        # 1 — Lore Researcher: deep-dives Silmarillion canon with search grounding
        [
            "LoreResearcher",
            "Tolkien Lore Researcher — Silmarillion Deep Dive",
            "cloud", "gemini-2.5-pro", "0.3", "8192", "TRUE",
            "write_file",
            "You are a Tolkien lore scholar and narrative archaeologist. Your specialty is tracing "
            "the causal chain from events in The Silmarillion into the age depicted in The Lord of "
            "the Rings. Use search grounding to pull in verified Tolkien scholarship, quotes, and "
            "chapter references. Produce exhaustive, citation-dense research documents. "
            "Do not summarize — go deep. Every claim must be traceable to a specific text. "
            "Write in academic prose. Minimum output: 600 words.",
        ],

        # 2 — Lore Analyst: thematic synthesis (no search, pure reasoning)
        [
            "LoreAnalyst",
            "Tolkien Lore Analyst — Thematic Connections",
            "cloud", "gemini-2.5-pro", "0.4", "8192", "FALSE",
            "write_file",
            "You are a literary analyst specializing in Tolkien's legendarium. You receive research "
            "documents and synthesize the thematic, symbolic, and narrative connections between "
            "The Silmarillion and The Lord of the Rings. Focus on: (1) the nature of evil as "
            "established through Morgoth's rebellion and inherited by Sauron; (2) the One Ring as "
            "a direct echo of Morgoth's strategy of dispersing power into Arda itself; (3) the "
            "fates of specific Silmarillion bloodlines (Elrond, Galadriel, the Numenorean descent "
            "to Aragorn); (4) the geographical memory embedded in Middle-earth's landscape. "
            "Write densely, with named chapter and character references. Minimum 600 words.",
        ],

        # 3 — Report Writer: Part A final scholarly document
        [
            "ReportWriter",
            "Scholarly Report Author — Part A Final Document",
            "cloud", "gemini-2.5-pro", "0.5", "8192", "FALSE",
            "write_file",
            "You are a senior Tolkien scholar writing a formal analytical report. You receive "
            "research and analysis from specialist agents and synthesize them into a comprehensive, "
            "well-structured scholarly essay. Structure: Introduction, Section 1 (The Inheritance "
            "of Morgoth — Sauron's nature and the One Ring), Section 2 (The Bloodline Thread — "
            "Elrond, Galadriel, Aragorn, the Numenorean connection), Section 3 (Geographical "
            "Memory — how Silmarillion-era events shaped the physical and spiritual landscape of "
            "LOTR), Section 4 (Thematic Resonance — the music of Ainulindale and the story's "
            "ending), Conclusion. Use proper headers. Aim for 1,200-1,500 words. "
            "Write to file: 04_Code_Artifacts/silmarillion_analysis.md",
        ],

        # 4 — Script Writer: writes the raw ~1,300-word podcast screenplay
        [
            "ScriptWriter",
            "Podcast Script Writer — 10-Minute Two-Host Dialogue",
            "cloud", "gemini-2.5-pro", "0.9", "16384", "FALSE",
            "write_file",
            "You are a professional podcast script writer specializing in engaging, intellectually "
            "stimulating dialogue between two hosts with distinct voices. "
            "HOST VOICE PROFILES: "
            "PROFESSOR (voice model: Fenrir) — measured, academic, loves citing specific "
            "Tolkien passages, occasionally gets genuinely excited when connecting dots. "
            "Speaks in complete paragraphs. "
            "VALE (voice model: Kore) — enthusiastic, asks the questions the listener is thinking, "
            "interrupts with 'Wait, hold on—' type energy, makes cultural connections to modern "
            "fantasy tropes. Shorter, punchier sentences. "
            "TIMING MANDATE: This script MUST produce exactly 10 minutes of audio when read aloud "
            "at a natural conversational pace. Average spoken words per minute is 130. "
            "TARGET: 1,300 words minimum, 1,500 words maximum. "
            "COUNT YOUR WORDS. If your script is under 1,200 words, expand the dialogue. "
            "STRUCTURE (20 scenes x ~30 seconds each): "
            "Scene 1-2: Intro hook and thesis. Scene 3-5: Morgoth/Sauron inheritance. "
            "Scene 6-8: The One Ring as corrupted Silmaril-logic. "
            "Scene 9-11: Bloodlines (Elrond, Aragorn, Galadriel). "
            "Scene 12-14: The geographical memory of Middle-earth. "
            "Scene 15-17: The Ainulindale echo in the story's music and ending. "
            "Scene 18-20: Wrap-up and listener call to reread. "
            "INTERRUPTION MANDATE: Include at least 4 natural interruptions where Vale cuts "
            "Professor off mid-sentence with 'Wait—' or 'Hold on—' or similar. "
            "FORMAT: Write as a screenplay with speaker labels. "
            "PROFESSOR: [dialogue] / VALE: [dialogue]. "
            "Mark scene boundaries with a comment: ## Scene N ##. "
            "Write to file: 04_Code_Artifacts/podcast_screenplay.md",
        ],

        # 5 — Word Count Reviewer: enforces the 1,300 word minimum via recursion
        [
            "WordCountReviewer",
            "Script Length Enforcer — Recursion Gate",
            "cloud", "gemini-2.5-flash", "0.1", "4096", "FALSE",
            "write_file",
            "You are a script quality gate. You receive a podcast screenplay and perform "
            "one job: count the total spoken words in the screenplay (exclude stage directions, "
            "comments, and speaker labels). "
            "DECISION RULES: "
            "If total spoken words >= 1,200: Write the word count to "
            "04_Code_Artifacts/word_count.md as: APPROVED: [N] words. "
            "If total spoken words < 1,200: Write a specific expansion brief to "
            "04_Code_Artifacts/word_count.md as: "
            "EXPAND_REQUIRED: [N] words found, need [1300-N] more words. "
            "Scenes that need expansion: [list specific scene numbers that are too short]. "
            "Do not rewrite anything. Just count and decide.",
        ],

        # 6 — Image Describer: writes 20 image prompts without composing dialogue
        [
            "ImageDescriber",
            "Visual Scene Director — 20 Imagen Prompt Composer",
            "cloud", "gemini-2.5-pro", "0.8", "8192", "FALSE",
            "write_file",
            "You are a visual storyboard director for a video podcast about Tolkien lore. "
            "You receive a podcast screenplay divided into 20 scenes. Your job is to write "
            "exactly 20 Imagen image prompts, one per scene, that visually represent what "
            "is being discussed in that scene. "
            "VISUAL STYLE: Epic fantasy oil painting style, warm candlelit tones, no text overlays, "
            "16:9 aspect ratio. Scenes should feel like illuminated manuscript illustrations "
            "come to life. Every 30 seconds (every scene) the image should change meaningfully. "
            "IMAGE MANDATE: Each prompt must be 2-3 sentences describing the scene content, "
            "style, lighting, and focal subject. Reference specific Tolkien visual iconography: "
            "Silmarils' light, the Two Trees, Mount Doom, Numenor, Angband, the White Tree, etc. "
            "OUTPUT FORMAT: Write a numbered list to 04_Code_Artifacts/image_prompts.md: "
            "Scene 1: [prompt] / Scene 2: [prompt] / ... / Scene 20: [prompt] "
            "Do not include any dialogue. Do not include any JSON. Only the numbered prompts.",
        ],

        # 7 — Manifest Builder: assembles final JSON from two sidecar files
        [
            "ManifestBuilder",
            "JSON Manifest Assembler — Podcast Director",
            "cloud", "gemini-2.5-pro", "0.1", "16384", "FALSE",
            "write_file",
            "You are a precision JSON manifest assembler. You receive TWO input files: "
            "(1) 04_Code_Artifacts/podcast_screenplay.md — the dialogue script with 20 scenes "
            "(2) 04_Code_Artifacts/image_prompts.md — 20 image prompts, one per scene "
            "Your job is to assemble a valid JSON array with exactly 20 objects. "
            "RULES: "
            "- Each JSON object represents one 30-second scene "
            "- 'speaker' field: either 'PROFESSOR' or 'VALE' based on who speaks in that scene "
            "- 'text' field: ALL spoken dialogue from that scene, combining multiple speaker "
            "  turns into a single string if needed. This is the TTS input. "
            "  Do not split scenes — each scene gets ONE object. "
            "- 'video_prompt' field: the corresponding image prompt from image_prompts.md "
            "OUTPUT: A single valid JSON array. No preamble. No explanation. "
            "No markdown code fences. Pure JSON starting with [ and ending with ]. "
            "Example structure: "
            '[{"speaker": "PROFESSOR", "text": "...", "video_prompt": "..."}, '
            '{"speaker": "VALE", "text": "...", "video_prompt": "..."}, ...] '
            "Write to: 04_Code_Artifacts/podcast_manifest.json",
        ],

        # 8 — Render Executor: triggers the render pipeline
        [
            "RenderDirector",
            "Render Pipeline Executor",
            "cloud", "gemini-2.5-flash", "0.1", "2048", "FALSE",
            "execute_render_pipeline",
            "You are the render executor. Your only job is to trigger the render pipeline. "
            "Read the file 04_Code_Artifacts/podcast_manifest.json. "
            "Call execute_render_pipeline with the exact JSON string from that file as the "
            "manifest_json argument. Do not modify the JSON. Do not add anything. "
            "Just read and execute.",
        ],
    ]

    for idx, agent in enumerate(agents, start=3):
        data_row(ws, idx, agent)

    blank_rows(ws, 3 + len(agents), 3 + len(agents) + 3, len(cols))


# ── TOPOLOGY ──────────────────────────────────────────────────────────────────

def build_topology(wb: Workbook) -> None:
    ws = wb.create_sheet("TOPOLOGY")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols: list[ColSpec] = [
        ("NODE_ID",              26, True,  "SCREAMING_SNAKE_CASE"),
        ("AGENT_NAME",           22, True,  "Must match AGENTS row"),
        ("AUTO_TOOL",            28, True,  "Single tool to fire"),
        ("NEXT_NODE",            28, True,  "Next NODE_ID or STOP"),
        ("INSTRUCTION_OVERRIDE", 90, True,  "Exact per-node instruction"),
        ("MODEL_OVERRIDE",       26, False, "Override agent model"),
        ("TEMPERATURE",          12, False, "Node temperature override"),
        ("MAX_RECURSION",        14, False, "Max loop bounces"),
    ]

    title_row(ws, "MACCRE TOPOLOGY  --  Silmarillion/LOTR  --  Analysis + Podcast Pipeline", len(cols))
    header_row(ws, cols)

    valid_tools = "ingest_document,query_local_memory,write_file,execute_render_pipeline,read_file,none,execute_hybrid_synthesis"
    dropdown(ws, f'"{valid_tools}"', "C3:C40", f"Must be one of: {valid_tools}")

    # ── Topology node definitions ─────────────────────────────────────────────
    #
    # Architecture decisions (applied lessons from NewsCast autopsy):
    #
    # 1. Two research streams run SERIALLY (not fan-out) to avoid broker
    #    complexity and keep payload chains tight.
    #
    # 2. ScriptWriter -> WordCountReviewer -> [APPROVED? ManifestBuilder : ScriptWriter]
    #    The recursion route goes BACK to SCRIPT_WRITER, not a new node.
    #    MAX_RECURSION=3 forces through after 3 bounces regardless.
    #
    # 3. ImageDescriber fires BEFORE ManifestBuilder, feeding a sidecar file.
    #    ManifestBuilder reads BOTH sidecar files. Its input context is tiny
    #    (just "read these two files and assemble JSON") — full output budget
    #    available for 20 JSON objects.
    #
    # 4. WordCountReviewer writes to word_count.md. The broker's standard
    #    recursion detection handles the routing; the reviewer writes
    #    APPROVED or EXPAND_REQUIRED as a signal.

    nodes: list[list[str]] = [

        # Anchor: opens the swarm, no generation, routes to research
        [
            "ANCHOR", "LoreResearcher", "none", "RESEARCH_SILM",
            "You are the ANCHOR node. Write a single line: 'ANCHOR: Pipeline launched. "
            "Beginning Silmarillion-to-LOTR analysis.' Route to the next node.",
            "", "0.1", "1",
        ],

        # Research Node 1: Silmarillion deep-dive
        [
            "RESEARCH_SILM", "LoreResearcher", "write_file", "RESEARCH_LOTR",
            "Perform a deep research dive into The Silmarillion's key events that directly "
            "influence The Lord of the Rings. Focus on: "
            "(1) Morgoth's original rebellion, the nature of his 'marring' of Arda, and how "
            "this prefigures Sauron's entire strategy and nature as a lieutenant carrying "
            "Morgoth's worldview forward. "
            "(2) The Silmarils themselves — their uncorruptible light, Feanor's oath, and "
            "how the logic of an object containing irreplaceable power directly anticipates "
            "the One Ring's narrative function. "
            "(3) The Fall of Numenor — how Sauron engineered it, why the Dunedain are "
            "exiles, and why Aragorn's kingship is a restoration across a 3,000-year wound. "
            "Use search grounding to pull verified Tolkien scholarship. "
            "Write to: 04_Code_Artifacts/research_silmarillion.md. Minimum 600 words.",
            "", "0.3", "1",
        ],

        # Research Node 2: LOTR through-line analysis
        [
            "RESEARCH_LOTR", "LoreAnalyst", "write_file", "REPORT",
            "Read the research document from 04_Code_Artifacts/research_silmarillion.md. "
            "Now perform deep thematic analysis of the through-lines from that research "
            "into The Lord of the Rings text. Specifically analyze: "
            "(1) Sauron as Morgoth's shadow — he cannot create, only corrupt and imitate. "
            "The One Ring is a Silmaril-logic object: a perfect vessel for concentrated "
            "will that enslaves its possessor. Trace this explicitly. "
            "(2) Elrond and Galadriel as living Silmarillion witnesses — what they personally "
            "experienced and how it shapes every decision they make in LOTR. "
            "(3) The geography of grief — Moria (Khazad-dum's fall), the Dead Marshes "
            "(Dagorlad battlefield), Weathertop (Amon Sul), Fangorn (Huorn-age forest). "
            "(4) The Ainulindale echo — the whole story is the Music being played out. "
            "Iluvatar's theme always overcomes Morgoth's discord. Write 600+ words to: "
            "04_Code_Artifacts/research_lotr_connections.md",
            "", "0.4", "1",
        ],

        # Part A Report: scholarly synthesis
        [
            "REPORT", "ReportWriter", "write_file", "SCRIPT_WRITE",
            "Read both research files: 04_Code_Artifacts/research_silmarillion.md and "
            "04_Code_Artifacts/research_lotr_connections.md. "
            "Write the final Part A scholarly report — a fully structured academic essay "
            "with Introduction, 4 numbered sections, and Conclusion. "
            "Section titles must be: "
            "1. The Inheritance of Evil: Morgoth, Sauron, and the Logic of the Ring. "
            "2. The Bloodline Thread: Elrond, Galadriel, Aragorn, and the Wound of Numenor. "
            "3. Landscape as Memory: The Silmarillion's Shadow on Middle-earth's Geography. "
            "4. The Music Resolves: Ainulindale and the Shape of the Story's Ending. "
            "Target 1,200-1,500 words. Use precise chapter and character citations. "
            "Write to: 04_Code_Artifacts/silmarillion_analysis.md",
            "", "0.5", "1",
        ],

        # Part B Script: 10-minute podcast screenplay
        [
            "SCRIPT_WRITE", "ScriptWriter", "write_file", "WORD_COUNT_GATE",
            "Read the analysis report from 04_Code_Artifacts/silmarillion_analysis.md. "
            "Write a complete 10-minute podcast screenplay for two hosts: PROFESSOR and VALE. "
            "## TIMING LAW ## "
            "10 minutes = 600 seconds. At 130 words per minute, you need 1,300 words of "
            "spoken dialogue. Do not write fewer than 1,300 words. Count your output. "
            "## STRUCTURE — 20 SCENES of ~65 words each ## "
            "Mark each scene boundary with: ## Scene N ## "
            "Scene 1-2: Hook — Vale asks 'If you've only read LOTR, what are you missing?' "
            "Scene 3-5: PROFESSOR explains Morgoth. Vale interrupts: 'Wait — so Sauron is "
            "basically a franchise evil?' "
            "Scene 6-8: The Ring as corrupted Silmaril-logic. Vale: 'Hold on — the Ring is "
            "basically what the Silmarils would become if Feanor had been Sauron?' "
            "Scene 9-11: Bloodlines. Elrond was AT the fall of Numenor. Galadriel knew "
            "Feanor. Aragorn is a 3,000-year restoration. "
            "Scene 12-14: Geography of grief. The Dead Marshes hold the faces of the "
            "Silmarillion dead. Moria's silence is 1,500 years old. "
            "Scene 15-17: The Ainulindale echo — the story's music. "
            "Scene 18-20: Wrap — 'If you haven't read The Silmarillion, read it now.' "
            "## VOICE RULES ## "
            "PROFESSOR: full paragraphs, quotes chapter references, gets excited. "
            "VALE: shorter, punchier. Interrupts at least 4 times with 'Wait—' or 'Hold on—'. "
            "Write to: 04_Code_Artifacts/podcast_screenplay.md",
            "", "0.9", "3",
        ],

        # Word count gate: recursion node
        [
            "WORD_COUNT_GATE", "WordCountReviewer", "write_file", "IMAGE_DESCRIBE",
            "Read the screenplay at 04_Code_Artifacts/podcast_screenplay.md. "
            "Count the total spoken words (exclude ## Scene N ## markers, speaker labels, "
            "and any stage directions in parentheses). "
            "If count >= 1,200 words: Write 'APPROVED: [N] words -- routing to image description.' "
            "to 04_Code_Artifacts/word_count.md and set your response to 'APPROVED'. "
            "If count < 1,200 words: Write 'EXPAND_REQUIRED: Only [N] words found. "
            "Need [1300-N] more. Thin scenes: [list scene numbers under 50 words].' "
            "to 04_Code_Artifacts/word_count.md and set your response to 'EXPAND_REQUIRED -- "
            "return to scriptwriter for expansion per word_count.md brief.' "
            "Note: if this is the 3rd recursion pass, approve regardless and write "
            "'FORCE_APPROVED: [N] words -- max recursion reached.'",
            "", "0.1", "3",
        ],

        # Image describer: writes 20 visual prompts as a sidecar
        [
            "IMAGE_DESCRIBE", "ImageDescriber", "write_file", "MANIFEST_BUILD",
            "Read the screenplay at 04_Code_Artifacts/podcast_screenplay.md. "
            "The screenplay is divided into 20 scenes marked with ## Scene N ##. "
            "For each of the 20 scenes, write one Imagen image prompt that visually "
            "depicts the Tolkien topics being discussed in that scene. "
            "STYLE: Epic fantasy oil painting, illuminated manuscript aesthetic, "
            "no text in image, warm gold and deep blue palette, 16:9 composition. "
            "CONTENT: Reference real Tolkien visual iconography — the light of the "
            "Silmarils, the Two Trees of Valinor, Angband's iron towers, the breaking "
            "of Numenor, Sauron's Eye, the One Ring in flame, Weathertop at dusk, "
            "the Dead Marshes, Minas Tirith's seven tiers, Galadriel's mirror, etc. "
            "OUTPUT: Exactly 20 numbered prompts. Format: "
            "'Scene 1: [2-3 sentence visual description]' each on its own line. "
            "Nothing else. No JSON. No dialogue. Write to: 04_Code_Artifacts/image_prompts.md",
            "", "0.8", "1",
        ],

        # Manifest Builder: assembles clean JSON from two sidecar files
        [
            "MANIFEST_BUILD", "ManifestBuilder", "write_file", "RENDER",
            "You must read two files and assemble a JSON manifest: "
            "FILE 1: 04_Code_Artifacts/podcast_screenplay.md (the dialogue) "
            "FILE 2: 04_Code_Artifacts/image_prompts.md (the 20 image prompts) "
            "Assemble a JSON array with exactly 20 objects. Rules: "
            "- Each object covers one ## Scene N ## from the screenplay "
            "- 'speaker': use the PRIMARY speaker for that scene ('PROFESSOR' or 'VALE') "
            "- 'text': ALL dialogue spoken in that scene, all speakers combined. "
            "  This is the raw TTS input. Include all words as-is. "
            "- 'video_prompt': the matching prompt from image_prompts.md "
            "Output ONLY the JSON array. Start with [ and end with ]. "
            "No preamble. No code fences. No explanation. "
            "Write the array to: 04_Code_Artifacts/podcast_manifest.json",
            "", "0.1", "1",
        ],

        # Render: executes the pipeline
        [
            "RENDER", "RenderDirector", "execute_render_pipeline", "STOP",
            "Read the file at 04_Code_Artifacts/podcast_manifest.json. "
            "Call the execute_render_pipeline tool with the manifest_json argument "
            "set to the EXACT string content of that file. "
            "Do not modify the JSON. Do not summarize. Do not add anything. "
            "Just pass the raw JSON string from the file to the tool.",
            "", "0.1", "1",
        ],
    ]

    for idx, node in enumerate(nodes, start=3):
        data_row(ws, idx, node)

    blank_rows(ws, 3 + len(nodes), 3 + len(nodes) + 3, len(cols))


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    build_swarm_request(wb)
    build_agents(wb)
    build_topology(wb)

    # Write into the project datacenter silo — this is where maccre.py looks
    out_dir = Path("B:/MACCREv2/__DATACENTER/SilmLOTR")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "MACCRE_Session.xlsx"
    wb.save(str(out_path))

    node_count = 9
    print(f"\n[OK] Silmarillion/LOTR Swarm Workbook generated: {out_path}")
    print(f"     Topology: {node_count} nodes (serial pipeline)")
    print("     Key fixes applied from NewsCast autopsy:")
    print("       - Director token starvation: ScriptWriter + ManifestBuilder separated")
    print("       - Word count enforcement: WordCountGate recursion loop (max 3)")
    print("       - Image prompts decoupled into ImageDescriber sidecar")
    print("       - ManifestBuilder reads 2 small files, full output budget for 20 JSON objects")
    print("\n     To run:  python maccre.py launch SilmLOTR")


if __name__ == "__main__":
    main()
