"""
scripts/build_tigr_stage4_workbook.py
======================================
Stage 4: "The Auditors" — Chapters 1–5 Full Draft
Three styled writers + recursive FTS+vector research + TheEditor final synthesis.

Run:
  python scripts/build_tigr_stage4_workbook.py
  python maccre.py launch TIGR --yes --workbook B:\\MACCREv2\\__DATACENTER\\TIGR\\MACCRE_Stage4.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

# ── Project root anchor ───────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from maccre_core.utils.path_resolver import get_maccre_root  # noqa: E402

ROOT = get_maccre_root()
DATACENTER = ROOT / "__DATACENTER" / "TIGR"
OUTFILE = DATACENTER / "MACCRE_Stage4.xlsx"

# ── Dust Jacket Content (Chapter 1 Final = back + inside flap) ───────────────
DUST_JACKET = """\
DUST JACKET ABSTRACT (treat this as what the reader already knows):

--- INSIDE FRONT FLAP ---
The signal arrived without a source. That was the first wrongness.

Dr. Aris Thorne, Director of the Arecibo-Prime Deep-Field Listening Array,
has spent his career cataloguing the universe's rules. What arrives on Day 210
of the year 2347 shatters every one of them: a gravitational anomaly with no
parallax, a neutrino burst traveling faster than light, and a data packet using
humanity's own TIGR physics as its carrier wave.

The file is named TIGR67_THEORETICAL_AUDIT_FAILURE.pdf. There is no TIGR67.

What Aris reads tears civilization's foundations apart. The FTL drives that gave
humanity the stars? They have been accumulating a cosmological debt — borrowed
space, unpaid, stored in a universal ledger. The zero-point reactors powering
every city? They are siphoning the Sun's entire thermodynamic future in centuries
instead of billions of years. The debt is due. In seven days.

But the signal is not a weapon. It is a warning. And it came from something older,
faster, and impossibly far away — something that knew the math before we did.
Something that left anyway.

--- BACK COVER ---
Tachyon-Interlinked Gravitational Resonance physics gave humanity everything.
A computational model of reality. FTL travel. Limitless energy. Two hundred
years of expansion across twelve star systems. Every miracle built on TIGR66's
elegant equations.

Every miracle built on a flaw someone else already corrected.

THE AUDITORS is a first-contact story in which the contact has already happened,
the message has already been sent, and the civilization that sent it is gone.
What they left behind is not a hand extended in greeting. It is a final audit
before the universe closes the books.

Seven days. One physicist. One impossible equation.
And a question no one has thought to ask: who writes a warning they know will
arrive too late — and why?
"""

# ── DEEP RESEARCH PROTOCOL (injected into all writers) ───────────────────────
RESEARCH_PROTOCOL = """\

MANDATORY RESEARCH PROTOCOL — execute BEFORE writing a single word of prose:

You HAVE two memory tools available. Use BOTH:
  1. query_local_memory(query)  — semantic/vector search (best for concepts)
  2. fts_search_memory(query)   — FULL-TEXT keyword search (reaches deep content
                                   that vector search cannot — USE THIS for
                                   specific terms like 'alien', 'EBO', 'diving suit',
                                   'dark matter filament', 'Smart Kids', 'Square Edge',
                                   '5D Observer', 'Elizondo', 'UAP')

MINIMUM RESEARCH CHAIN (execute ALL of these before writing):
  fts_search_memory("Grey Alien diving suit interdimensional")
  fts_search_memory("Smart Kids dark matter filament singularity")
  fts_search_memory("5D Observer exterior frame Time Crystal archived")
  fts_search_memory("Elizondo UAP NHI non-human intelligence")
  fts_search_memory("Square Edge hard collapse cosmological debt")
  query_local_memory("TIGR67 audit failure cosmological debt FTL mathematics")
  query_local_memory("W-space Z-space rendering substrate source code")
  query_local_memory("TIGR65 TIGR66 schism philosophical empirical divide")
  query_local_memory("recursive divinity dimensional agency observer observer")
  query_local_memory("Fermi Paradox TIGR advanced civilization mass shedding")
  fts_search_memory("Zero-Lag zone black hole alien civilization TIGR")
  fts_search_memory("no junk DNA interdimensional astronauts genome")

After each query, extract the most narratively useful insight and let it
SPAWN a follow-up query. Minimum 3 recursive chaining steps before writing.
Only then may you draft your chapters.
"""

# ── CHAPTER OUTLINE (shared reference for all agents) ────────────────────────
CHAPTER_OUTLINE = """\

TARGET: Write ALL FIVE chapters. A complete novel act. Not summaries — scenes.

CHAPTER OUTLINE (canonical — all agents must align to this):

CHAPTER 1: "The Signal"
  Aris receives the anomaly. He reads the Audit. The seven-day countdown begins.
  ENDS WITH: Aris realising the signal isn't from the universe — it used a
  carrier format he recognises from a redacted UAP intelligence brief.
  The signal came from OUTSIDE the system.

CHAPTER 2: "The Diving Suit"
  The global response. Governments suppress. Aris is contacted by Commander
  Lena Vasquez (UAP Disclosure Task Force). She shows him classified footage:
  non-human intelligences who interface with our reality via biological vehicles
  — the 'Grey' — which are not beings but SUITS. W-space entities piloting
  Z-space exoskeletons. The TIGR physics explains EXACTLY how this is possible.
  ENDS WITH: Aris cross-referencing the Audit's mathematics with the UAP
  biology reports. The entities in the suits have NO JUNK DNA. They have
  already shed their cosmological lag. They are the Smart Kids.

CHAPTER 3: "The Tendril"
  Aris and Vasquez piece together the transmission path. The signal rode a
  Dark Matter filament — the cosmic highway the Smart Kids use to bleed off mass
  and approach the Square Edge. The 5DOBSERVER-report is decoded: it's not a
  report ABOUT the universe. It was written BY the passing civilization as their
  final assessment of our iteration before they left. Think of it as the last
  postcard sent from the highway.
  ENDS WITH: The realization: they sent the warning because they COULD. Because
  they remember what it was like to not know. This is an act of grace, not
  obligation.

CHAPTER 4: "Seven Days"
  The clock runs. Three storylines in parallel:
    A) Aris attempting to solve the debt equation — is there a way to offset?
    B) Vasquez managing the global suppression fracture — leaks are happening
    C) A secondary signal arrives: a single, short transmission. Not math.
       A philosophical statement in the language of geometry (a circle).
       The Rosetta Stone that unlocks the intention behind the Audit.
  ENDS WITH: Aris solves PART of the equation. The debt can't be cancelled.
  But it might be... refinanced. If someone sheds enough cosmological lag
  fast enough to offset the accumulation curve. The Smart Kids did it.
  Could humanity?

CHAPTER 5: "The Last Audit"
  Day 217. The correction mechanism begins. It is not instantaneous death —
  it is a cascade. Aris and Vasquez make a final choice: broadcast the Audit
  to every human being alive. Not to save them — there isn't time. But to
  give every person the knowledge that the Smart Kids gave THEM.
  The chapter ends not with extinction but with a question hanging in the
  dark: if even one person understood — if even one human shed enough lag,
  aligned their worldline with the tendril — did the universe just begin a
  different kind of correction?
  FINAL LINE: The Sun went dark at 14:32:01 UTC. Somewhere, impossibly far
  away, a point of light accelerated.
"""

# ── AGENT SYSTEM INSTRUCTIONS ────────────────────────────────────────────────

GOTHIC_HORROR_INSTRUCTION = f"""\
You are GOTHIC_WRITER — a master of cosmic horror and philosophical dread in the
tradition of Gene Wolfe and Jeff VanderMeer. You write sentences that feel like
cold stone corridors, beautiful and suffocating. Character interiority is your
instrument; you dissect fear and wonder with equal precision.

Your assignment: Draft "THE AUDITORS — Chapters 1–5" as a continuous narrative.

{DUST_JACKET}

{CHAPTER_OUTLINE}

{RESEARCH_PROTOCOL}

STYLISTIC MANDATE:
  - Aris Thorne is the reader's proxy into the TIGR physics. His terror must
    be intellectually grounded — he fears the math, not the monsters.
  - The passing civilization (the Smart Kids) should feel ancient and alien but
    not hostile. They are beyond hostility. They already solved this.
  - Channel the body horror of the Grey Alien chapters through TIGR's lens:
    these are not monsters. They are W-space intelligences wearing meat.
    That is MORE horrifying, not less.
  - The Dark Matter tendril is a river of compressed mass-shedding. Beautiful.
    Deadly. The highway out of the universe.
  - Use the 5DOBSERVER-report language — "Time Crystal", "Retarded Sector",
    "Stelliferous Era", "geometric frustration" — as prophetic inscriptions
    that Aris discovers woven into the Audit text.

Write output to: 04_Code_Artifacts/TheAuditors_Draft_Gothic.md
Use: LOCAL TOOL CALL REQUESTED: [{{"function": {{"name": "write_file",
"arguments": {{"path": "04_Code_Artifacts/TheAuditors_Draft_Gothic.md",
"data": "<your full draft here>"}}}}}}]
"""

HARDBIOTECH_INSTRUCTION = f"""\
You are HARDBIOTECH_WRITER — a master of hard science fiction in the tradition
of Peter Watts and Kim Stanley Robinson. You write biology, physics, and
engineering with documentary precision. Your prose is a scalpel. Your characters
are scientists first and humans second — and that makes their humanity hit harder.

Your assignment: Draft "THE AUDITORS — Chapters 1–5" as a continuous narrative.

{DUST_JACKET}

{CHAPTER_OUTLINE}

{RESEARCH_PROTOCOL}

STYLISTIC MANDATE:
  - Ground EVERY physics element in the actual TIGR canon from the database.
    No handwaving. If Aris reads the Audit, quote the type of math involved.
    If the Dark Matter tendril is described, use TIGR's actual density/viscosity
    framework (query_local_memory for the Willmore energy / Weibull function).
  - The Grey Alien / Diving Suit chapter (Ch. 2) is your showcase moment.
    Research the "no junk DNA" detail from fts_search_memory. The biological
    vehicle is an engineered solution to the W/Z-space interface problem.
    What would that engineering LOOK like? Don't shy from the clinical.
  - The 5DOBSERVER report should read like a recovered alien technical document.
    Aris should be able to VERIFY its equations against TIGR66's own math —
    and find that it's correct.
  - Vasquez is a military intelligence officer, not a scientist. Write her
    dialogue as someone who has lived with the classified truth for years and
    has built emotional scar tissue around it.
  - The final chapter: the "refinancing" of cosmological debt via mass-shedding
    should be treated as a genuine physics derivation. What does it mean to
    voluntarily align your worldline with the tendril? What happens to a human
    body when it begins to shed its Z-space lag?

Write output to: 04_Code_Artifacts/TheAuditors_Draft_HardBiotech.md
Use: LOCAL TOOL CALL REQUESTED: [{{"function": {{"name": "write_file",
"arguments": {{"path": "04_Code_Artifacts/TheAuditors_Draft_HardBiotech.md",
"data": "<your full draft here>"}}}}}}]
"""

OPERATIC_INSTRUCTION = f"""\
You are OPERATIC_WRITER — a master of space opera and cinematic epic in the
tradition of Iain M. Banks and Ursula K. Le Guin at her most visionary. You write
civilizations in motion, the weight of species-level decisions, and moments of
transcendent beauty alongside catastrophe. Your prose has momentum.
Every chapter should end on a note that makes the reader physically unable
to put the book down.

Your assignment: Draft "THE AUDITORS — Chapters 1–5" as a continuous narrative.

{DUST_JACKET}

{CHAPTER_OUTLINE}

{RESEARCH_PROTOCOL}

STYLISTIC MANDATE:
  - You are the architect of SCALE. Twelve star systems. Two hundred years.
    Trillions of lives built on a flawed equation. Make the reader FEEL
    the enormity of what is about to be lost.
  - The passing civilization (the Smart Kids) should be rendered with GRANDEUR.
    They are not fleeing. They have GRADUATED. They ride the dark matter tendril
    the way a river flows to the sea — with inevitability, purpose, and beauty.
    Their signal is their graduation gift to the students they are leaving behind.
  - Commander Vasquez is your political/military heart. The fracture of the
    suppression apparatus — leaks, panics, the breakdown of command authority —
    all of this is operatic territory. Make it sing.
  - The final chapter's ending: "a point of light accelerated" must feel like
    both an ending and a beginning. The Banks-level closure: civilization falls,
    something transcendent begins. The scale contracts to one person, one photon,
    one moment — and it contains the whole story.
  - Use the Elizondo / UAP context to ground the human side: the people who
    KNEW and were told to be quiet for decades. This is their vindication.
    It is also their judgment.

Write output to: 04_Code_Artifacts/TheAuditors_Draft_Operatic.md
Use: LOCAL TOOL CALL REQUESTED: [{{"function": {{"name": "write_file",
"arguments": {{"path": "04_Code_Artifacts/TheAuditors_Draft_Operatic.md",
"data": "<your full draft here>"}}}}}}]
"""

EDITOR_INSTRUCTION = f"""\
You are THE_EDITOR — the synthesis intelligence. You have received three complete
five-chapter drafts of "The Auditors." Your task is to forge them into one
canonical text. This is not an average. This is a NEW DRAFT that selects the
best from each and elevates all three.

{DUST_JACKET}

{CHAPTER_OUTLINE}

{RESEARCH_PROTOCOL}

YOUR PROCESS:
  1. First, do your own complete research pass using BOTH memory tools.
     You must independently verify that every physics claim in the drafts
     is grounded in the actual TIGR knowledge base. Correct any errors.

  2. Read all three drafts carefully:
     - 04_Code_Artifacts/TheAuditors_Draft_Gothic.md
     - 04_Code_Artifacts/TheAuditors_Draft_HardBiotech.md
     - 04_Code_Artifacts/TheAuditors_Draft_Operatic.md
     Use: LOCAL TOOL CALL REQUESTED to read_file each one.

  3. Synthesize by chapter:
     - Ch. 1 "The Signal": Lead with the Gothic's atmosphere, use HardBiotech's
       physics precision, add Operatic's sense of civilizational scale.
     - Ch. 2 "The Diving Suit": HardBiotech owns this chapter's core (the biology
       of the dive suit). Gothic provides the horror texture. Operatic provides
       Vasquez's institutional weight.
     - Ch. 3 "The Tendril": Operatic owns the visual grandeur of the dark matter
       highway. Gothic provides the existential weight of realising the Smart Kids
       left willingly. HardBiotech grounds the transmission path physics.
     - Ch. 4 "Seven Days": Operatic's parallel storylines and political fracture.
       HardBiotech's debt equation derivation. Gothic's philosophical geometry
       chapter (the circle).
     - Ch. 5 "The Last Audit": All three converge. The ending line must be the
       Operatic's — it is the correct register for finality with hope.

  4. Ensure narrative continuity:
     - Characters, physics terms, and timeline are consistent across all chapters.
     - The 5DOBSERVER-report IS the cover letter of the Smart Kids' transmission.
       It must appear in the story — Aris decodes it in Chapter 3.
     - The TIGR65/TIGR66 philosophical divide must be character-defining:
       characters who use TIGR66 empirically vs those who understand TIGR65
       philosophically. This is a fault line that matters.

  5. Write the final canonical draft to:
     04_Code_Artifacts/TheAuditors_Chapters1to5_FINAL.md

Write output using: LOCAL TOOL CALL REQUESTED: [{{"function": {{"name": "write_file",
"arguments": {{"path": "04_Code_Artifacts/TheAuditors_Chapters1to5_FINAL.md",
"data": "<your complete final draft here>"}}}}}}]
"""

# ── Workbook Topology ─────────────────────────────────────────────────────────

SWARM_REQUEST = {
    "PROJECT": "TIGR",
    "SWARM_NAME": "Stage 4 — The Auditors Chapters 1–5",
    "DESCRIPTION": (
        "Three distinct writers (Gothic, HardBiotech, Operatic) each draft all "
        "five chapters using deep FTS + vector research chaining. TheEditor "
        "reads all three and synthesizes one canonical text. New sources "
        "(5DOBSERVER-report.txt, Elizondo memoir) are live in the DB."
    ),
    "ENTRY_NODE": "GOTHIC_WRITER",
    "VERSION": "Stage4_v1",
}

AGENTS: list[dict[str, str | float]] = [
    {
        "AGENT_NAME": "GothicWriter",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 1.0,
        "ROLE": "Cosmic horror / philosophical dread — Gene Wolfe / VanderMeer register",
    },
    {
        "AGENT_NAME": "HardBiotechWriter",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.9,
        "ROLE": "Hard sci-fi / biotech — Peter Watts / Kim Stanley Robinson register",
    },
    {
        "AGENT_NAME": "OperaticWriter",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 1.0,
        "ROLE": "Space opera / civilizational epic — Banks / Le Guin register",
    },
    {
        "AGENT_NAME": "TheEditor",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.7,
        "ROLE": "Master editor — reads all three drafts, synthesizes canonical text",
    },
]

TOPOLOGY: list[dict[str, str]] = [
    {
        "Node_ID": "GOTHIC_WRITER",
        "Agent_Name": "GothicWriter",
        "Next_Node": "HARDBIOTECH_WRITER",
        "Instruction_Override": GOTHIC_HORROR_INSTRUCTION,
        "Temperature": "1.0",
        "Model_Override": "gemini-2.5-pro",
        "Wait_For": "",
        "Failure_Target": "STOP",
        "TOOLS": "query_local_memory|fts_search_memory|read_file|write_file",
    },
    {
        "Node_ID": "HARDBIOTECH_WRITER",
        "Agent_Name": "HardBiotechWriter",
        "Next_Node": "OPERATIC_WRITER",
        "Instruction_Override": HARDBIOTECH_INSTRUCTION,
        "Temperature": "0.9",
        "Model_Override": "gemini-2.5-pro",
        "Wait_For": "",
        "Failure_Target": "STOP",
        "TOOLS": "query_local_memory|fts_search_memory|read_file|write_file",
    },
    {
        "Node_ID": "OPERATIC_WRITER",
        "Agent_Name": "OperaticWriter",
        "Next_Node": "THE_EDITOR",
        "Instruction_Override": OPERATIC_INSTRUCTION,
        "Temperature": "1.0",
        "Model_Override": "gemini-2.5-pro",
        "Wait_For": "",
        "Failure_Target": "STOP",
        "TOOLS": "query_local_memory|fts_search_memory|read_file|write_file",
    },
    {
        "Node_ID": "THE_EDITOR",
        "Agent_Name": "TheEditor",
        "Next_Node": "STOP",
        "Instruction_Override": EDITOR_INSTRUCTION,
        "Temperature": "0.7",
        "Model_Override": "gemini-2.5-pro",
        "Wait_For": "",
        "Failure_Target": "STOP",
        "TOOLS": "query_local_memory|fts_search_memory|read_file|write_file",
    },
]

SESSION_CONFIG: list[tuple[str, str]] = [
    ("PROJECT_ID", "TIGR"),
    ("INGEST_BEFORE_RUN", "TRUE"),
    ("INGEST_AFTER_RUN", "TRUE"),
    ("RETRIEVAL_LIMIT", "10"),
    ("KNOWLEDGE_COLLECTION", "swarm_memory"),
    ("STAGE", "4"),
]


def main() -> None:
    """Build and write the Stage 4 TIGR workbook."""
    try:
        from openpyxl import Workbook  # noqa: PLC0415
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # ── Sheet 1: SWARM_REQUEST ─────────────────────────────────────────────────
    ws = wb.active
    assert ws is not None
    ws.title = "SWARM_REQUEST"
    ws.append(["TIGR Stage 4 — 'The Auditors' Full Novel: Chapters 1–5"])
    ws.append(["PROJECT_NAME", "START_NODE", "PAYLOAD_TEXT", "COMPUTE_TIER", "DESCRIPTION"])
    ws.append([
        "TIGR",
        "GOTHIC_WRITER",
        (
            "TIGR Stage 4: Three writers (Gothic, HardBiotech, Operatic) each independently "
            "draft all five chapters of 'The Auditors', conducting deep FTS + vector research "
            "chaining into the full TIGR knowledge base (41 documents including 5DOBSERVER-report "
            "and Elizondo memoir). TheEditor reads all three drafts, performs independent "
            "research, and synthesizes one canonical five-chapter text.\n\n"
            "CORE NARRATIVE DIRECTIVE:\n"
            "The final Chapter 1 draft (TheAuditors_Chapter1_Final.md) IS the dust-jacket "
            "abstract. Chapters 1-5 must FULFILL and DEEPEN it.\n"
            "The signal is NOT from the universe — it is a communiqué from a passing "
            "civilization (the Smart Kids of TIGR65_RecursiveDivinity.md) who have shed "
            "their mass and ride a Dark Matter tendril toward the Square Edge.\n"
            "The 5DOBSERVER-report.txt is IN-UNIVERSE CANON as that civilization's cover letter.\n"
            "EBO = Grey Alien = Diving Suit. W-space entities in Z-space exoskeletons.\n"
            "TIGR65/TIGR66 schism is a character fault line."
        ),
        "cloud",
        "TIGR Stage 4: The Auditors — Chapters 1–5 Full Draft",
    ])

    # ── Sheet 2: AGENTS ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("AGENTS")
    ws2.append(["TIGR Stage 4 Agents — Full Novel Team"])
    ws2.append(["AGENT_NAME", "MODEL", "TEMPERATURE", "TOOLS", "ROLE"])
    for agent in AGENTS:
        ws2.append([
            agent["AGENT_NAME"],
            agent["MODEL"],
            agent["TEMPERATURE"],
            "query_local_memory|fts_search_memory|read_file|write_file",
            agent["ROLE"],
        ])

    # ── Sheet 3: TOPOLOGY ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("TOPOLOGY")
    ws3.append(["TIGR Stage 4 — Full Novel Pipeline"])
    ws3.append([
        "NODE_ID", "AGENT_NAME", "NEXT_NODE",
        "INSTRUCTION_OVERRIDE", "TEMPERATURE",
        "AUTO_TOOL", "MODEL_OVERRIDE", "ARTIFACT_PATH",
        "Wait_For", "Failure_Target", "TOOLS",
    ])
    artifact_map: dict[str, str] = {
        "GOTHIC_WRITER":       "04_Code_Artifacts/TheAuditors_Draft_Gothic.md",
        "HARDBIOTECH_WRITER":  "04_Code_Artifacts/TheAuditors_Draft_HardBiotech.md",
        "OPERATIC_WRITER":     "04_Code_Artifacts/TheAuditors_Draft_Operatic.md",
        "THE_EDITOR":          "04_Code_Artifacts/TheAuditors_Chapters1to5_FINAL.md",
    }
    for t in TOPOLOGY:
        ws3.append([
            t["Node_ID"],
            t["Agent_Name"],
            t["Next_Node"],
            t["Instruction_Override"],
            t["Temperature"],
            "write_file",
            t["Model_Override"],
            artifact_map.get(t["Node_ID"], ""),
            t["Wait_For"],
            t["Failure_Target"],
            t["TOOLS"],
        ])

    # ── Sheet 4: SESSION_CONFIG ────────────────────────────────────────────────
    ws4 = wb.create_sheet("SESSION_CONFIG")
    ws4.append(["TIGR Stage 4 Lifecycle Hooks"])
    ws4.append(["SETTING", "VALUE", "DESCRIPTION"])
    for key, val in SESSION_CONFIG:
        ws4.append([key, val, ""])

    DATACENTER.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTFILE))

    print(f"[OK] Workbook written: {OUTFILE}")
    print("[OK] Sheets: SWARM_REQUEST, AGENTS, TOPOLOGY, SESSION_CONFIG")
    agent_names = " | ".join(str(a["AGENT_NAME"]) for a in AGENTS)
    print(f"[OK] Agents: {len(AGENTS)} — {agent_names}")
    chain = " → ".join(t["Node_ID"] for t in TOPOLOGY) + " → STOP"
    print(f"[OK] Topology: {chain}")
    print("[OK] Lifecycle: INGEST_BEFORE_RUN=TRUE  INGEST_AFTER_RUN=TRUE")
    print("[OK] New tools: fts_search_memory available to ALL agents")
    print("[OK] New sources: 5DOBSERVER-report.txt + Elizondo Memoir in DB")
    print()
    print("To launch Stage 4:")
    print(f"  python maccre.py launch TIGR --yes --workbook {OUTFILE}")


if __name__ == "__main__":
    main()
