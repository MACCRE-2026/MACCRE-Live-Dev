"""
scripts/populate_alistair_test.py
===================================
Populates MACCRE_Global.xlsx for the AlistairFinch project.
- 6 agents: 3 Collaborators, 2 Synthesizers, 1 Final Editor
- Reads Chapter 1-4 and 5-7 from Desktop TestStory directory
- Constructs a 15-node linear topology representing 6 group brainstorm turns
  followed by a draft/refine/finalize arc for chapters 8, 9, 10.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT    = Path(__file__).parent.parent
WB_PATH = ROOT / "MACCRE_Global.xlsx"

DESKTOP_STORY_DIR = Path.home() / "Desktop" / "TestStory"
FILE_1 = DESKTOP_STORY_DIR / "Chapter_1_to_4.md"
FILE_2 = DESKTOP_STORY_DIR / "Chapter_5_to_7.md"

MODEL = "gemini-3.1-pro-preview"

AGENTS = [
    {
        "AGENT_NAME": "Collab_A",
        "ROLE": "Creative Collaborator",
        "PERSONA": "You are a creative narrative collaborator. Your job is to brainstorm and draft compelling story ideas, expanding the world and adding creative depth while respecting the existing narrative.",
    },
    {
        "AGENT_NAME": "Collab_B",
        "ROLE": "Creative Collaborator",
        "PERSONA": "You are a creative narrative collaborator. You critique, expand, and inject unique creative angles into the story, finding emotional resonance and character-driven moments.",
    },
    {
        "AGENT_NAME": "Collab_C",
        "ROLE": "Creative Collaborator",
        "PERSONA": "You are a creative narrative collaborator. You focus on pacing, thematic payoff, and writing engaging prose that pushes the plot forward naturally.",
    },
    {
        "AGENT_NAME": "Synth_Structure",
        "ROLE": "Synthesizer - Structure and Continuity",
        "PERSONA": "You are a master synthesizer focusing on structure, logical continuity, and narrative arc. Your job is to ensure all ideas fit seamlessly without plot holes, making the outline robust and tight.",
    },
    {
        "AGENT_NAME": "Synth_Tone",
        "ROLE": "Synthesizer - Tone and Delivery",
        "PERSONA": "You are a master synthesizer focusing on creativity and narrative delivery. Your sole purpose is to ensure the tone precisely matches the first 7 chapters, adjusting prose to sound authentic to the established world.",
    },
    {
        "AGENT_NAME": "Final_Editor",
        "ROLE": "Final Editor",
        "PERSONA": "You are the Final Editor and ultimate decision-maker. You take all previous feedback and write the definitive, finalized version of the story components. You have the final say on all creative choices.",
    }
]

TOPOLOGY = [
    # --- PHASE 1: 6 GROUP TURNS TO SETTLE ON A CONCLUSION ---
    {
        "NODE_ID": "PLAN_TURN_1",
        "AGENT_NAME": "Collab_A",
        "NEXT_NODE": "PLAN_TURN_2",
        "INSTRUCTION_OVERRIDE": "Review chapters 1-7. Propose a high-level creative direction for chapters 8, 9, and 10 to conclude this arc, but ensuring there is room left for a sequel concerning the Architect and his lab. Output your proposal.",
    },
    {
        "NODE_ID": "PLAN_TURN_2",
        "AGENT_NAME": "Synth_Structure",
        "NEXT_NODE": "PLAN_TURN_3",
        "INSTRUCTION_OVERRIDE": "Review Collab_A's proposal. Critique it for structural integrity and continuity with chapters 1-7. Re-organize the ideas into a solid, logical outline for chapters 8-10.",
    },
    {
        "NODE_ID": "PLAN_TURN_3",
        "AGENT_NAME": "Collab_B",
        "NEXT_NODE": "PLAN_TURN_4",
        "INSTRUCTION_OVERRIDE": "Review the structured outline. Add deep emotional beats, character moments, and unique creative twists to chapters 8-10. Enhance the sequel hook regarding the Architect's lab.",
    },
    {
        "NODE_ID": "PLAN_TURN_4",
        "AGENT_NAME": "Collab_C",
        "NEXT_NODE": "PLAN_TURN_5",
        "INSTRUCTION_OVERRIDE": "Review the expanded outline. Focus on pacing and thematic payoff. Adjust the flow to ensure a satisfying climax in Ch 9 and resolution in Ch 10. Output the improved outline.",
    },
    {
        "NODE_ID": "PLAN_TURN_5",
        "AGENT_NAME": "Synth_Tone",
        "NEXT_NODE": "PLAN_TURN_6",
        "INSTRUCTION_OVERRIDE": "Review the current outline. Analyze whether these ideas will match the established tone of chapters 1-7. Make adjustments to ensure the narrative delivery remains consistent with the original author's voice.",
    },
    {
        "NODE_ID": "PLAN_TURN_6",
        "AGENT_NAME": "Final_Editor",
        "NEXT_NODE": "DRAFT_CH_8",
        "INSTRUCTION_OVERRIDE": "You are the Final Editor. Review the multi-round brainstorming. Settle the conversation and lock in the exact, finalized outline for Chapters 8, 9, and 10 to conclude the story, establishing the Architect's lab sequel hook.",
    },
    # --- PHASE 2: CHAPTER 8 ---
    {
        "NODE_ID": "DRAFT_CH_8",
        "AGENT_NAME": "Collab_A",
        "NEXT_NODE": "REFINE_CH_8",
        "INSTRUCTION_OVERRIDE": "Using the finalized outline, draft the complete Chapter 8. Write prose only. Establish the beginning of the conclusion arc.",
    },
    {
        "NODE_ID": "REFINE_CH_8",
        "AGENT_NAME": "Synth_Tone",
        "NEXT_NODE": "FINAL_CH_8",
        "INSTRUCTION_OVERRIDE": "Review the draft of Chapter 8. Refine and rewrite sections to ensure the prose completely matches the tone and narrative delivery of chapters 1-7.",
    },
    {
        "NODE_ID": "FINAL_CH_8",
        "AGENT_NAME": "Final_Editor",
        "NEXT_NODE": "DRAFT_CH_9",
        "INSTRUCTION_OVERRIDE": "Review the refined Chapter 8. Make the final decisions, polish the prose, and output the definitive finalized version of Chapter 8.",
    },
    # --- PHASE 3: CHAPTER 9 ---
    {
        "NODE_ID": "DRAFT_CH_9",
        "AGENT_NAME": "Collab_B",
        "NEXT_NODE": "REFINE_CH_9",
        "INSTRUCTION_OVERRIDE": "Using the finalized outline and the finalized Chapter 8, draft the complete Chapter 9. Write prose only. This should be the climax of the story.",
    },
    {
        "NODE_ID": "REFINE_CH_9",
        "AGENT_NAME": "Synth_Structure",
        "NEXT_NODE": "FINAL_CH_9",
        "INSTRUCTION_OVERRIDE": "Review the draft of Chapter 9. Refine the structure, action flow, and continuity. Ensure it logically connects from Ch 8 and sets up Ch 10 perfectly.",
    },
    {
        "NODE_ID": "FINAL_CH_9",
        "AGENT_NAME": "Final_Editor",
        "NEXT_NODE": "DRAFT_CH_10",
        "INSTRUCTION_OVERRIDE": "Review the refined Chapter 9. Make the final decisions, polish the prose, and output the definitive finalized version of Chapter 9.",
    },
    # --- PHASE 4: CHAPTER 10 ---
    {
        "NODE_ID": "DRAFT_CH_10",
        "AGENT_NAME": "Collab_C",
        "NEXT_NODE": "REFINE_CH_10",
        "INSTRUCTION_OVERRIDE": "Using the finalized outline, C8, and C9, draft the complete Chapter 10. Write prose only. Conclude the narrative satisfyingly but introduce the sequel hook for the Architect and his lab.",
    },
    {
        "NODE_ID": "REFINE_CH_10",
        "AGENT_NAME": "Synth_Tone",
        "NEXT_NODE": "FINAL_CH_10",
        "INSTRUCTION_OVERRIDE": "Review the draft of Chapter 10. Refine the prose to perfectly match the emotional tone and narrative voice of the original chapters. Ensure the sequel hook is subtle and compelling.",
    },
    {
        "NODE_ID": "FINAL_CH_10",
        "AGENT_NAME": "Final_Editor",
        "NEXT_NODE": "DONE",
        "INSTRUCTION_OVERRIDE": "Review the refined Chapter 10. Make the final decisions, polish the prose, and output the definitive finalized version of Chapter 10. This is the end of the run.",
    },
]

def build_payload() -> str:
    parts: list[str] = []
    
    if FILE_1.exists():
        parts.append(FILE_1.read_text(encoding="utf-8").strip())
    else:
        print(f"Warning: {FILE_1} not found")
        
    if FILE_2.exists():
        parts.append(FILE_2.read_text(encoding="utf-8").strip())
    else:
        print(f"Warning: {FILE_2} not found")
        
    original = "\n\n---\n\n".join(parts)

    directive = (
        "USER DIRECTIVE — ALISTAIR FINCH COLLABORATIVE SESSIONS\n"
        "====================================================\n"
        "The original story chapters 1-7 are above. We need to conclude this story with chapters 8, 9, and 10.\n"
        "Your goal is to converse, outline, and write the definitive end of this narrative arc.\n"
        "Crucial Requirement: The story must close appropriately, but leave room for a sequel specifically concerning the Architect and his lab.\n"
    )

    return f"ORIGINAL STORY — CHAPTERS 1-7\n{'=' * 48}\n{original}\n\n{'=' * 48}\n\n{directive}"


def _clear_data_rows(ws: object, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for cell in ws[r]:  # type: ignore[index]
            cell.value = None  # type: ignore[union-attr]

def _header_col(ws: object, target: str, header_row: int = 2) -> int | None:
    for cell in ws[header_row]:  # type: ignore[index]
        if cell.value and target.lower() in str(cell.value).lower():
            return int(cell.column)
    return None

def _write_row(ws: object, row: int, data: dict[str, str], header_row: int = 2) -> None:
    for key, value in data.items():
        col = _header_col(ws, key, header_row)
        if col is not None:
            ws.cell(row=row, column=col, value=value)  # type: ignore[union-attr]


def populate(wb_path: Path) -> None:
    payload = build_payload()
    print(f"[POPULATE] Payload length: {len(payload)} chars")

    print(f"[POPULATE] Loading workbook: {wb_path}")
    wb = load_workbook(filename=str(wb_path))

    # PROJECT_DEFINITION
    ws_pd = wb["PROJECT_DEFINITION"]
    _clear_data_rows(ws_pd, 3, 12)
    kv: dict[str, str] = {
        "PROJECT_NAME":   "AlistairFinch",
        "DESCRIPTION":    "A collaborative 6-turn brainstorm and drafting session for Ch 8-10 with Architect lab sequel hook.",
        "VERSION":        "0.1",
        "AUTHOR":         "MACCREv2 Architect Swarm",
        "COMPUTE_TIER":   "cloud",
        "OUTPUT_FORMAT":  "markdown",
    }
    for idx, (k, v) in enumerate(kv.items()):
        ws_pd.cell(row=3 + idx, column=1, value=k)
        ws_pd.cell(row=3 + idx, column=2, value=v)
    
    # AGENTS
    ws_ag = wb["AGENTS"]
    _clear_data_rows(ws_ag, 3, 15)
    for i, ag in enumerate(AGENTS):
        _write_row(ws_ag, 3 + i, {
            "AGENT_NAME":  ag["AGENT_NAME"],
            "MODEL":       MODEL,
            "TEMPERATURE": "1.0",
            "TOOLS":       "write_file",
            "ROLE":        ag["ROLE"],
            "PERSONA":     ag["PERSONA"],
        })
        
    # TOPOLOGY
    ws_tp = wb["TOPOLOGY"]
    _clear_data_rows(ws_tp, 3, 25)
    for i, node in enumerate(TOPOLOGY):
        node["AUTO_TOOL"] = "write_file"
        node["MODEL_OVERRIDE"] = ""
        node["TEMPERATURE"] = "0.8" if "Final" in node["AGENT_NAME"] else "1.0"
        _write_row(ws_tp, 3 + i, node)

    # SWARM_REQUEST
    ws_sr = wb["SWARM_REQUEST"]
    _clear_data_rows(ws_sr, 3, 6)
    _write_row(ws_sr, 3, {
        "PROJECT_NAME":   "AlistairFinch",
        "DESCRIPTION":    "Drafting Chapters 8-10 with Architect lab sequel hook.",
        "COMPUTE_TIER":   "cloud",
        "PAYLOAD_TEXT":   payload,
        "PAYLOAD_PATH":   "",
        "START_NODE":     "PLAN_TURN_1",
        "OUTPUT_FOLDER":  "04_Code_Artifacts",
        "NOTIFY_WEBHOOK": "",
    })

    # EXECUTION_PLAN
    if "EXECUTION_PLAN" in wb.sheetnames:
        ws_ep = wb["EXECUTION_PLAN"]
        for row in ws_ep.iter_rows(min_row=3):
            first = row[0].value
            if first and "SWARM" in str(first).upper():
                if len(row) >= 3:
                    row[2].value = "YES"

    wb.save(str(wb_path))
    print(f"\n[POPULATE] Workbook saved: {wb_path}")

if __name__ == "__main__":
    populate(WB_PATH)
