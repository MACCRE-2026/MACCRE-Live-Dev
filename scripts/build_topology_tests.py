"""
scripts/build_topology_tests.py
================================
Builds test workbooks for T1-T7 edge-case topology validation.

Each test gets its own project silo under __DATACENTER and a populated
MACCRE_Swarm_Request.xlsx. Run all tests via:

    python maccre.py launch T1_SMOKE    --yes
    python maccre.py launch T2_FAILURE  --yes
    python maccre.py launch T3_DIAMOND  --yes
    python maccre.py launch T4_RACE     --yes
    python maccre.py launch T5_LOOP     --yes
    python maccre.py launch T6_VALIDATE --yes  # should FAIL at pre-flight
    python maccre.py launch T7_DIALOGUE --yes  # DialogueRunner mid-chain

MACCREv2 Law Rev 19.0 compliant.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

# ── Root anchor ───────────────────────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_SCRIPT_DIR.parent))

from maccre_core.utils.path_resolver import get_maccre_root

ROOT         = get_maccre_root()
DATACENTER   = ROOT / "__DATACENTER"
TEMPLATE_SRC = DATACENTER / "EXO_TEST" / "MACCRE_Swarm_Request.xlsx"

# ── openpyxl import ───────────────────────────────────────────────────────────
try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ImportError:
    print("[ERROR] openpyxl not installed.  Run: pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED MINIMAL AGENT PERSONA
# ══════════════════════════════════════════════════════════════════════════════
_TEST_PERSONA = (
    "You are a test agent in the MACCREv2 engine validation suite. "
    "Your job is to produce a SHORT, clear response that proves you received context correctly. "
    "State what you received, state your node name, and call write_file to save your output. "
    "Be concise — 3-5 sentences maximum. No filler."
)

_AGENT_COLS = [
    "AGENT_NAME", "ROLE", "COMPUTE_TIER", "MODEL", "TEMPERATURE",
    "TOP_P", "TOP_K", "MAX_OUTPUT_TOKENS", "THINKING_BUDGET",
    "SEARCH_GROUNDING", "BRAVE_SEARCH", "URL_CONTEXT",
    "RESPONSE_FORMAT", "SAFETY_LEVEL", "TOOLS", "PERSONA",
]

_TOPO_COLS = [
    "NODE_ID", "AGENT_NAME", "NEXT_NODE", "INSTRUCTION_OVERRIDE",
    "MODEL_OVERRIDE", "TEMPERATURE", "MAX_RECURSION",
    "WAIT_FOR", "FAILURE_TARGET", "ARTIFACT_PATH",
    "DIALOGUE_PARTNER", "DIALOGUE_ROUNDS",
]

_REQ_COLS = [
    "PROJECT_NAME", "DESCRIPTION", "COMPUTE_TIER",
    "PAYLOAD_TEXT", "PAYLOAD_PATH", "START_NODE",
    "OUTPUT_FOLDER", "NOTIFY_WEBHOOK",
]


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _copy_template(project_id: str) -> Path:
    """Copy EXO_TEST template xlsx into a fresh project silo."""
    dest_dir = DATACENTER / project_id
    dest_dir.mkdir(parents=True, exist_ok=True)
    (dest_dir / "01_Raw_Source").mkdir(exist_ok=True)
    dest_wb  = dest_dir / "MACCRE_Swarm_Request.xlsx"
    shutil.copy2(TEMPLATE_SRC, dest_wb)
    return dest_wb


def _hmap(ws: object) -> dict[str, int]:
    """Build header-name → 1-based column-index map from row 2."""
    m: dict[str, int] = {}
    for idx, cell in enumerate(ws[2], start=1):  # type: ignore[index]
        raw = str(cell.value or "").lstrip("★* ").strip().upper().replace(" ", "_")
        if raw:
            m[raw] = idx
    return m


def _ensure_cols(ws: object, needed: list[str]) -> None:
    """Append any missing columns to the header row."""
    hm = _hmap(ws)
    max_col: int = ws.max_column  # type: ignore[union-attr]
    for col in needed:
        if col.upper() not in hm:
            max_col += 1
            ws.cell(row=2, column=max_col, value=col)  # type: ignore[union-attr]


def _write_row(ws: object, row_idx: int, cols: list[str], values: list[str]) -> None:
    """Write values keyed by column name."""
    hm = _hmap(ws)
    for col, val in zip(cols, values):
        ci = hm.get(col.upper().replace(" ", "_"))
        if ci is not None:
            ws.cell(row=row_idx, column=ci, value=val)  # type: ignore[union-attr]


def _clear_data(ws: object) -> None:
    for row in ws.iter_rows(min_row=3):  # type: ignore[union-attr]
        for cell in row:
            cell.value = None  # type: ignore[union-attr]


def _populate(
    wb_path: Path,
    project_id: str,
    description: str,
    start_node: str,
    payload: str,
    agents: list[list[str]],
    topology: list[list[str]],
) -> None:
    """Populate all three sheets of a test workbook."""
    # Write payload text to input.md so the launch system finds it
    input_path = wb_path.parent / "01_Raw_Source" / "input.md"
    input_path.parent.mkdir(parents=True, exist_ok=True)
    input_path.write_text(payload, encoding="utf-8")

    wb = load_workbook(filename=str(wb_path))

    # ── SWARM_REQUEST ──────────────────────────────────────────────────────────
    ws_req = wb["SWARM_REQUEST"]
    _clear_data(ws_req)
    _write_row(ws_req, 3, _REQ_COLS, [
        project_id, description, "cloud",
        "", "input.md", start_node, "", "",
    ])

    # ── AGENTS ────────────────────────────────────────────────────────────────
    ws_ag = wb["AGENTS"]
    _clear_data(ws_ag)
    for i, agent in enumerate(agents, start=3):
        _write_row(ws_ag, i, _AGENT_COLS, agent)

    # ── TOPOLOGY ──────────────────────────────────────────────────────────────
    ws_tp = wb["TOPOLOGY"]
    _ensure_cols(ws_tp, ["MAX_RECURSION", "WAIT_FOR", "FAILURE_TARGET",
                          "ARTIFACT_PATH", "DIALOGUE_PARTNER", "DIALOGUE_ROUNDS"])
    _clear_data(ws_tp)
    for i, node in enumerate(topology, start=3):
        # Pad to 12 columns (matching TOPO_COLS)
        while len(node) < 12:
            node.append("")
        _write_row(ws_tp, i, _TOPO_COLS, node)

    wb.save(str(wb_path))
    wb.close()
    print(f"  [OK] {wb_path}")


# ══════════════════════════════════════════════════════════════════════════════
# T1 — SINGLE NODE SMOKE TEST
# ══════════════════════════════════════════════════════════════════════════════
def build_t1() -> None:
    """
    Topology:  START:[SOLO] → STOP

    Tests the absolute minimum: one agent, one write_file call, one artifact.
    Pass: artifact written, ledger created, cost > $0.
    """
    wb = _copy_template("T1_SMOKE")
    _populate(
        wb_path=wb,
        project_id="T1_SMOKE",
        description="Single node smoke test — minimum viable execution",
        start_node="SOLO",
        payload="Topology test T1. You are node SOLO. Confirm you received this message and write your output.",
        agents=[
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
        ],
        topology=[
            # NODE_ID, AGENT_NAME, NEXT_NODE, OVERRIDE, MODEL, TEMP, MAXREC, WAIT, FAIL, ARTIFACT, DLG_P, DLG_R
            ["SOLO", "TestAgent", "STOP",
             "You are node SOLO in topology test T1. Confirm receipt of this payload, "
             "state your node name, and call write_file to save your output to: "
             "04_Code_Artifacts/T1_SMOKE/solo_output.md",
             "", "0.7", "1", "none", "FAILED", "04_Code_Artifacts/T1_SMOKE/solo_output.md",
             "", ""],
        ],
    )


# ══════════════════════════════════════════════════════════════════════════════
# T2 — LINEAR CHAIN WITH FAILURE INJECTION
# ══════════════════════════════════════════════════════════════════════════════
def build_t2() -> None:
    """
    Topology:
        NODE_A → NODE_B (model_override=INVALID → forces RuntimeError)
                     ↓FAILURE_TARGET
               RECOVERY → STOP
        NODE_C should NOT execute (it is NODE_B's success target, never reached).

    Tests: FAILURE_TARGET routing fires on model exhaustion.
    Pass: RECOVERY ledger exists. NODE_C ledger does NOT exist.
    """
    wb = _copy_template("T2_FAILURE")
    _populate(
        wb_path=wb,
        project_id="T2_FAILURE",
        description="Failure injection — FAILURE_TARGET routing must fire when NODE_B exhausts its model chain",
        start_node="NODE_A",
        payload="Topology test T2. Chain: NODE_A → NODE_B (will fail) → RECOVERY.",
        agents=[
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
        ],
        topology=[
            ["NODE_A", "TestAgent", "NODE_B",
             "You are NODE_A. Confirm receipt and pass a short message downstream. "
             "Call write_file to save output to: 04_Code_Artifacts/T2_FAILURE/node_a.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T2_FAILURE/node_a.md", "", ""],
            # NODE_B uses a deliberately invalid model — all failover attempts will
            # exhaust and raise RuntimeError, which routes to FAILURE_TARGET=RECOVERY.
            ["NODE_B", "TestAgent", "NODE_C",
             "You are NODE_B. This node is INTENTIONALLY BROKEN for test T2. "
             "This text should never be sent to a model.",
             "invalid-model-t2-failtest-xyz9999", "0.7", "1",
             "none", "RECOVERY", "", "", ""],
            # NODE_C should NEVER execute — it is NODE_B's success path (unreachable).
            ["NODE_C", "TestAgent", "STOP",
             "You are NODE_C. If this executes, T2 FAILED — FAILURE_TARGET was not honoured.",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T2_FAILURE/node_c_SHOULD_NOT_EXIST.md", "", ""],
            ["RECOVERY", "TestAgent", "STOP",
             "You are the RECOVERY node. NODE_B failed as expected. "
             "Confirm you received the failure payload and call write_file to save: "
             "04_Code_Artifacts/T2_FAILURE/recovery.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T2_FAILURE/recovery.md", "", ""],
        ],
    )


# ══════════════════════════════════════════════════════════════════════════════
# T3 — DIAMOND FAN-OUT / FAN-IN
# ══════════════════════════════════════════════════════════════════════════════
def build_t3() -> None:
    """
    Topology:
              ┌→ BRANCH_A ─┐
    ROOT ───→ ┤             ├→ MERGE (wait_for=BRANCH_A,BRANCH_B) → STOP
              └→ BRANCH_B ─┘

    Tests: MERGE correctly blocks until BOTH branches complete.
    Pass: MERGE ledger contains [GATHERED ARTIFACT: BRANCH_A] AND [GATHERED ARTIFACT: BRANCH_B].
    """
    wb = _copy_template("T3_DIAMOND")
    _populate(
        wb_path=wb,
        project_id="T3_DIAMOND",
        description="Diamond fan-out/fan-in — MERGE must receive both branch artifacts via wait_for",
        start_node="ROOT",
        payload="Topology test T3. Diamond pattern: ROOT fans out to BRANCH_A and BRANCH_B, both feed MERGE.",
        agents=[
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
        ],
        topology=[
            ["ROOT", "TestAgent", "BRANCH_A,BRANCH_B",
             "You are ROOT in topology test T3. Generate a short research question on any topic. "
             "Call write_file to save it to: 04_Code_Artifacts/T3_DIAMOND/root_output.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T3_DIAMOND/root_output.md", "", ""],
            ["BRANCH_A", "TestAgent", "MERGE",
             "You are BRANCH_A. The [PREVIOUS NODE OUTPUT] above is the ROOT topic. "
             "Write a short response SUPPORTING the topic. "
             "Call write_file to save to: 04_Code_Artifacts/T3_DIAMOND/branch_a.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T3_DIAMOND/branch_a.md", "", ""],
            ["BRANCH_B", "TestAgent", "MERGE",
             "You are BRANCH_B. The [PREVIOUS NODE OUTPUT] above is the ROOT topic. "
             "Write a short response CHALLENGING the topic. "
             "Call write_file to save to: 04_Code_Artifacts/T3_DIAMOND/branch_b.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T3_DIAMOND/branch_b.md", "", ""],
            # MERGE waits for both branches — fan-in injection pre-loads both artifacts.
            ["MERGE", "TestAgent", "STOP",
             "You are MERGE. The [GATHERED ARTIFACT] blocks above contain outputs from "
             "BRANCH_A and BRANCH_B. Confirm BOTH are present by quoting one line from each. "
             "Call write_file to save your merged summary to: "
             "04_Code_Artifacts/T3_DIAMOND/merge_output.md",
             "", "0.7", "1", "BRANCH_A,BRANCH_B", "FAILED",
             "04_Code_Artifacts/T3_DIAMOND/merge_output.md", "", ""],
        ],
    )


# ══════════════════════════════════════════════════════════════════════════════
# T4 — UNEVEN DIAMOND (RACE CONDITION PROBE)
# ══════════════════════════════════════════════════════════════════════════════
def build_t4() -> None:
    """
    Topology:
              ┌→ FAST_NODE (1 sentence, max_recursion=1) ──────────────────────┐
    ROOT ───→ ┤                                                                 ├→ MERGE → STOP
              └→ SLOW_NODE (detailed multi-step, max_recursion=5) ─────────────┘

    Tests: MERGE does not fire on FAST_NODE completion — it waits for SLOW_NODE.
    Pass: MERGE receives both artifacts. No deadlock. No premature fire.
    """
    wb = _copy_template("T4_RACE")
    _populate(
        wb_path=wb,
        project_id="T4_RACE",
        description="Uneven diamond — MERGE must wait for slow branch, not fire early on fast branch",
        start_node="ROOT",
        payload="Topology test T4. One fast branch, one slow branch. MERGE waits for both.",
        agents=[
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
            ["DetailAgent", "Detailed analysis agent", "cloud",
             "gemini-2.5-flash", "1.0", "", "", "4096", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file",
             "You are a thorough analyst. Always produce detailed, multi-paragraph responses."],
        ],
        topology=[
            ["ROOT", "TestAgent", "FAST_NODE,SLOW_NODE",
             "You are ROOT in topology test T4. Write one sentence: 'The capital of France is Paris.' "
             "Call write_file to save to: 04_Code_Artifacts/T4_RACE/root_output.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T4_RACE/root_output.md", "", ""],
            # FAST_NODE: trivial task, completes in ~1 turn
            ["FAST_NODE", "TestAgent", "MERGE",
             "You are FAST_NODE. Write exactly one sentence confirming you ran. "
             "Call write_file to save to: 04_Code_Artifacts/T4_RACE/fast_output.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T4_RACE/fast_output.md", "", ""],
            # SLOW_NODE: detailed multi-step task with higher max_recursion
            ["SLOW_NODE", "DetailAgent", "MERGE",
             "You are SLOW_NODE in a race condition test. Write a detailed 5-paragraph analysis "
             "of why distributed systems require careful synchronization. Cover: (1) race conditions, "
             "(2) deadlocks, (3) starvation, (4) consistency models, (5) practical mitigations. "
             "Call write_file to save your complete analysis to: "
             "04_Code_Artifacts/T4_RACE/slow_output.md",
             "", "1.0", "5", "none", "FAILED",
             "04_Code_Artifacts/T4_RACE/slow_output.md", "", ""],
            ["MERGE", "TestAgent", "STOP",
             "You are MERGE in topology test T4. "
             "The [GATHERED ARTIFACT] blocks above contain FAST_NODE and SLOW_NODE outputs. "
             "Confirm BOTH are present. Quote the first line from each. "
             "Call write_file to save your confirmation to: "
             "04_Code_Artifacts/T4_RACE/merge_output.md",
             "", "0.7", "1", "FAST_NODE,SLOW_NODE", "FAILED",
             "04_Code_Artifacts/T4_RACE/merge_output.md", "", ""],
        ],
    )


# ══════════════════════════════════════════════════════════════════════════════
# T5 — BOUNDED LOOP (BACKWARD ROUTING TEST)
# ══════════════════════════════════════════════════════════════════════════════
def build_t5() -> None:
    """
    Topology:
        INIT → REFINER → JUDGE → STOP
                  ↑          |
                  └──────────┘ (JUDGE routes back to REFINER if not satisfied)

    JUDGE is instructed to reject the first REFINER output and route back.
    max_recursion=3 on JUDGE bounds the total loop depth.

    Tests: Whether backward routing (next_node pointing to a prior topology row)
    correctly re-enqueues the node without deadlocking.

    Pass:  REFINER ledger appears at least twice (different row numbers).
           JUDGE eventually reaches STOP. No deadlock.
    Fail signal: System hangs, or REFINER only appears once, or JUDGE never reaches STOP.
    """
    wb = _copy_template("T5_LOOP")
    _populate(
        wb_path=wb,
        project_id="T5_LOOP",
        description="Bounded backward-routing loop — JUDGE re-enqueues REFINER up to max_recursion times",
        start_node="INIT",
        payload="Topology test T5. REFINER produces a draft. JUDGE evaluates and routes back if insufficient. Loop is bounded.",
        agents=[
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
            ["JudgeAgent", "Strict quality judge", "cloud",
             "gemini-2.5-flash", "0.5", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file",
             "You are a strict quality judge. You reject the first attempt and accept the second. "
             "Be decisive. Route to REFINER if rejecting. Route to STOP if accepting."],
        ],
        topology=[
            ["INIT", "TestAgent", "REFINER",
             "You are INIT. Write a one-sentence prompt: 'Write a haiku about distributed systems.' "
             "Call write_file to save to: 04_Code_Artifacts/T5_LOOP/init.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T5_LOOP/init.md", "", ""],
            ["REFINER", "TestAgent", "JUDGE",
             "You are REFINER writing a haiku about distributed systems. "
             "TWO-PHASE RULE — read your context first: "
             "Phase 1 (no CRITIQUE block in context): Write a haiku where the MIDDLE line "
             "has exactly 6 syllables (intentionally wrong — do NOT correct this). "
             "Example structure: Line1=5syl, Line2=6syl(wrong), Line3=5syl. "
             "Call write_file to save to: 04_Code_Artifacts/T5_LOOP/refiner_draft.md "
             "Phase 2 (CRITIQUE block IS in context): Fix the middle line to exactly 7 syllables. "
             "Prefix your haiku with 'REVISED:'. "
             "Call write_file to overwrite: 04_Code_Artifacts/T5_LOOP/refiner_draft.md",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T5_LOOP/refiner_draft.md", "", ""],
            # JUDGE: rejects when middle line ≠ 7 syllables, accepts when it sees 'REVISED:'.
            # CRITICAL: emit ROUTE_TO: as the FIRST word of plain text — no write_file on reject.
            ["JUDGE", "JudgeAgent", "STOP",
             "You are JUDGE evaluating a haiku. Read the [PREVIOUS NODE OUTPUT] haiku carefully. "
             "TASK: Count the syllables in each line of the haiku. "
             "DECISION RULE — choose exactly one path: "
             "PATH A — The middle line does NOT have 7 syllables (it has 6 or another wrong count): "
             "Your entire response must be: ROUTE_TO:REFINER\n\nCRITIQUE: [count each line's syllables, state which line is wrong and by how many]. "
             "Do NOT call write_file in this path. "
             "PATH B — The haiku starts with 'REVISED:' AND the middle line has exactly 7 syllables: "
             "Your response must start with: ROUTE_TO:STOP\n\nThe haiku is accepted. "
             "Then call write_file to save the accepted haiku to: 04_Code_Artifacts/T5_LOOP/accepted_haiku.md",
             "", "0.5", "3", "none", "FAILED",
             "04_Code_Artifacts/T5_LOOP/accepted_haiku.md", "", ""],
        ],
    )


# ══════════════════════════════════════════════════════════════════════════════
# T6 — ORPHAN + DEAD-END DETECTION (PRE-FLIGHT MUST REJECT)
# ══════════════════════════════════════════════════════════════════════════════
def build_t6() -> None:
    """
    Topology (intentionally broken):
        VALID_NODE → STOP                     ← reachable, valid
        ORPHAN_NODE → NONEXISTENT_TARGET      ← unreachable (nothing routes here)
                                              ← dead-end (target doesn't exist)

    Tests: Pre-flight validate() catches BOTH problems BEFORE any API calls.
    Pass:  maccre.py launch exits with [VALIDATE] FAILED and zero API cost.
    Fail:  Any agent runs, any ledger is created, any cost is incurred.
    """
    wb = _copy_template("T6_VALIDATE")
    _populate(
        wb_path=wb,
        project_id="T6_VALIDATE",
        description="Pre-flight validation test — orphan node and dead-end must be caught before launch",
        start_node="VALID_NODE",
        payload="Topology test T6. Pre-flight must reject this topology before any agent runs.",
        agents=[
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "512", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
        ],
        topology=[
            # Valid node — reachable from START
            ["VALID_NODE", "TestAgent", "STOP",
             "Valid node in T6. Should never run — pre-flight must block launch.",
             "", "0.7", "1", "none", "FAILED",
             "04_Code_Artifacts/T6_VALIDATE/valid_output.md", "", ""],
            # Orphan: nothing routes here. Dead-end: NONEXISTENT_TARGET not defined.
            # Also has a bad wait_for target to exercise Phase 3 check 5.
            ["ORPHAN_NODE", "TestAgent", "NONEXISTENT_TARGET",
             "Orphan node — unreachable. Routes to a nonexistent target. Also waits for a ghost node.",
             "", "0.7", "1", "GHOST_NODE", "FAILED", "", "", ""],
        ],
    )


# ══════════════════════════════════════════════════════════════════════════════
# T7 — DIALOGUE MID-CHAIN
# ══════════════════════════════════════════════════════════════════════════════
def build_t7() -> None:
    """
    Topology:
        BRIEFER → DEBATER (DialogueRunner: ANALYST ↔ SKEPTIC, 2 rounds) → VERDICT → STOP

    Tests: DialogueRunner fires correctly when it is NOT the first node.
    The BRIEFER output is the payload fed into the dialogue context.
    Pass:
      - DEBATER ledger contains multi-turn dialogue (ANALYST/SKEPTIC turns visible)
      - VERDICT ledger exists and references the debate output
      - No SESSION_ID token in any artifact path
    """
    _ANALYST_PERSONA = (
        "You are ANALYST, a rigorous researcher who presents evidence-based arguments. "
        "Build on prior turns, cite the briefing context, and be concise (2-3 sentences per turn)."
    )
    _SKEPTIC_PERSONA = (
        "You are SKEPTIC, a critical thinker who challenges assumptions and identifies gaps. "
        "Push back on ANALYST's claims with sharp, specific counter-arguments (2-3 sentences)."
    )

    wb = _copy_template("T7_DIALOGUE")
    _populate(
        wb_path=wb,
        project_id="T7_DIALOGUE",
        description="DialogueRunner mid-chain — BRIEFER non-dialogue node feeds DEBATER DialogueRunner",
        start_node="BRIEFER",
        payload=(
            "Topology test T7. Topic for debate: "
            "'Is a fully autonomous multi-agent AI system safer with static topologies "
            "or adaptive conditional routing?' "
            "BRIEFER will summarise the question, DEBATER will debate it, VERDICT will decide."
        ),
        agents=[
            # BRIEFER uses generic test agent
            ["TestAgent", "Minimal test agent", "cloud",
             "gemini-2.5-flash", "0.7", "", "", "1024", "0",
             "FALSE", "FALSE", "", "text", "minimal", "write_file", _TEST_PERSONA],
            # ANALYST — dialogue side A
            ["AnalystAgent", "Evidence-based researcher", "cloud",
             "gemini-2.5-flash", "0.9", "", "", "1024", "0",
             "FALSE", "FALSE", "", "text", "minimal", "none", _ANALYST_PERSONA],
            # SKEPTIC — dialogue side B (dialogue_partner)
            ["SkepticAgent", "Critical challenger", "cloud",
             "gemini-2.5-flash", "0.9", "", "", "1024", "0",
             "FALSE", "FALSE", "", "text", "minimal", "none", _SKEPTIC_PERSONA],
        ],
        topology=[
            # BRIEFER: non-dialogue node, writes a structured brief from the payload
            ["BRIEFER", "TestAgent", "DEBATER",
             "You are BRIEFER in topology test T7. The [PREVIOUS NODE OUTPUT] contains a debate topic. "
             "Write a concise 3-paragraph research brief: "
             "(1) Frame the question. (2) Arguments for static topologies. "
             "(3) Arguments for adaptive/conditional routing. "
             "Call write_file to save to: 04_Code_Artifacts/T7_DIALOGUE/briefing.md",
             "", "0.7", "2", "none", "FAILED",
             "04_Code_Artifacts/T7_DIALOGUE/briefing.md", "", ""],
            # DEBATER: DialogueRunner mid-chain — ANALYST opens, SKEPTIC responds, 2 rounds
            ["DEBATER", "AnalystAgent", "VERDICT",
             "You are ANALYST opening a structured debate on the briefing you have received. "
             "State your position on adaptive/conditional routing being safer. "
             "Be direct and evidence-based. Your debate partner SKEPTIC will challenge you.",
             "", "0.9", "3", "none", "FAILED",
             "04_Code_Artifacts/T7_DIALOGUE/debate_log.md",
             "SkepticAgent", "2"],
            # VERDICT: fan-in waits for DEBATER artifact, writes final decision
            ["VERDICT", "TestAgent", "STOP",
             "You are VERDICT in topology test T7. "
             "The [GATHERED ARTIFACT: DEBATER] block above contains a full multi-turn debate. "
             "Read both sides and write a 2-paragraph verdict: which position is stronger and why. "
             "Call write_file to save to: 04_Code_Artifacts/T7_DIALOGUE/verdict.md",
             "", "0.7", "2", "DEBATER", "FAILED",
             "04_Code_Artifacts/T7_DIALOGUE/verdict.md", "", ""],
        ],
    )

# ══════════════════════════════════════════════════════════════════════════════
def main() -> None:
    print("\n[BUILD TOPOLOGY TESTS]")
    print(f"  Template source: {TEMPLATE_SRC}")
    if not TEMPLATE_SRC.exists():
        print(f"  [ERROR] Template not found: {TEMPLATE_SRC}")
        sys.exit(1)

    builders = [
        ("T1_SMOKE",    build_t1),
        ("T2_FAILURE",  build_t2),
        ("T3_DIAMOND",  build_t3),
        ("T4_RACE",     build_t4),
        ("T5_LOOP",     build_t5),
        ("T6_VALIDATE", build_t6),
        ("T7_DIALOGUE", build_t7),
    ]

    for name, fn in builders:
        print(f"\n  Building {name}...")
        fn()

    print("\n[DONE] All test workbooks written.")
    print("\nLaunch order (after EXO_TEST run completes):")
    print("  python maccre.py launch T1_SMOKE    --yes   # should complete clean")
    print("  python maccre.py launch T2_FAILURE  --yes   # RECOVERY must fire, NODE_C must NOT")
    print("  python maccre.py launch T3_DIAMOND  --yes   # MERGE must see both GATHERED ARTIFACTs")
    print("  python maccre.py launch T4_RACE     --yes   # MERGE must wait for SLOW_NODE")
    print("  python maccre.py launch T5_LOOP     --yes   # REFINER ledger must appear 2+ times")
    print("  python maccre.py launch T6_VALIDATE --yes   # must FAIL at pre-flight, zero cost")
    print("  python maccre.py launch T7_DIALOGUE --yes   # DialogueRunner must fire mid-chain")


if __name__ == "__main__":
    main()
