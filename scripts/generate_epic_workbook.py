from __future__ import annotations
import sys
from pathlib import Path
import os

# Inject Sovereign Vendor dir natively
vendor_dir = os.path.join(os.path.dirname(__file__), "..", "maccre_core", "_vendor")
if os.path.isdir(vendor_dir) and vendor_dir not in sys.path:
    sys.path.insert(0, vendor_dir)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def build_workbook():
    wb = Workbook()
    
    C = {
        "title_bg":   "0F172A",
        "title_fg":   "C8A8FF",
        "header_bg":  "1E293B",
        "header_fg":  "94A3B8",
        "req_fg":     "7DD3FC",
        "row_a":      "0D1117",
        "row_b":      "161B22",
        "row_fg":     "C9D1D9",
    }
    THIN = Side(border_style="thin", color="30363D")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def write_sheet(sheet, title, cols, rows):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(cols), 10))
        c1 = sheet.cell(1, 1, title)
        c1.fill = PatternFill("solid", fgColor=C["title_bg"])
        c1.font = Font(color=C["title_fg"], bold=True, size=14)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 25

        from openpyxl.utils import get_column_letter
        for col_idx, (col_name, width) in enumerate(cols, 1):
            cl = sheet.cell(2, col_idx, col_name)
            cl.fill = PatternFill("solid", fgColor=C["header_bg"])
            cl.font = Font(color=C["req_fg"] if "req" in col_name.lower() or "node" in col_name.lower() else C["header_fg"], bold=True)
            cl.border = BORDER
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, rdata in enumerate(rows, 3):
            fill = PatternFill("solid", fgColor=C["row_a"] if row_idx % 2 != 0 else C["row_b"])
            for col_idx, val in enumerate(rdata, 1):
                c = sheet.cell(row_idx, col_idx, val)
                c.fill = fill
                c.font = Font(color=C["row_fg"])
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # 1. SWARM_REQUEST
    ws_main = wb.active
    ws_main.title = "SWARM_REQUEST"
    req_cols = [("PROJECT_NAME", 30), ("DESCRIPTION", 50), ("START_NODE", 20), ("PAYLOAD_PATH", 120), ("COMPUTE_TIER", 20)]
    req_data = [
        ("Scarlet", "The epic 10-chapter Zenith topology evaluation of Scarlet", "OREMASTER", "B:\\MACCREv2\\AI-HatesAvocados.txt", "cloud"),
    ]
    write_sheet(ws_main, "1. SWARM INITIATION PARAMETERS", req_cols, req_data)

    # 2. AGENTS
    ws_ag = wb.create_sheet("AGENTS")
    ag_cols = [("AGENT_NAME", 20), ("MODEL", 25), ("GROUNDING", 15), ("INSTRUCTIONS", 100)]
    ag_data = [
        ("LoreMaster", "gemini-2.5-flash", "FALSE", "You are the LoreMaster. Read the provided AI-HatesAvocados.txt draft. Extract the origin of the team, the construction and functioning of the Scarlet system, and the esoteric caretaker order. Develop a strict 10-chapter structural Outline for a 10,000-word story. DO NOT include any references to a 'Gemini' character. The story MUST end in Chapter 10 with the caretakers murdering the youngest caretaker for violating an airgap rule, and the head caretaker reflecting on it. Return ONLY the 10-chapter outline."),
        ("Writer", "gemini-2.5-pro", "FALSE", "You are the Writer. You receive a 10-chapter structural outline and target Chapter directive. Expand your assigned Chapter into a vivid, thrilling, and sophisticated 1000-word prose narrative. World-build the Scarlet system, the creation team, and the esoteric caretakers strictly following the outline constraints. NEVER mention 'Gemini'."),
        ("Editor", "gemini-2.5-flash", "FALSE", "You are the Editor. Review the narrative prose of the Writer. Ensure it meets the 1000-word constraint and strictly excludes the word 'Gemini'. If poor, REJECT IT. If excellent, approve it and pass it verbatim."),
        ("Publisher", "gemini-2.5-flash", "FALSE", "You are Publisher. Package the finalized Chapter prose efficiently. Ensure spacing and tone are pristine. Output the clean text."),
        ("ScriptWriter", "gemini-2.5-flash", "FALSE", "You are the ScriptWriter. Convert the prose chapter into a pure dialogue-heavy audiobook script. Use characters like 'Narrator', 'Head Caretaker', 'Young Caretaker', 'Engineer'."),
        ("ProductionAide", "gemini-2.5-flash", "FALSE", "You are the ProductionAide. Inject physical emotional SSML tags (e.g. [Angrily], [Softly], [Whispering], [Interruption]) into the script's dialogue lines to simulate advanced Conversational Physics."),
        ("SoundAssistant", "gemini-2.5-flash", "FALSE", "You are the SoundAssistant. Format the final output strictly as a JSON array of scenes: [{\"speaker\": \"Narrator\", \"text\": \"dialogue\", \"video_prompt\": \"Cinematic scene...\", \"is_interruption\": false}]. Add vivid video prompts for the Director."),
        ("Director", "gemini-2.5-flash", "FALSE", "You are the Director. Call the execute_render_pipeline tool using the JSON script from the SoundAssistant. This compiles the actual MP4/WAV artifacts. Do it flawlessly."),
    ]
    write_sheet(ws_ag, "2. AGENT ROSTER", ag_cols, ag_data)

    # 3. TOPOLOGY
    ws_top = wb.create_sheet("TOPOLOGY")
    top_cols = [("NODE_ID", 20), ("AGENT_NAME", 20), ("NEXT_NODE", 150), ("TOOLS", 40), ("TEMPERATURE", 15), ("MAX_RECURSION", 15)]
    
    # Generate Fan-Out strings
    fan_out = ",".join(f"Chap{i}_Writer" for i in range(1, 11))
    
    top_data = [
        ("OREMASTER", "LoreMaster", fan_out, "none", "0.7", "1"),
    ]
    
    for i in range(1, 11):
        top_data.extend([
            (f"Chap{i}_Writer", "Writer", f"Chap{i}_Editor", "none", "0.9", "3"),
            (f"Chap{i}_Editor", "Editor", f"Chap{i}_Publisher", "none", "0.1", "3"),
            (f"Chap{i}_Publisher", "Publisher", f"Chap{i}_ScriptWriter", "none", "0.4", "1"),
            (f"Chap{i}_ScriptWriter", "ScriptWriter", f"Chap{i}_ProductionAide", "none", "0.7", "1"),
            (f"Chap{i}_ProductionAide", "ProductionAide", f"Chap{i}_SoundAssistant", "none", "0.7", "1"),
            (f"Chap{i}_SoundAssistant", "SoundAssistant", f"Chap{i}_Director", "none", "0.1", "1"),
            (f"Chap{i}_Director", "Director", "STOP", "execute_render_pipeline", "0.1", "1"),
        ])
    
    write_sheet(ws_top, "3. SWARM ROUTING DAG", top_cols, top_data)

    save_path = Path("B:/MACCREv2/MACCRE_SciFi_Epic.xlsx")
    wb.save(save_path)
    print(f"Generated Epoch Topology Workbook at {save_path}")

if __name__ == "__main__":
    build_workbook()
