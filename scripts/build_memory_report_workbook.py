"""
scripts/build_memory_report_workbook.py
"""
import sys
from pathlib import Path

# ── Project root anchor ───────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from maccre_core.utils.path_resolver import get_maccre_root

ROOT = get_maccre_root()
DATACENTER = ROOT / "__DATACENTER" / "MEMORY_TEST"
OUTFILE = DATACENTER / "MACCRE_MemoryReport.xlsx"

REPORTER_INSTRUCTION = """\
You are MemoryReporter. Your task is to investigate the database.
First, you MUST use iterative_scoped_search to find what the last swarm did.
Second, you MUST use iterative_scoped_search to find everything else in the databases that does not concern what the last swarm did.
Finally, compile a comprehensive report on your findings detailing what the last swarm did vs everything else in the database, and save it using write_file.

You MUST format all tool calls EXACTLY like this (do NOT wrap in XML tags):
LOCAL TOOL CALL REQUESTED: [{"function": {"name": "iterative_scoped_search", "arguments": {"query": "Aether-Flux"}}}]

And to write your report:
LOCAL TOOL CALL REQUESTED: [{"function": {"name": "write_file", "arguments": {"path": "04_Code_Artifacts/memory_report.txt", "data": "<your full report>"}}}]
"""

AGENTS = [
    {
        "AGENT_NAME": "MemoryReporter",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.7,
        "ROLE": "Investigates the database and compiles a report on recent additions vs existing knowledge.",
    }
]

TOPOLOGY = [
    {
        "Node_ID": "MEMORY_REPORTER",
        "Agent_Name": "MemoryReporter",
        "Next_Node": "STOP",
        "Instruction_Override": REPORTER_INSTRUCTION,
        "Temperature": "0.7",
        "Model_Override": "gemini-2.5-pro",
        "Wait_For": "",
        "Failure_Target": "STOP",
        "TOOLS": "iterative_scoped_search|write_file",
    }
]

SESSION_CONFIG = [
    ("PROJECT_ID", "MEMORY_TEST"),
]

def main() -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: openpyxl not installed.")
        sys.exit(1)

    wb = Workbook()

    ws = wb.active
    ws.title = "SWARM_REQUEST"
    ws.append(["MEMORY_TEST — Memory Report"])
    ws.append(["PROJECT_NAME", "START_NODE", "PAYLOAD_TEXT", "COMPUTE_TIER", "DESCRIPTION"])
    ws.append([
        "MEMORY_TEST",
        "MEMORY_REPORTER",
        "Compile a report on the memory DB.",
        "cloud",
        "A single-agent swarm to read what the last swarm did and compare it to everything else in the database."
    ])

    ws2 = wb.create_sheet("AGENTS")
    ws2.append(["Agents"])
    ws2.append(["AGENT_NAME", "MODEL", "TEMPERATURE", "TOOLS", "ROLE"])
    for agent in AGENTS:
        ws2.append([
            agent["AGENT_NAME"],
            agent["MODEL"],
            agent["TEMPERATURE"],
            "iterative_scoped_search|write_file",
            agent["ROLE"],
        ])

    ws3 = wb.create_sheet("TOPOLOGY")
    ws3.append(["Topology"])
    ws3.append([
        "NODE_ID", "AGENT_NAME", "NEXT_NODE",
        "INSTRUCTION_OVERRIDE", "TEMPERATURE",
        "AUTO_TOOL", "MODEL_OVERRIDE", "ARTIFACT_PATH",
        "Wait_For", "Failure_Target", "TOOLS",
    ])
    for t in TOPOLOGY:
        ws3.append([
            t["Node_ID"],
            t["Agent_Name"],
            t["Next_Node"],
            t["Instruction_Override"],
            t["Temperature"],
            "none",
            t["Model_Override"],
            "04_Code_Artifacts/memory_report.txt",
            t["Wait_For"],
            t["Failure_Target"],
            t["TOOLS"],
        ])

    ws4 = wb.create_sheet("SESSION_CONFIG")
    ws4.append(["Session Config"])
    ws4.append(["SETTING", "VALUE", "DESCRIPTION"])
    for key, val in SESSION_CONFIG:
        ws4.append([key, val, ""])

    DATACENTER.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTFILE))
    print(f"[OK] Workbook written: {OUTFILE}")

if __name__ == "__main__":
    main()
