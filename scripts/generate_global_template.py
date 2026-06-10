"""
scripts/generate_global_template.py
=====================================
Generates MACCRE_Global.xlsx at the MACCREv2 project root.

Sheets produced:
  PROJECT_DEFINITION  — project name, description, label, linked projects
  AGENTS              — agent roster with extended AI Studio params
  TOPOLOGY            — swarm pipeline node configuration
  SWARM_REQUEST       — payload config + start node
  SESSION_CONFIG      — lifecycle hooks (ingest_before_run, canonize_after_run)
  EXECUTION_PLAN      — per-section readiness checkboxes + FinOps estimates
  PIPELINE_CONFIG     — key/value runtime settings (inherits from legacy template)
  VAULT_KEYS          — key/value vault references

Usage:
  python scripts/generate_global_template.py             # writes MACCRE_Global.xlsx to root
  python scripts/generate_global_template.py --session   # writes MACCRE_Session.xlsx to root
  python scripts/generate_global_template.py --project FOO  # writes to __DATACENTER/FOO/
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

# Ensure maccre_core is importable from this script directory
sys.path.insert(0, str(Path(__file__).parent.parent))

# Vendored openpyxl produces guaranteed Excel-compatible OOXML output.
# The sovereign ooxml.py writer is used for swarm output artifacts (not UI workbooks).
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent.parent / "maccre_core" / "_vendor"))
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]
from maccre_core.workbook_data import (
    load_model_ids, load_project_names, load_topology_csv, load_node_ids,
    load_agent_roster_csv, load_recent_sessions,
    load_tool_names, load_all_agents_across_projects,
)
from scripts._panel_content import (
    PANEL_PROJECT, PANEL_AGENTS, PANEL_TOPOLOGY, PANEL_SWARM, PANEL_SESSION,
)

# ── Colour Palette ─────────────────────────────────────────────────────────────
_COL_HEADER_BG  = "1A1A2E"   # dark navy
_COL_HEADER_FG  = "E0E0E0"   # light grey
_COL_TITLE_BG   = "16213E"   # deeper navy
_COL_TITLE_FG   = "00FF99"   # mint accent
_COL_REQUIRED   = "E63946"   # required field indicator
_COL_OPTIONAL   = "4A9EBF"   # optional field indicator
_COL_EXEC_BG    = "0F3460"   # execution plan bg
_COL_ROW_ALT    = "F8F9FA"   # alternating data row tint
_COL_PANEL_HDR  = "0F3460"   # panel title background
_COL_PANEL_SEC  = "16213E"   # panel section header background
_COL_PANEL_RULE = "F0F0F8"   # panel rule / code row tint
_COL_PANEL_BODY = "FAFAFD"   # panel body row background
_COL_PANEL_GAP  = "6B6B8A"   # gap-column separator fill


def _header_font(bold: bool = True, color: str = _COL_HEADER_FG) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=10)


def _title_font() -> Font:
    return Font(name="Calibri", bold=True, color=_COL_TITLE_FG, size=12)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _add_dv(ws: object, sqref: str, formula1: str) -> None:
    """Add a strict-stop list DataValidation dropdown (free-text blocked)."""
    dv: DataValidation = DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showDropDown=False,   # False = SHOW the dropdown arrow (OOXML is inverted)
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid Selection",
        error="Choose a value from the dropdown list.",
    )
    ws.add_data_validation(dv)  # type: ignore[union-attr]
    dv.add(sqref)


def _add_dv_warn(
    ws: object,
    sqref: str,
    formula1: str,
    error_title: str = "New Project?",
    error_msg: str = "This name is not in the existing project list. Continue to auto-provision a new project silo.",
) -> None:
    """Add a warning-mode list DataValidation dropdown (free-text allowed).

    Use for fields like PROJECT_NAME where the dropdown shows existing values
    as suggestions, but the user must be able to type a brand-new name to
    auto-provision a fresh project silo.
    """
    dv: DataValidation = DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorStyle="warning",          # warning = allow anyway, don't hard-block
        errorTitle=error_title,
        error=error_msg,
    )
    ws.add_data_validation(dv)  # type: ignore[union-attr]
    dv.add(sqref)



def _write_sheet_header(ws: object, title: str, columns: list[str]) -> None:  # type: ignore[type-arg]
    """Write a decorative title row (row 1) and a header row (row 2)."""

    # Row 1 — title banner
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")  # type: ignore[union-attr]
    title_cell = ws.cell(row=1, column=1, value=f"◈  {title}  ◈")  # type: ignore[union-attr]
    title_cell.font = _title_font()  # type: ignore[union-attr]
    title_cell.fill = _fill(_COL_TITLE_BG)  # type: ignore[union-attr]
    title_cell.alignment = Alignment(horizontal="center", vertical="center")  # type: ignore[union-attr]
    ws.row_dimensions[1].height = 24  # type: ignore[union-attr]

    # Row 2 — column headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)  # type: ignore[union-attr]
        is_required = col_name.startswith("★")
        cell.font  = Font(name="Calibri", bold=True, size=10,   # type: ignore[union-attr]
                          color=_COL_REQUIRED if is_required else _COL_OPTIONAL)
        cell.fill  = _fill(_COL_HEADER_BG)   # type: ignore[union-attr]
        cell.border = _border()              # type: ignore[union-attr]
        cell.alignment = Alignment(horizontal="center", wrap_text=True)  # type: ignore[union-attr]


def _set_col_widths(ws: object, widths: list[int]) -> None:  # type: ignore[type-arg]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w  # type: ignore[union-attr]


def _write_side_panel(
    ws: object,
    gap_col: int,
    panel_col: int,
    content: list[tuple[str, str]],
    start_row: int = 2,
    gap_width: float = 1.5,
    panel_width: float = 58.0,
) -> None:
    """Write a styled reference panel to the right of the data entry area.

    The panel starts at *start_row* (default 2, alongside column headers) and
    grows downward one row per content entry.  A narrow gap column provides a
    visual separator.  Row heights are left at Excel defaults so the panel does
    not distort the data rows that share the same row index.

    Args:
        ws:          Target worksheet.
        gap_col:     Column index of the visual separator (e.g. 15 = col O).
        panel_col:   Column index of the main panel text (e.g. 16 = col P).
        content:     list[tuple[style, text]] from ``scripts._panel_content``.
        start_row:   First row to write into (default 2).
        gap_width:   Width of the gap column in Excel units.
        panel_width: Width of the panel column in Excel units.
    """
    import openpyxl.worksheet.worksheet as _wst  # noqa: PLC0415
    _ws: _wst.Worksheet = ws  # type: ignore[assignment]
    _ws.column_dimensions[get_column_letter(gap_col)].width = gap_width
    _ws.column_dimensions[get_column_letter(panel_col)].width = panel_width

    _style_map: dict[str, tuple[str, str, bool, int]] = {
        # style: (bg_hex, fg_hex, bold, font_size)
        "title":   (_COL_PANEL_HDR,  "E8D8FF", True,  11),
        "section": (_COL_PANEL_SEC,  "C8C8E8", True,  9),
        "rule":    (_COL_PANEL_RULE, "222244", False, 9),
        "body":    (_COL_PANEL_BODY, "444466", False, 9),
        "blank":   (_COL_PANEL_BODY, "FAFAFD", False, 9),
    }

    for row_offset, (style, text) in enumerate(content):
        row = start_row + row_offset
        bg, fg, bold, sz = _style_map.get(style, _style_map["body"])

        # Gap column — separator fill, no text
        gap_cell = _ws.cell(row=row, column=gap_col)
        gap_cell.fill = _fill(_COL_PANEL_GAP)

        # Panel text cell
        cell = _ws.cell(row=row, column=panel_col, value=text)
        cell.fill = _fill(bg)
        cell.font = Font(
            name="Courier New" if style == "rule" else "Calibri",
            bold=bold, size=sz, color=fg,
        )
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(style == "body"))


# ── Sheet Builders ─────────────────────────────────────────────────────────────


def _build_project_definition(wb: Workbook) -> None:
    ws = wb.create_sheet("PROJECT_DEFINITION")
    _write_sheet_header(ws, "PROJECT DEFINITION", ["★ SETTING", "★ VALUE", "Notes"])
    rows = [
        ("PROJECT_NAME",    "",  "Unique project silo identifier (alphanumeric + underscores)"),
        ("DESCRIPTION",     "",  "Brief description of the project's purpose"),
        ("SESSION_LABEL",   "",  "Optional human-readable tag for this run (e.g. 'chapter_2_outline')"),
        ("LINKED_PROJECTS", "",  "Comma-separated project names for Synaptic Bridge memory federation"),
    ]
    for i, (setting, val, note) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=setting).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=3, value=note).font = Font(name="Calibri", color="888888", size=9, italic=True)
    _set_col_widths(ws, [28, 40, 60])


def _build_agents(wb: Workbook) -> None:
    ws = wb.create_sheet("AGENTS")
    cols = [
        "★ AGENT_NAME", "★ MODEL", "ROLE", "★ PERSONA",
        "TEMPERATURE", "TOOLS", "TOP_P", "TOP_K",
        "MAX_OUTPUT_TOKENS", "THINKING_BUDGET", "SEARCH_GROUNDING",
        "RESPONSE_FORMAT", "SAFETY_LEVEL", "COMPUTE_TIER",
    ]
    _write_sheet_header(ws, "AGENT ROSTER", cols)
    # Example row
    example = [
        "DirectorFincher", "gemini-2.5-flash", "Script Director",
        "You are DirectorFincher, a precise narrative architect...",
        "1.0", "write_file", "", "", "", "", "FALSE",
        "markdown", "standard", "cloud",
    ]
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10, color="888888", italic=True)
        cell.fill = _fill(_COL_ROW_ALT)
    widths = [20, 20, 18, 50, 12, 20, 8, 8, 18, 16, 17, 16, 13, 14]
    _set_col_widths(ws, widths)


def _build_swarm_request(wb: Workbook, project_id: str = "", start_node: str = "") -> None:
    ws = wb.create_sheet("SWARM_REQUEST")
    cols = [
        "★ PROJECT_NAME", "DESCRIPTION", "COMPUTE_TIER",
        "PAYLOAD_TEXT", "PAYLOAD_PATH", "★ START_NODE",
        "OUTPUT_FOLDER", "NOTIFY_WEBHOOK",
    ]
    _write_sheet_header(ws, "SWARM REQUEST", cols)
    data_row = [project_id, "", "cloud", "", "", start_node, "", ""]
    for ci, val in enumerate(data_row, start=1):
        ws.cell(row=3, column=ci, value=val)
    _set_col_widths(ws, [22, 40, 14, 60, 50, 16, 30, 40])
    _write_side_panel(ws, gap_col=9, panel_col=10, content=PANEL_SWARM)


def _build_session_config(wb: Workbook) -> None:
    ws = wb.create_sheet("SESSION_CONFIG")
    _write_sheet_header(ws, "SESSION CONFIG", ["★ SETTING", "VALUE", "Notes"])
    rows = [
        ("PROJECT_NAME",        "",      "Must match the project silo created by 'maccre.py global'"),
        ("SESSION_LABEL",       "",      "Optional tag for this session (used in sessionID and filenames)"),
        ("INGEST_BEFORE_RUN",   "FALSE", "TRUE = run hash-aware ingest of 01_Raw_Source before swarm"),
        ("CANONIZE_AFTER_RUN",  "FALSE", "TRUE = auto-canonize session memory after swarm completes"),
        ("OUTPUT_FORMATS",      "md",    "Comma-separated: md, txt, json, mp3, mp4"),
    ]
    for i, (setting, val, note) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=setting).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=3, value=note).font = Font(name="Calibri", color="888888", size=9, italic=True)
    _set_col_widths(ws, [24, 20, 70])
    _write_side_panel(ws, gap_col=4, panel_col=5, content=PANEL_SESSION)


def _build_execution_plan(wb: Workbook, sections: list[str]) -> None:
    ws = wb.create_sheet("EXECUTION_PLAN")
    _write_sheet_header(ws, "EXECUTION PLAN & FINOPS", [
        "★ SECTION", "STATUS", "EXECUTE", "EST_COST_USD", "NOTES"
    ])
    for i, section in enumerate(sections, start=3):
        ws.cell(row=i, column=1, value=section).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=2, value="PENDING")
        ws.cell(row=i, column=3, value="TRUE")
        ws.cell(row=i, column=4, value=0.0)
        ws.cell(row=i, column=5, value="Populated by 'maccre.py global --dry-run' (Phase 8)")
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = _border()
    _set_col_widths(ws, [26, 12, 10, 14, 60])


def _build_pipeline_config(wb: Workbook) -> None:
    ws = wb.create_sheet("PIPELINE_CONFIG")
    _write_sheet_header(ws, "PIPELINE CONFIG", ["SETTING", "VALUE"])
    rows = [
        ("MAX_CYCLES",         "60"),
        ("TIMEOUT_SECONDS",    "600"),
        ("RETRY_ON_FAIL",      "TRUE"),
        ("MAX_RETRIES",        "3"),
        ("LOG_LEVEL",          "INFO"),
        ("TELEMETRY_ENABLED",  "TRUE"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _set_col_widths(ws, [28, 20])


def _build_vault_keys(wb: Workbook) -> None:
    ws = wb.create_sheet("VAULT_KEYS")
    _write_sheet_header(ws, "VAULT KEY REFERENCES", ["KEY_NAME", "VAULT_REF", "Notes"])
    rows = [
        ("GEMINI_API_KEY",      "MACCRE_Sovereign",      "Google Gemini API key — required for all cloud inference"),
        ("BRAVE_SEARCH_API_KEY","BRAVE_SEARCH_API_KEY",  "Brave Search API key — required for OSINT_BRAVE web search tool"),
        ("DRIVE_CREDS",         "MACCRE_Drive",          "Google Drive service-account credentials for state sync"),
    ]
    for i, (k, v, note) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=note).font = Font(name="Calibri", color="888888", size=9, italic=True)
    _set_col_widths(ws, [26, 28, 65])


def _build_instructions(wb: Workbook, wb_type: str) -> None:
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.column_dimensions["A"].width = 120
    ws.merge_cells("A1:A1")
    ws.cell(row=1, column=1, value=f"◈  MACCRE_{wb_type.upper()}.xlsx — Quick Reference  ◈").font = _title_font()
    ws.cell(row=1, column=1).fill = _fill(_COL_TITLE_BG)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    lines = [
        "",
        "WORKFLOW:",
        f"  1. Fill PROJECT_DEFINITION with your project name{' and SESSION_CONFIG' if wb_type == 'session' else ''}.",
        "  2. Define your AGENTS (one agent per row, ★ fields required).",
        "  3. Set up TOPOLOGY (one node per row, wire NEXT_NODE → STOP to terminate).",
        "  4. Set SWARM_REQUEST: paste payload text or point to PAYLOAD_PATH.",
        "  5. Use EXECUTION_PLAN to toggle which sections run and review cost.",
        "",
        "CLI COMMANDS:",
        f"  python maccre.py {'global' if wb_type == 'global' else 'launch <project_name>'}",
        "  python maccre.py ingest <project_name>   # embed 01_Raw_Source docs",
        "  python maccre.py status                  # check queue",
        "  python maccre.py canonize --project P --session S",
        "",
        "RULES:",
        "  • ★ fields are REQUIRED — the completeness engine will flag missing ones.",
        "  • TEMPERATURE: 0.1 = critic/extractor, 1.0 = creative generator.",
        "  • TOOLS: pipe-separated tool names (e.g. search_web|write_file|read_file).",
        "  •         Browse the hidden _TOOLS sheet for the full 33-tool reference.",
        "  • NEXT_NODE / FAILURE_TARGET: use STOP, DONE, or FAILED to terminate.",
        "  • PAYLOAD_TEXT takes priority over PAYLOAD_PATH if both are filled.",
    ]
    for row_idx, line in enumerate(lines, start=2):
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = Font(name="Calibri", size=10)


# ── Main Generator ─────────────────────────────────────────────────────────────


# ── Hidden reference sheet ─────────────────────────────────────────────────────

def _build_hidden_models(wb: Workbook, models: list[str]) -> int:
    """Write model IDs into a hidden _MODELS sheet; return last data row index."""
    ws = wb.create_sheet("_MODELS")
    for i, m in enumerate(models, start=1):
        ws.cell(row=i, column=1, value=m)
    ws.column_dimensions["A"].width = 40
    return len(models)


def _build_hidden_tools(wb: Workbook, tools: list[str]) -> int:
    """Write operator-facing tool names into a hidden _TOOLS reference sheet.

    The TOOLS column in AGENTS does not get a dropdown (pipe-separated multi-select
    is not supported by Excel DV without VBA).  This sheet is a visible reference
    operators can browse by unhiding it.  The column header comment on AGENTS
    directs operators here.
    """
    ws = wb.create_sheet("_TOOLS")
    ws.cell(row=1, column=1, value="TOOL_NAME").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=1, column=2, value="USAGE").font = Font(name="Calibri", bold=True, size=10)
    tool_notes: dict[str, str] = {
        "search_web": "Brave Search API — SEARCH_GROUNDING: Brave or Hybrid",
        "read_url_content": "Fetch and parse a URL's text content",
        "execute_hybrid_synthesis": "Parallel Google+Brave search, agentic dedup, source-labelled output",
        "query_local_memory": "Semantic vector search in project ChromaDB silo",
        "fts_search_memory": "BM25 full-text search — reaches content unavailable via vector",
        "ingest_document": "Hash-aware ingest of a file into the project vector store",
        "query_foreign_memory": "Semantic search in a linked project's vector silo",
        "import_foreign_vectors": "Pull vectors from a linked silo into this project",
        "prune_semantic_memory": "Remove stale or low-relevance embeddings from the silo",
        "read_file": "Read a file from the DATACENTER (project-aware path resolution)",
        "write_file": "Write content to 04_Code_Artifacts or 05_Rendered_Media",
        "file_exists": "Check whether a DATACENTER path exists",
        "trash_file": "Move a file to trash (requires elevation via request_elevation)",
        "query_thoughts": "Read agent L1 thoughts from the session ledger",
        "query_telemetry_matrix": "Query the telemetry DB for cost, node, and event data",
        "read_local_codebase": "Read Python source files for self-audit or code review",
        "generate_telemetry_report": "Produce a formatted FinOps + telemetry summary",
        "rotate_logs": "Archive and purge Op-logs and Bug-logs",
        "estimate_manifest_cost": "Estimate USD cost of a render manifest before execution",
        "execute_render_pipeline": "Fire TTS + image gen + FFmpeg stitching pipeline",
        "mint_agent": "Create a new agent definition in the roster",
        "build_topology": "Construct a topology CSV from a node spec",
        "link_projects": "Establish a Synaptic Bridge between two project silos",
        "ignite_swarm": "Inject a payload into the queue and start the swarm worker",
        "initialize_workspace": "Provision a new 6-tier DATACENTER project silo",
        "switch_workspace": "Change the active project context",
        "request_elevation": "Request PIN-gated elevation for privileged operations",
        "promote_topology_to_library": "Save current topology to the named library",
        "recall_topology": "Load a named topology from the library into the project",
        "design_swarm": "Agentic swarm design tool (generates topology + roster from a brief)",
        "run_swarm": "Execute the swarm worker loop for the active project",
        "create_persona_card": "Generate a structured agent persona card from a description",
        "export_project_nugget": "Export project knowledge nugget to Google Drive",
        "import_project_nuggets": "Import knowledge nuggets from Google Drive",
        "list_project_nuggets": "List available nuggets in Google Drive for this project",
        "ensure_project_workbook": "Guarantee a MACCRE_Session.xlsx exists in the project silo",
    }
    for i, tool in enumerate(tools, start=2):
        ws.cell(row=i, column=1, value=tool).font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=2, value=tool_notes.get(tool, "")).font = Font(
            name="Calibri", size=9, color="888888", italic=True
        )
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 65
    return len(tools)


def _build_hidden_agents(wb: Workbook, agent_entries: list[str]) -> int:
    """Write grouped agent entries into a hidden _AGENTS sheet.

    Each entry is ``[ProjectName] AgentName``.  The AGENT_NAME column in AGENTS
    uses warning-mode DV pointing here so existing agents are suggested in the
    dropdown while new names are still accepted.
    """
    ws = wb.create_sheet("_AGENTS")
    for i, entry in enumerate(agent_entries, start=1):
        ws.cell(row=i, column=1, value=entry)
    ws.column_dimensions["A"].width = 40
    return max(len(agent_entries), 1)


# ── Session Log ────────────────────────────────────────────────────────────────

def _build_session_log(wb: Workbook, sessions: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("SESSION_LOG")
    cols = ["SESSION_ID", "PROJECT", "LABEL", "STATUS", "EST_COST", "ACTUAL_COST", "CREATED"]
    _write_sheet_header(ws, "SESSION LOG", cols)
    for i, s in enumerate(sessions, start=3):
        ws.cell(row=i, column=1, value=str(s.get("session_id", "")))
        ws.cell(row=i, column=2, value=str(s.get("project_name", "")))
        ws.cell(row=i, column=3, value=str(s.get("label", "")))
        ws.cell(row=i, column=4, value=str(s.get("status", "")))
        ws.cell(row=i, column=5, value=float(s.get("est_cost_usd", 0) or 0))
        ws.cell(row=i, column=6, value=float(s.get("actual_cost_usd", 0) or 0))
        ws.cell(row=i, column=7, value=str(s.get("created_at", "")))
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = _border()
    _set_col_widths(ws, [34, 20, 20, 12, 10, 12, 26])


# ── Main Builder ───────────────────────────────────────────────────────────────

def build_global_workbook(output_path: Path, project_id: str = "") -> None:
    """Build and write MACCRE_Global.xlsx with live data and dropdowns."""
    # Load live data
    models      = load_model_ids()
    projects    = load_project_names()
    topo_rows   = load_topology_csv(project_id)
    agent_rows  = load_agent_roster_csv(project_id)
    node_ids    = load_node_ids(project_id)
    sessions    = load_recent_sessions(project_id)
    tools       = load_tool_names()
    all_agents  = load_all_agents_across_projects()
    model_count = len(models)
    model_range = f"'_MODELS'!$A$1:$A${model_count}"
    agent_count = max(len(all_agents), 1)
    agent_range = f"'_AGENTS'!$A$1:$A${agent_count}"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Hidden reference sheets first (must precede visible sheets that reference them)
    _build_hidden_models(wb, models)
    wb["_MODELS"].sheet_state = "hidden"  # type: ignore[index]
    _build_hidden_tools(wb, tools)
    wb["_TOOLS"].sheet_state = "hidden"   # type: ignore[index]
    _build_hidden_agents(wb, all_agents)
    wb["_AGENTS"].sheet_state = "hidden"  # type: ignore[index]

    # ── PROJECT_DEFINITION with project dropdown ────────────────────────────
    ws_proj = wb.create_sheet("PROJECT_DEFINITION")
    _write_sheet_header(ws_proj, "PROJECT DEFINITION", ["★ SETTING", "★ VALUE", "Notes"])
    proj_rows = [
        ("PROJECT_NAME",    project_id, "Select existing project or type a new name to auto-provision"),
        ("DESCRIPTION",     "",         "Brief description of the project purpose"),
        ("SESSION_LABEL",   "",         "Optional tag for this run (e.g. 'chapter_2_outline')"),
        ("SAVE_TO_LIBRARY", "TRUE",     "TRUE = save agents + topology to library on successful fire"),
        ("LINKED_PROJECTS", "",         "Comma-separated project names for memory federation"),
    ]
    proj_inline = ",".join(projects[:50])  # inline list max ~255 chars
    for i, (setting, val, note) in enumerate(proj_rows, start=3):
        ws_proj.cell(row=i, column=1, value=setting).font = Font(name="Calibri", bold=True, size=10)
        ws_proj.cell(row=i, column=2, value=val)
        ws_proj.cell(row=i, column=3, value=note).font = Font(name="Calibri", color="888888", size=9, italic=True)
    _add_dv_warn(ws_proj, "B3", f'"{proj_inline[:250]}"')  # PROJECT_NAME: suggest existing, allow new
    _add_dv(ws_proj, "B6", '"TRUE,FALSE"')
    _set_col_widths(ws_proj, [28, 40, 70])
    _write_side_panel(ws_proj, gap_col=4, panel_col=5, content=PANEL_PROJECT)

    # ── AGENTS with model + tool dropdowns ─────────────────────────────────
    ws_ag = wb.create_sheet("AGENTS")
    ag_cols = [
        "★ AGENT_NAME", "★ MODEL", "ROLE", "★ PERSONA",
        "TEMPERATURE", "TOOLS", "TOP_P", "TOP_K",
        "MAX_OUTPUT_TOKENS", "THINKING_BUDGET", "SEARCH_GROUNDING",
        "RESPONSE_FORMAT", "SAFETY_LEVEL", "COMPUTE_TIER",
    ]
    _write_sheet_header(ws_ag, "AGENT ROSTER", ag_cols)
    start_row = 3
    if agent_rows:
        for i, ag in enumerate(agent_rows, start=start_row):
            ws_ag.cell(row=i, column=1, value=str(ag.get("Agent_Name", ag.get("agent_name", ""))))
            ws_ag.cell(row=i, column=2, value=str(ag.get("Model_Override", ag.get("model", ""))))
            ws_ag.cell(row=i, column=3, value=str(ag.get("Role", ag.get("role", ""))))
            ws_ag.cell(row=i, column=5, value=str(ag.get("Temperature", ag.get("temperature", "1.0"))))
            ws_ag.cell(row=i, column=6, value=str(ag.get("Tools_Allowed", ag.get("tools_allowed", ""))))
            ws_ag.cell(row=i, column=11, value=str(ag.get("Search_Grounding", "Local Only")))
            ws_ag.cell(row=i, column=12, value=str(ag.get("Response_Format", "markdown")))
            ws_ag.cell(row=i, column=13, value=str(ag.get("Safety_Level", "standard")))
            ws_ag.cell(row=i, column=14, value=str(ag.get("Compute_Tier", "cloud")))
    else:
        eg = ["MyAgent", "gemini-2.5-flash", "Researcher",
              "You are MyAgent, a precise research agent...",
              "1.0", "search_web|write_file", "", "", "", "",
              "Local Only", "markdown", "standard", "cloud"]
        for ci, v in enumerate(eg, start=1):
            c = ws_ag.cell(row=3, column=ci, value=v)
            c.font = Font(name="Calibri", size=10, color="888888", italic=True)
            c.fill = _fill(_COL_ROW_ALT)
        start_row = 4
    end_row = max(start_row + len(agent_rows), 200)
    # ── AGENT_NAME: warning-mode — existing agents in dropdown, new names allowed.
    #    Excel shows a Yes/No dialog for unrecognised names:
    #      YES  → keep the typed name (new agent, will be minted on fire)
    #      NO   → revert to previous value
    _add_dv_warn(
        ws_ag, f"A3:A{end_row}", agent_range,
        error_title="New Agent Name",
        error_msg=(
            "This agent is not in the existing roster. "
            "Click YES to keep this name — the agent will be defined by the row settings "
            "and minted into the project on successful workbook execution. "
            "Click NO to pick from the dropdown instead."
        ),
    )
    # ── MODEL: strict stop — must be a registered model ID ───────────────────
    _add_dv(ws_ag, f"B3:B{end_row}", model_range)
    # ── ROLE: strict stop — standardised role vocabulary ─────────────────────
    _add_dv(ws_ag, f"C3:C{end_row}",
            '"Researcher,Writer,Synthesiser,Director,Critic,Extractor,Verifier,Renderer,Archivist,Router"')
    # ── SEARCH_GROUNDING: 5-mode list (replaces stale TRUE/FALSE) ────────────
    _add_dv(ws_ag, f"K3:K{end_row}",
            '"Local Only,Google,Brave,Hybrid,Hybrid-Synthesis"')
    # ── RESPONSE_FORMAT: deterministic ───────────────────────────────────────
    _add_dv(ws_ag, f"L3:L{end_row}", '"markdown,json,text"')
    # ── SAFETY_LEVEL: deterministic ──────────────────────────────────────────
    _add_dv(ws_ag, f"M3:M{end_row}", '"standard,strict,permissive"')
    # ── COMPUTE_TIER: deterministic ───────────────────────────────────────────
    _add_dv(ws_ag, f"N3:N{end_row}", '"cloud,edge,local"')
    _set_col_widths(ws_ag, [22, 22, 15, 50, 12, 35, 8, 8, 18, 16, 18, 14, 13, 12])
    _write_side_panel(ws_ag, gap_col=15, panel_col=16, content=PANEL_AGENTS)

    # ── TOPOLOGY with 10-column schema ──────────────────────────────────────
    ws_topo = wb.create_sheet("TOPOLOGY")
    topo_cols = [
        "★ NODE_ID", "★ AGENT_NAME", "★ NEXT_NODE",
        "MODEL_OVERRIDE", "TEMPERATURE", "INSTRUCTION_OVERRIDE",
        "WAIT_FOR", "FAILURE_TARGET", "MAX_RECURSION", "ARTIFACT_PATH",
        "DIALOGUE_PARTNER", "DIALOGUE_ROUNDS"
    ]
    _write_sheet_header(ws_topo, "SWARM TOPOLOGY", topo_cols)
    if topo_rows:
        for i, row in enumerate(topo_rows, start=3):
            ws_topo.cell(row=i, column=1,  value=str(row.get("Node_ID", "")))
            ws_topo.cell(row=i, column=2,  value=str(row.get("Agent_Name", "")))
            ws_topo.cell(row=i, column=3,  value=str(row.get("Next_Node", "STOP")))
            ws_topo.cell(row=i, column=4,  value=str(row.get("Model_Override", "")))
            ws_topo.cell(row=i, column=5,  value=str(row.get("Temperature", "1.0")))
            ws_topo.cell(row=i, column=6,  value=str(row.get("Instruction_Override", "")))
            ws_topo.cell(row=i, column=7,  value=str(row.get("Wait_For", "")))
            ws_topo.cell(row=i, column=8,  value=str(row.get("Failure_Target", "")))
            ws_topo.cell(row=i, column=9,  value=str(row.get("Max_Recursion", "3")))
            ws_topo.cell(row=i, column=10, value=str(row.get("Artifact_Path", "")))
            ws_topo.cell(row=i, column=11, value=str(row.get("Dialogue_Partner", "")))
            ws_topo.cell(row=i, column=12, value=str(row.get("Dialogue_Rounds", "0")))
    else:
        eg2 = ["NODE_01", "MyAgent", "STOP", "", "1.0",
               "Write a brief from the payload.", "", "STOP", "3", "", "", "0"]
        for ci, v in enumerate(eg2, start=1):
            c = ws_topo.cell(row=3, column=ci, value=v)
            c.font = Font(name="Calibri", size=10, color="888888", italic=True)
            c.fill = _fill(_COL_ROW_ALT)
    topo_end = max(3 + len(topo_rows), 50)
    agent_names_inline = ",".join(
        str(r.get("Agent_Name", r.get("agent_name", ""))) for r in agent_rows
    ) or "MyAgent"
    next_nodes_inline = ",".join(node_ids + ["STOP", "DONE", "FAILED"]) or "STOP,DONE,FAILED"
    _add_dv_warn(
        ws_topo, f"B3:B{topo_end}", f'"{agent_names_inline[:250]}"',
        error_title="Agent Not in Roster",
        error_msg=(
            "This agent is not in the current project roster. "
            "Click YES to use this name - the agent must be defined "
            "in the AGENTS sheet and will be minted on first execution."
        ),
    )
    _add_dv(ws_topo, f"C3:C{topo_end}", f'"{next_nodes_inline[:250]}"')
    _add_dv(ws_topo, f"D3:D{topo_end}", model_range)
    _add_dv(ws_topo, f"H3:H{topo_end}", f'"{next_nodes_inline[:250]}"')  # FAILURE_TARGET
    _set_col_widths(ws_topo, [18, 22, 22, 22, 12, 45, 28, 22, 14, 35])
    _write_side_panel(ws_topo, gap_col=13, panel_col=14, content=PANEL_TOPOLOGY)

    # ── Remaining sheets (unchanged structure) ──────────────────────────────
    # Derive the default start node from the first topology row
    _default_start = str(topo_rows[0].get("Node_ID", "")) if topo_rows else ""
    _build_swarm_request(wb, project_id=project_id, start_node=_default_start)
    _build_session_config(wb)
    _build_execution_plan(wb, ["PROJECT_DEFINITION", "AGENTS", "TOPOLOGY", "SWARM_REQUEST"])
    _build_pipeline_config(wb)
    _build_vault_keys(wb)
    _build_session_log(wb, sessions)
    _build_instructions(wb, "global")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    _stamp_execution_plan_status(output_path)
    print(f"[OK] MACCRE_Global.xlsx -> {output_path}")
    print(f"     Project: {project_id or 'GLOBAL'} | Models: {model_count} | Sheets: {len(wb.sheetnames)}")


# ── Live EXECUTION_PLAN Stamper ────────────────────────────────────────────────


def _stamp_execution_plan_status(output_path: Path) -> None:
    """Re-open a just-saved workbook and stamp live STATUS, EST_COST, and colours
    into the EXECUTION_PLAN sheet rows.

    Runs check_workbook_completeness() on the file, then overwrites columns
    2 (STATUS) / 4 (EST_COST_USD) / 5 (NOTES) with real values and Excel-standard
    conditional-formatting colours so the operator can see readiness at a glance.
    """
    # Excel standard conditional-formatting palette: (fill_rgb, text_rgb)
    _STATUS_STYLE: dict[str, tuple[str, str]] = {
        "READY":      ("C6EFCE", "276221"),  # green
        "PARTIAL":    ("FFEB9C", "9C5700"),  # amber
        "INCOMPLETE": ("FFC7CE", "9C0006"),  # red
        "MISSING":    ("FFC7CE", "9C0006"),  # red
    }
    _GREY = ("D9D9D9", "595959")  # unknown / pending

    try:
        from maccre_core.tools.workbook_engine import check_workbook_completeness  # noqa: PLC0415
        plan = check_workbook_completeness(output_path, wb_type="global")
    except Exception as exc:  # noqa: BLE001
        print(f"[WARN] Completeness engine unavailable — EXECUTION_PLAN will show PENDING: {exc}")
        return

    section_map: dict[str, Any] = {s.name: s for s in plan.sections}

    wb2 = load_workbook(str(output_path))  # type: ignore[no-untyped-call]
    if "EXECUTION_PLAN" not in wb2.sheetnames:
        wb2.close()
        return

    ws_ep = wb2["EXECUTION_PLAN"]
    # Layout: col1=SECTION  col2=STATUS  col3=EXECUTE  col4=EST_COST_USD  col5=NOTES
    for row_idx in range(3, (ws_ep.max_row or 3) + 1):
        section_cell = ws_ep.cell(row=row_idx, column=1)
        section_name = str(section_cell.value or "").strip()
        if not section_name:
            continue
        result = section_map.get(section_name)
        if result is None:
            continue

        bg, fg = _STATUS_STYLE.get(result.status, _GREY)

        status_cell = ws_ep.cell(row=row_idx, column=2)
        status_cell.value = result.status
        status_cell.fill = PatternFill("solid", fgColor=bg)  # type: ignore[call-arg]
        status_cell.font = Font(name="Calibri", bold=True, size=10, color=fg)

        cost_cell = ws_ep.cell(row=row_idx, column=4)
        cost_cell.value = round(result.est_cost_usd, 6)

        notes_cell = ws_ep.cell(row=row_idx, column=5)
        notes_cell.value = "; ".join(result.notes) if result.notes else "OK"
        notes_cell.font = Font(name="Calibri", size=9, color="888888", italic=True)

    wb2.save(str(output_path))
    wb2.close()


# ── CLI ────────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate MACCRE_Global.xlsx with live data and dropdowns.",
    )
    parser.add_argument(
        "--project", default="",
        help="Pre-populate from this project silo (default: GLOBAL)"
    )
    parser.add_argument(
        "--out", default="",
        help="Override output path (default: MACCRE_Global.xlsx at root)"
    )
    args = parser.parse_args()
    root = Path(__file__).parent.parent
    out = Path(args.out) if args.out else root / "MACCRE_Global.xlsx"
    build_global_workbook(out, project_id=args.project)


if __name__ == "__main__":
    main()
