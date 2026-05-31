"""
scripts/populate_phase0_test.py
================================
Populates MACCRE_Global.xlsx with the Phase 0 live-fire test:
  Three narrator agents in a linear chain producing a 3-chapter story.

Run once. Safe to re-run — overwrites only the data rows we own.

Usage:
    python scripts/populate_phase0_test.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
WB_PATH = ROOT / "MACCRE_Global.xlsx"

NARRATOR_PERSONA = (
    "You are a skilled narrative author participating in a collaborative, "
    "linear story-writing pipeline. You receive the original story premise "
    "and all previously written chapters. You write ONLY your assigned chapter "
    "— clean, compelling prose with no meta-commentary, no labels, no preamble. "
    "Begin writing immediately. Match the tone and voice of preceding chapters. "
    "End your chapter at a natural moment that the next author can continue from."
)

CHAPTER_1_OVERRIDE = (
    "You are writing CHAPTER ONE. This is the OPENING THIRD of the story arc. "
    "Introduce the protagonist (the boy), the setting (Easter morning), and the "
    "inciting incident (discovering the solid gold egg in his Easter basket). "
    "Establish the emotional tone and end the chapter with the boy just beginning "
    "to comprehend what he has found. Do not resolve anything."
)

CHAPTER_2_OVERRIDE = (
    "You are writing CHAPTER TWO. This is the MIDDLE THIRD of the story arc. "
    "The boy now has the golden egg and must decide what to do. Complications "
    "arise — reactions from family, questions of value and meaning, perhaps a "
    "temptation or external pressure. Build tension. End on the cusp of the "
    "story's climax but do not resolve it."
)

CHAPTER_3_OVERRIDE = (
    "You are writing CHAPTER THREE. This is the FINAL THIRD of the story arc. "
    "Bring the story to a satisfying, complete resolution. Address what the golden "
    "egg truly means to the boy — not just its monetary value. The ending should "
    "feel earned and emotionally resonant. This is the final chapter; close all "
    "open threads with intention."
)

PAYLOAD = (
    "A young boy wakes up on Easter morning to find a single egg in his basket. "
    "It is not chocolate. It is not plastic. It is solid gold — warm to the touch, "
    "heavier than anything he has held before, and engraved on one side with his "
    "initials. No one in his family put it there. No one knows where it came from."
)


def _find_data_start(ws: object, header_row: int = 2) -> int:  # type: ignore[type-arg]
    """Return the first data row index (header_row + 1), 1-based."""
    return header_row + 1


def _header_col(ws: object, target: str, header_row: int = 2) -> int | None:  # type: ignore[type-arg]
    """Return the 1-based column index for a header cell matching target (case-insensitive)."""
    for cell in ws[header_row]:  # type: ignore[index]
        if cell.value and target.lower() in str(cell.value).lower():
            return int(cell.column)
    return None


def _write_row(ws: object, row: int, data: dict[str, str], header_row: int = 2) -> None:  # type: ignore[type-arg]
    """Write key→value pairs into the given row using the header map."""
    for key, value in data.items():
        col = _header_col(ws, key, header_row)
        if col is not None:
            ws.cell(row=row, column=col, value=value)  # type: ignore[union-attr]


def _clear_data_rows(ws: object, start_row: int, end_row: int) -> None:  # type: ignore[type-arg]
    """Blank out data rows so stale content does not bleed through."""
    for r in range(start_row, end_row + 1):
        for cell in ws[r]:  # type: ignore[index]
            cell.value = None  # type: ignore[union-attr]


def populate(wb_path: Path) -> None:
    print(f"[POPULATE] Loading workbook: {wb_path}")
    wb = load_workbook(filename=str(wb_path))

    # ── PROJECT_DEFINITION ────────────────────────────────────────────────────
    ws_pd = wb["PROJECT_DEFINITION"]
    _clear_data_rows(ws_pd, 3, 10)
    kv_map: dict[str, str] = {
        "PROJECT_NAME":   "GOLDEN_EGG",
        "DESCRIPTION":    "Phase 0 live-fire test: three-chapter narrative pipeline.",
        "VERSION":        "0.1",
        "AUTHOR":         "MACCREv2 Phase 0",
        "COMPUTE_TIER":   "cloud",
        "OUTPUT_FORMAT":  "markdown",
    }
    for idx, (key, val) in enumerate(kv_map.items()):
        ws_pd.cell(row=3 + idx, column=1, value=key)
        ws_pd.cell(row=3 + idx, column=2, value=val)
    print("[POPULATE] PROJECT_DEFINITION -> OK")

    # ── AGENTS ────────────────────────────────────────────────────────────────
    ws_ag = wb["AGENTS"]
    _clear_data_rows(ws_ag, 3, 10)
    # All three narrators are identical by persona — instruction override is per topology node
    for i, agent_name in enumerate(["Narrator_1", "Narrator_2", "Narrator_3"], start=0):
        _write_row(ws_ag, 3 + i, {
            "AGENT_NAME":  agent_name,
            "MODEL":       "gemini-2.5-flash",
            "TEMPERATURE": "1.0",
            "TOOLS":       "write_file",
            "ROLE":        "Narrative author — writes one chapter of the story",
            "PERSONA":     NARRATOR_PERSONA,
        })
    print("[POPULATE] AGENTS -> OK (3 narrators)")

    # ── TOPOLOGY ──────────────────────────────────────────────────────────────
    ws_tp = wb["TOPOLOGY"]
    _clear_data_rows(ws_tp, 3, 10)

    nodes: list[dict[str, str]] = [
        {
            "NODE_ID":              "CHAPTER_1",
            "AGENT_NAME":          "Narrator_1",
            "AUTO_TOOL":           "write_file",
            "NEXT_NODE":           "CHAPTER_2",
            "INSTRUCTION_OVERRIDE": CHAPTER_1_OVERRIDE,
            "MODEL_OVERRIDE":      "",
            "TEMPERATURE":         "1.0",
        },
        {
            "NODE_ID":              "CHAPTER_2",
            "AGENT_NAME":          "Narrator_2",
            "AUTO_TOOL":           "write_file",
            "NEXT_NODE":           "CHAPTER_3",
            "INSTRUCTION_OVERRIDE": CHAPTER_2_OVERRIDE,
            "MODEL_OVERRIDE":      "",
            "TEMPERATURE":         "1.0",
        },
        {
            "NODE_ID":              "CHAPTER_3",
            "AGENT_NAME":          "Narrator_3",
            "AUTO_TOOL":           "write_file",
            "NEXT_NODE":           "DONE",
            "INSTRUCTION_OVERRIDE": CHAPTER_3_OVERRIDE,
            "MODEL_OVERRIDE":      "",
            "TEMPERATURE":         "0.9",
        },
    ]
    for i, node in enumerate(nodes):
        _write_row(ws_tp, 3 + i, node)
    print("[POPULATE] TOPOLOGY -> OK (CHAPTER_1 -> CHAPTER_2 -> CHAPTER_3 -> DONE)")

    # ── SWARM_REQUEST ─────────────────────────────────────────────────────────
    ws_sr = wb["SWARM_REQUEST"]
    _clear_data_rows(ws_sr, 3, 5)
    _write_row(ws_sr, 3, {
        "PROJECT_NAME":  "GOLDEN_EGG",
        "DESCRIPTION":   "Three-chapter narrative: a boy finds a solid gold egg in his Easter basket.",
        "COMPUTE_TIER":  "cloud",
        "PAYLOAD_TEXT":  PAYLOAD,
        "PAYLOAD_PATH":  "",
        "START_NODE":    "CHAPTER_1",
        "OUTPUT_FOLDER": "04_Code_Artifacts",
        "NOTIFY_WEBHOOK": "",
    })
    print("[POPULATE] SWARM_REQUEST -> OK")

    # ── EXECUTION_PLAN — mark SWARM_REQUEST as execute=YES ───────────────────
    if "EXECUTION_PLAN" in wb.sheetnames:
        ws_ep = wb["EXECUTION_PLAN"]
        for row in ws_ep.iter_rows(min_row=3):
            first = row[0].value
            if first and "SWARM" in str(first).upper():
                # Column C is typically the EXECUTE checkbox
                if len(row) >= 3:
                    row[2].value = "YES"
        print("[EXECUTION_PLAN] SWARM_REQUEST marked EXECUTE=YES")

    wb.save(str(wb_path))
    print(f"\n[POPULATE] Workbook saved: {wb_path}")
    print("[POPULATE] Ready to run:  python maccre.py global")


if __name__ == "__main__":
    populate(WB_PATH)
