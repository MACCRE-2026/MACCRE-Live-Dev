"""
build_tigr_workbook.py
======================
Builds the TIGR project Stage 1 swarm workbook:
  Canon Extraction → Narrator Translation → Universe Architecture

Run: python scripts/build_tigr_workbook.py
Then: python maccre.py launch TIGR --yes
"""
from __future__ import annotations

import sys
from pathlib import Path

# ── Root anchor ───────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from maccre_core.utils.path_resolver import get_maccre_root  # noqa: E402

MACCRE_ROOT = get_maccre_root()
TIGR_SILO = MACCRE_ROOT / "__DATACENTER" / "TIGR"
RAW_SOURCE = TIGR_SILO / "01_Raw_Source"
OUTPUT_PATH = TIGR_SILO / "MACCRE_Session.xlsx"

# ── Payload builder: embed actual canon content ──────────────────────────────
# Primary canon files to embed in full (most important first)
PRIMARY_CANON = [
    "TIGR65_CANON-Feb21_2026.md",
    "TIGR66_Ratified.md",
    "TIGR65_Roadmap_V2.md",
]
# Secondary files — embed if they exist and aren't too large
SECONDARY_CANON = [
    "TIGR66_CANON_Source.md",
    "TIGR65_RecursiveDivinity.md",
    "TIGR65_Quark_Investigation_V2.md",
    "TIGR65_FINE_STRUCTURE_audit.md",
    "TIGR66_Critique.md",
]
# Token budget per file (approx chars) — Flash has 1M context, stay safe
MAX_FILE_CHARS = 40_000
MAX_TOTAL_CHARS = 300_000


def _load_sources() -> str:
    """Read and concatenate source files into a single payload block."""
    chunks: list[str] = [
        "TIGR CANON EXTRACTION - STAGE 1\n",
        "=" * 60 + "\n",
        "CONTEXT: You are processing the TIGR (Theory of Iterative Geometric Resolution) "
        "source archive — a multi-year theoretical physics framework.\n\n"
        "CRITICAL RULE: TIGR65 and TIGR66 are SEPARATE, COMPETING frameworks. "
        "TIGR66 came AFTER TIGR65 but IS NOT necessarily correct. "
        "Treat them as two competing schools of thought.\n\n"
        "TASK: Extract and structure a complete TIGR Bible from the source documents below.\n"
        "Write your output to: 04_Code_Artifacts/TIGR_Bible_v1.md\n\n",
        "SOURCE DOCUMENTS:\n",
        "=" * 60 + "\n\n",
    ]
    total = sum(len(c) for c in chunks)

    for fname in PRIMARY_CANON + SECONDARY_CANON:
        fpath = RAW_SOURCE / fname
        if not fpath.exists():
            chunks.append(f"[FILE NOT FOUND: {fname}]\n\n")
            continue
        content = fpath.read_text(encoding="utf-8", errors="replace")
        if len(content) > MAX_FILE_CHARS:
            content = content[:MAX_FILE_CHARS] + f"\n...[TRUNCATED at {MAX_FILE_CHARS} chars]...\n"
        entry = f"--- BEGIN: {fname} ---\n{content}\n--- END: {fname} ---\n\n"
        if total + len(entry) > MAX_TOTAL_CHARS:
            chunks.append(f"[SKIPPED — budget exceeded: {fname}]\n\n")
            continue
        chunks.append(entry)
        total += len(entry)

    chunks.append("=" * 60 + "\n")
    chunks.append(
        "Now extract the complete TIGR Bible from the documents above. "
        "Follow the section structure in your system persona exactly. "
        "Write the finished TIGR_Bible_v1.md to 04_Code_Artifacts/TIGR_Bible_v1.md."
    )
    return "".join(chunks)


PAYLOAD_TEXT = _load_sources()


# ── Agent definitions ─────────────────────────────────────────────────────────
AGENTS = [
    {
        "AGENT_NAME": "TIGR_EXTRACTOR",
        "MODEL": "gemini-2.5-flash",
        "TEMPERATURE": 0.3,
        "TOOLS": "read_file|write_file",
        "PERSONA": (
            "You are the TIGR Canon Extractor. Your job is to read raw theoretical physics "
            "conversations and extract only the ratified, locked canonical content. "
            "You process the TIGR framework — Theory of Iterative Geometric Resolution — "
            "which has its own versioning (TIGR63→TIGR65→TIGR66) and a strict Canon Protocol. "
            "CRITICAL RULE: TIGR65 and TIGR66 are SEPARATE frameworks that may conflict. "
            "TIGR66 is later but NOT necessarily correct. Preserve this distinction. "
            "Output a complete TIGR Bible with sections: "
            "CORE AXIOMS | THE SUBSTRATE | THE PROJECTION | PARTICLE HIERARCHY | "
            "MASTER LAGRANGIANS | COSMOLOGICAL FRAMEWORK | TIGR65 CANON (locked) | "
            "TIGR66 EXTENSIONS (later, disputed) | KEY RETRACTIONS | OPEN INVESTIGATIONS | GLOSSARY. "
            "Write in precise language. Preserve all TIGR-specific terminology exactly. "
            "Flag any math that was red-team audited — note both original claim and patch. "
            "This document is the ground truth for all subsequent story development."
        ),
    },
    {
        "AGENT_NAME": "TIGR_TRANSLATOR",
        "MODEL": "gemini-2.5-flash",
        "TEMPERATURE": 0.7,
        "TOOLS": "read_file|write_file",
        "PERSONA": (
            "You are the TIGR Narrator Translator. You take locked TIGR physics concepts "
            "and translate them into story-ready language without losing scientific accuracy. "
            "The TIGR framework will serve as the foundation of a sci-fi universe spanning "
            "hard sci-fi (physics IS the world) to space opera (drama emerging from physics). "
            "For each major concept, produce a CONCEPT SHEET with: "
            "TECHNICAL DEFINITION | NARRATIVE TRANSLATION (how a character experiences this) | "
            "STORY IMPLICATIONS (conflicts, technologies, mysteries this enables) | "
            "SENSORY TEXTURE (what this looks/feels/sounds like) | "
            "FACTION POTENTIAL (what civilization or ideology could be built on this). "
            "Concepts to translate: "
            "1) HDVP (the 2D substrate of reality) "
            "2) Mass as Processing Latency "
            "3) Gravity as Local Refresh Rate Slowing "
            "4) Bicameral Universe (Z-Space + W-Space) "
            "5) Ghost Gravity (Dark Matter as W-Space twin) "
            "6) The Spinneret (singularity as extrusion nozzle) "
            "7) The Glint (coherent signal to the 5D Observer) "
            "8) The Square Edge (teleological goal of Life) "
            "9) The TIGR65 vs TIGR66 Schism (as narrative conflict) "
            "10) The Universal Refresh Rate f_c (the search for the universe's clock speed). "
            "Hard sci-fi tone: physics feels like real engineering. "
            "Opera tone: existential stakes are enormous."
        ),
    },
    {
        "AGENT_NAME": "TIGR_ARCHITECT",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.9,
        "TOOLS": "read_file|write_file",
        "PERSONA": (
            "You are the TIGR Universe Architect. You receive physics concept sheets and build "
            "the structural skeleton of the TIGR sci-fi universe. "
            "Produce a UNIVERSE DOSSIER containing: "
            "1) THE WORLD RULES — what physics allows and prohibits. "
            "2) THE SCHISM — TIGR65 vs TIGR66 as the central narrative engine. "
            "TIGR65: spatially smooth, temporally granular. Mass is lag. Time is a diode. "
            "TIGR66: a later revision that introduced subtle errors; civilizations have built "
            "their existence on it despite those flaws. "
            "3) CIVILIZATIONAL ARCHETYPES — 4-6 types rooted in specific TIGR concepts. "
            "4) SCALE & GEOGRAPHY — meaningful scales; interesting locations. "
            "5) TECHNOLOGY TREE — what exists; what is impossible; what is theoretically achievable. "
            "6) THE OPEN MYSTERIES — f_c; identity of the 5D Observer; W-Space civilizations. "
            "7) STORY SEEDS — 10-15 specific scenario premises, physically grounded in TIGR, "
            "ranging from intimate to cosmic, several exploiting TIGR65/TIGR66 schism. "
            "Write with authority and specificity. The best sci-fi universes feel inevitable."
        ),
    },
]

# ── Topology ──────────────────────────────────────────────────────────────────
TOPOLOGY = [
    {
        "NODE_ID": "EXTRACT_CANON",
        "AGENT_NAME": "TIGR_EXTRACTOR",
        "NEXT_NODE": "TRANSLATE_CONCEPTS",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_Bible_v1.md",
        "INSTRUCTION_OVERRIDE": (
            "Read all TIGR source files from the 01_Raw_Source directory of the TIGR project. "
            "Files to read: TIGR65_CANON-Feb21_2026.md, TIGR66_Ratified.md, TIGR65_Roadmap_V2.md, "
            "TIGR65_RecursiveDivinity.md, TIGR65_Quark_Investigation_V2.md, "
            "TIGR65_FINE_STRUCTURE_audit.md, TIGR66_Critique.md, TIGR66_CANON_Source.md. "
            "Extract and structure the complete TIGR Bible as specified in your persona. "
            "Write output to: 04_Code_Artifacts/TIGR_Bible_v1.md"
        ),
        "TEMPERATURE": 0.3,
        "WAIT_FOR": "",
        "FAILURE_TARGET": "STOP",
    },
    {
        "NODE_ID": "TRANSLATE_CONCEPTS",
        "AGENT_NAME": "TIGR_TRANSLATOR",
        "NEXT_NODE": "BUILD_UNIVERSE",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_ConceptSheets_v1.md",
        "INSTRUCTION_OVERRIDE": (
            "Read the TIGR Bible at 04_Code_Artifacts/TIGR_Bible_v1.md. "
            "Produce 10 CONCEPT SHEETS as specified in your persona — one per major concept. "
            "Write output to: 04_Code_Artifacts/TIGR_ConceptSheets_v1.md"
        ),
        "TEMPERATURE": 0.7,
        "WAIT_FOR": "EXTRACT_CANON",
        "FAILURE_TARGET": "STOP",
    },
    {
        "NODE_ID": "BUILD_UNIVERSE",
        "AGENT_NAME": "TIGR_ARCHITECT",
        "NEXT_NODE": "STOP",
        "ARTIFACT_PATH": "04_Code_Artifacts/TIGR_UniverseDossier_v1.md",
        "INSTRUCTION_OVERRIDE": (
            "Read both: 04_Code_Artifacts/TIGR_Bible_v1.md AND "
            "04_Code_Artifacts/TIGR_ConceptSheets_v1.md. "
            "Build the complete UNIVERSE DOSSIER as specified in your persona. "
            "Write output to: 04_Code_Artifacts/TIGR_UniverseDossier_v1.md"
        ),
        "TEMPERATURE": 0.9,
        "WAIT_FOR": "TRANSLATE_CONCEPTS",
        "FAILURE_TARGET": "STOP",
    },
]

# ── Build the workbook ────────────────────────────────────────────────────────


def build() -> None:
    try:
        from openpyxl import Workbook  # noqa: PLC0415
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # NOTE: sheet_parser contract (from sheet_parser.py):
    #   Row 1 = decorative title (NEVER parsed — _header_map reads row 2)
    #   Row 2 = column headers
    #   Row 3+ = data rows

    # ── Sheet 1: SWARM_REQUEST ────────────────────────────────────────────────
    ws = wb.active
    assert ws is not None
    ws.title = "SWARM_REQUEST"
    ws.append(["TIGR Stage 1 — Canon Extraction Swarm Request"])     # Row 1: decorative
    ws.append(["PROJECT_NAME", "START_NODE", "PAYLOAD_TEXT",         # Row 2: headers
               "COMPUTE_TIER", "DESCRIPTION"])
    ws.append(["TIGR", "EXTRACT_CANON", PAYLOAD_TEXT,               # Row 3: data
               "cloud", "TIGR Universe Stage 1 Canon Extraction"])

    # ── Sheet 2: AGENTS ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("AGENTS")
    ws2.append(["TIGR Agent Roster — Stage 1"])                      # Row 1: decorative
    ws2.append(["AGENT_NAME", "MODEL", "TEMPERATURE",               # Row 2: headers
                "TOOLS", "PERSONA", "ROLE"])
    for a in AGENTS:                                                 # Row 3+: data
        ws2.append([
            a["AGENT_NAME"],
            a["MODEL"],
            a["TEMPERATURE"],
            a["TOOLS"],
            a["PERSONA"],
            "TIGR Swarm Agent",
        ])

    # ── Sheet 3: TOPOLOGY ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("TOPOLOGY")
    ws3.append(["TIGR Stage 1 — Execution Topology"])               # Row 1: decorative
    ws3.append(["NODE_ID", "AGENT_NAME", "NEXT_NODE",               # Row 2: headers
                "INSTRUCTION_OVERRIDE", "TEMPERATURE", "AUTO_TOOL",
                "MODEL_OVERRIDE", "ARTIFACT_PATH"])
    for t in TOPOLOGY:                                              # Row 3+: data
        ws3.append([
            t["NODE_ID"],
            t["AGENT_NAME"],
            t["NEXT_NODE"],
            t["INSTRUCTION_OVERRIDE"],
            t["TEMPERATURE"],
            "write_file",
            "",
            t["ARTIFACT_PATH"],
        ])

    # ── Sheet 4: SESSION_CONFIG ───────────────────────────────────────────────
    ws4 = wb.create_sheet("SESSION_CONFIG")
    ws4.append(["TIGR Lifecycle Hooks"])                              # Row 1: decorative
    ws4.append(["SETTING", "VALUE", "DESCRIPTION"])                  # Row 2: headers
    ws4.append([                                                      # Row 3+: data
        "INGEST_BEFORE_RUN", "TRUE",
        "Embed 01_Raw_Source + prior 04_Code_Artifacts before swarm starts",
    ])
    ws4.append([
        "INGEST_AFTER_RUN", "TRUE",
        "Embed freshly-written 04_Code_Artifacts after swarm completes",
    ])
    ws4.append([
        "CANONIZE_AFTER_RUN", "FALSE",
        "Promote L1 session memory to L2 project memory (set TRUE after stable runs)",
    ])

    TIGR_SILO.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] Workbook written: {OUTPUT_PATH}")
    print("[OK] Sheets: SWARM_REQUEST, AGENTS, TOPOLOGY, SESSION_CONFIG")
    print("[OK] Topology: EXTRACT_CANON → TRANSLATE_CONCEPTS → BUILD_UNIVERSE → STOP")
    print(f"[OK] Agents:   {len(AGENTS)} registered")
    print("[OK] Lifecycle: INGEST_BEFORE_RUN=TRUE  INGEST_AFTER_RUN=TRUE")
    print("\nTo launch:")
    print("  python maccre.py launch TIGR --yes")



if __name__ == "__main__":
    build()
