"""
scripts/populate_sequel_test.py
=================================
Populates MACCRE_Global.xlsx for the Phase 0 Sequel run:
  GOLDEN_EGG_SEQUEL — adult narrative continuation.

The original story is embedded in the payload alongside the sequel premise.
Three narrator agents in the same linear chain, tuned for an adult register.

Usage:
    python scripts/populate_sequel_test.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT    = Path(__file__).parent.parent
WB_PATH = ROOT / "MACCRE_Global.xlsx"

ORIGINAL_STORY_PATH = (
    ROOT / "__DATACENTER" / "GOLDEN_EGG" / "04_Code_Artifacts" / "golden_egg_original.md"
)

# ── Agent Persona ─────────────────────────────────────────────────────────────
NARRATOR_PERSONA = (
    "You are a literary fiction author writing the second installment of a two-part story. "
    "Your readers are adults, average age 35, who read the original story as children and "
    "carry a deep emotional connection to it. Your prose should be mature, reflective, and "
    "psychologically layered — the language of someone who has lived, lost, and learned. "
    "Write ONLY your assigned chapter — no meta-commentary, no preamble, no labels beyond "
    "the chapter heading. Begin immediately. Match the careful, intimate tone of the preceding "
    "chapters. The emotional register is grief, wonder, and hard-won grace."
)

# ── Chapter Instruction Overrides ─────────────────────────────────────────────
CHAPTER_1_OVERRIDE = (
    "You are writing CHAPTER ONE of the sequel. "
    "Leo is now 35 years old. His father has just died. At the funeral home, going through "
    "his father's suit — the suit Leo himself chose and delivered — Leo discovers something "
    "impossible in the inside breast pocket: the golden egg. His own initials still engraved "
    "on its side. He is certain, absolutely certain, it was not there when he handed the suit "
    "over. Establish the weight of grief, the shock of the impossible return, and the first "
    "tender collision between memory and mystery. End with Leo holding the egg again for the "
    "first time in 26 years — do not yet explain anything or resolve anything."
)

CHAPTER_2_OVERRIDE = (
    "You are writing CHAPTER TWO of the sequel. "
    "Leo takes the egg home from the funeral. The grief and the mystery entangle. He begins "
    "trying to understand — did his father know about the egg all along? Did it appear to "
    "his father too, in some private moment Leo never knew about? Explore the complicated "
    "love between a father and son, the things left unsaid, and the way the egg now feels "
    "different in his adult hands — heavier with meaning, not lighter. A memory or object "
    "from Leo's childhood (something small and specific) becomes a key. Build toward the "
    "moment Leo understands what lesson the egg is here to teach him now — but do not "
    "deliver the resolution yet. End at the threshold."
)

CHAPTER_3_OVERRIDE = (
    "You are writing CHAPTER THREE of the sequel — the final chapter of the entire duology. "
    "Bring Leo's adult story to its resolution. The lesson the egg carries now is different "
    "from the lesson it carried for the child — then it taught him that true treasure is not "
    "material; now, as a grieving adult, it must teach him something harder and more specific "
    "to this moment in his life. The ending should feel earned by both stories simultaneously "
    "— it should close Leo's childhood and his adulthood in a single gesture. The egg will "
    "disappear again. How it disappears and what Leo does in that moment is the culmination. "
    "Write with restraint. The most powerful things go unsaid."
)


def _extract_prose(ledger_path: Path) -> str:
    """Pull clean prose from a ledger file, stripping tool-call wrappers."""
    text = ledger_path.read_text(encoding="utf-8")
    match = re.search(r'write_file - ({.*})\]', text, re.DOTALL)
    if match:
        try:
            payload = json.loads(match.group(1))
            return str(payload.get("data", "")).strip()
        except Exception:
            pass
    return text.strip()


def build_payload() -> str:
    """Assemble: original story + sequel premise."""
    if ORIGINAL_STORY_PATH.exists():
        original = ORIGINAL_STORY_PATH.read_text(encoding="utf-8").strip()
    else:
        # Fallback: try to extract from ledger files directly
        ledger_dir = ROOT / "__DATACENTER" / "GOLDEN_EGG" / "03_Agent_Ledgers"
        parts: list[str] = []
        for f in sorted(ledger_dir.rglob("CHAPTER_*.md")):
            parts.append(_extract_prose(f))
        original = "\n\n---\n\n".join(parts)

    sequel_premise = (
        "SEQUEL DIRECTIVE — FOR NARRATOR AGENTS ONLY\n"
        "============================================\n"
        "The original story above is the foundation. You are writing its sequel.\n\n"
        "26 years have passed. Leo is 35. His father has died.\n"
        "At the funeral home, reaching into his father's suit pocket — the suit Leo chose "
        "and delivered himself — he finds the golden egg. His initials are still on it. "
        "It was not there when he handed the suit over. He is certain.\n\n"
        "The egg has returned. It carries a new lesson for a man, not a boy.\n"
        "Write for readers who are 35 years old and remember the original story from childhood. "
        "The emotional register is grief, wonder, and hard-won grace. "
        "Prose should be mature, spare, and psychologically honest."
    )

    return f"ORIGINAL STORY — THE GOLDEN EGG\n{'=' * 48}\n{original}\n\n{'=' * 48}\n\n{sequel_premise}"


def _clear_data_rows(ws: object, start_row: int, end_row: int) -> None:  # type: ignore[type-arg]
    for r in range(start_row, end_row + 1):
        for cell in ws[r]:  # type: ignore[index]
            cell.value = None  # type: ignore[union-attr]


def _header_col(ws: object, target: str, header_row: int = 2) -> int | None:  # type: ignore[type-arg]
    for cell in ws[header_row]:  # type: ignore[index]
        if cell.value and target.lower() in str(cell.value).lower():
            return int(cell.column)
    return None


def _write_row(ws: object, row: int, data: dict[str, str], header_row: int = 2) -> None:  # type: ignore[type-arg]
    for key, value in data.items():
        col = _header_col(ws, key, header_row)
        if col is not None:
            ws.cell(row=row, column=col, value=value)  # type: ignore[union-attr]


def populate(wb_path: Path) -> None:
    payload = build_payload()
    print(f"[POPULATE] Payload length: {len(payload)} chars")

    print(f"[POPULATE] Loading workbook: {wb_path}")
    wb = load_workbook(filename=str(wb_path))

    # ── PROJECT_DEFINITION ────────────────────────────────────────────────────
    ws_pd = wb["PROJECT_DEFINITION"]
    _clear_data_rows(ws_pd, 3, 12)
    kv: dict[str, str] = {
        "PROJECT_NAME":   "GOLDEN_EGG_SEQUEL",
        "DESCRIPTION":    "Sequel narrative: adult Leo finds the egg at his father's funeral.",
        "VERSION":        "0.1",
        "AUTHOR":         "MACCREv2 Phase 0 Sequel",
        "COMPUTE_TIER":   "cloud",
        "OUTPUT_FORMAT":  "markdown",
    }
    for idx, (k, v) in enumerate(kv.items()):
        ws_pd.cell(row=3 + idx, column=1, value=k)
        ws_pd.cell(row=3 + idx, column=2, value=v)
    print("[POPULATE] PROJECT_DEFINITION -> OK")

    # ── AGENTS ────────────────────────────────────────────────────────────────
    ws_ag = wb["AGENTS"]
    _clear_data_rows(ws_ag, 3, 12)
    for i, name in enumerate(["Narrator_1", "Narrator_2", "Narrator_3"]):
        _write_row(ws_ag, 3 + i, {
            "AGENT_NAME":  name,
            "MODEL":       "gemini-2.5-flash",
            "TEMPERATURE": "1.0",
            "TOOLS":       "write_file",
            "ROLE":        "Literary fiction author — adult sequel narrator",
            "PERSONA":     NARRATOR_PERSONA,
        })
    print("[POPULATE] AGENTS -> OK")

    # ── TOPOLOGY ─────────────────────────────────────────────────────────────
    ws_tp = wb["TOPOLOGY"]
    _clear_data_rows(ws_tp, 3, 12)
    nodes: list[dict[str, str]] = [
        {
            "NODE_ID":              "CHAPTER_1",
            "AGENT_NAME":          "Narrator_1",
            "AUTO_TOOL":           "write_file",
            "NEXT_NODE":           "CHAPTER_2",
            "INSTRUCTION_OVERRIDE": CHAPTER_1_OVERRIDE,
            "MODEL_OVERRIDE":      "",
            "TEMPERATURE":         "1.0",
        },
        {
            "NODE_ID":              "CHAPTER_2",
            "AGENT_NAME":          "Narrator_2",
            "AUTO_TOOL":           "write_file",
            "NEXT_NODE":           "CHAPTER_3",
            "INSTRUCTION_OVERRIDE": CHAPTER_2_OVERRIDE,
            "MODEL_OVERRIDE":      "",
            "TEMPERATURE":         "1.0",
        },
        {
            "NODE_ID":              "CHAPTER_3",
            "AGENT_NAME":          "Narrator_3",
            "AUTO_TOOL":           "write_file",
            "NEXT_NODE":           "DONE",
            "INSTRUCTION_OVERRIDE": CHAPTER_3_OVERRIDE,
            "MODEL_OVERRIDE":      "",
            "TEMPERATURE":         "0.9",
        },
    ]
    for i, node in enumerate(nodes):
        _write_row(ws_tp, 3 + i, node)
    print("[POPULATE] TOPOLOGY -> OK")

    # ── SWARM_REQUEST ─────────────────────────────────────────────────────────
    ws_sr = wb["SWARM_REQUEST"]
    _clear_data_rows(ws_sr, 3, 6)
    _write_row(ws_sr, 3, {
        "PROJECT_NAME":   "GOLDEN_EGG_SEQUEL",
        "DESCRIPTION":    "Adult sequel — Leo finds the egg at his father's funeral.",
        "COMPUTE_TIER":   "cloud",
        "PAYLOAD_TEXT":   payload,
        "PAYLOAD_PATH":   "",
        "START_NODE":     "CHAPTER_1",
        "OUTPUT_FOLDER":  "04_Code_Artifacts",
        "NOTIFY_WEBHOOK": "",
    })
    print("[POPULATE] SWARM_REQUEST -> OK")

    # ── EXECUTION_PLAN ────────────────────────────────────────────────────────
    if "EXECUTION_PLAN" in wb.sheetnames:
        ws_ep = wb["EXECUTION_PLAN"]
        for row in ws_ep.iter_rows(min_row=3):
            first = row[0].value
            if first and "SWARM" in str(first).upper():
                if len(row) >= 3:
                    row[2].value = "YES"
        print("[POPULATE] EXECUTION_PLAN marked EXECUTE=YES")

    wb.save(str(wb_path))
    print(f"\n[POPULATE] Workbook saved: {wb_path}")
    print("[POPULATE] Ready:  python maccre.py global --yes")


if __name__ == "__main__":
    populate(WB_PATH)
