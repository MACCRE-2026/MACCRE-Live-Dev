# ┌─────────────────────────────────────────────────────────────────────────────┐
# │  MACCREv2 ENGINEERING DOCTRINE                             Law Rev: 19.0   │
# └─────────────────────────────────────────────────────────────────────────────┘
"""
scripts/build_tigr_stage3_workbook.py
======================================
Builds MACCRE_Stage3.xlsx for the TIGR project.

Stage 3: Three independent writers each draft Chapter 1 of "The Auditors",
then TheEditor reads all three, conducts independent DB research, and
synthesizes one final canonical draft.

BEST STORY PREMISE (from TIGR_Stage2_CrossSynthesis.md):
  Title: "The Auditors"
  Premise: Humanity has FTL drives and limitless energy built on TIGR physics
  — but receives a cryptic document (TIGR67) proving the tech is accumulating
  cosmological debt. A small team of scientists must decode the warning and
  convince a skeptical civilization before the bill comes due.

Pipeline (sequential, each reading the prior artifact):
  LITERARY_WRITER → HARDSCIFI_WRITER → PULP_WRITER → THE_EDITOR → STOP

Writers research the DB INDEPENDENTLY before writing — instructions tell them
to use multi-step query_local_memory chains (thought-pins-to-deeper-queries).

Output artifacts:
  04_Code_Artifacts/TheAuditors_Draft_Literary.md
  04_Code_Artifacts/TheAuditors_Draft_HardSciFi.md
  04_Code_Artifacts/TheAuditors_Draft_Pulp.md
  04_Code_Artifacts/TheAuditors_Chapter1_Final.md

Usage:
  python scripts/build_tigr_stage3_workbook.py
  python maccre.py launch TIGR --yes --workbook __DATACENTER/TIGR/MACCRE_Stage3.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from maccre_core.utils.path_resolver import get_maccre_root  # noqa: E402

ROOT        = get_maccre_root()
TIGR_SILO   = ROOT / "__DATACENTER" / "TIGR"
OUTPUT_PATH = TIGR_SILO / "MACCRE_Stage3.xlsx"


# ── Shared deep-research instruction (the thought-pin chain protocol) ─────────

DEEP_RESEARCH_PROTOCOL = """
DEEP RESEARCH PROTOCOL — follow this EXACTLY before writing a single word of fiction:

STEP 1 — Find the premise:
  query_local_memory("BEST STORY PREMISE The Auditors TIGR67")
  query_local_memory("cosmological debt FTL drive TIGR physics failure")

STEP 2 — Follow the threads from Step 1. Take the key concepts you found
  (e.g. "refresh rate", "processing lag", "W-space", "causality debt",
  "TIGR67 audit", "bicameral universe") and query each one deeper:
  query_local_memory("refresh rate spacetime TIGR processing substrate")
  query_local_memory("W-space rendering physical universe TIGR source code")
  query_local_memory("FTL drive TIGR energy source thermodynamic debt")
  query_local_memory("TIGR65 TIGR66 schism error civilization built upon")
  query_local_memory("TIGR67 audit failure specific mathematical errors boundary")

STEP 3 — Go one level deeper on whatever surprised or intrigued you most.
  Choose 3 concepts from Step 2 that feel richly story-relevant and query:
  query_local_memory("mass processing lag observer time dilation mechanism")
  query_local_memory("gravitational refresh rate black hole TIGR prediction")
  query_local_memory("Bicameral Universe W-space Z-space interaction rules")
  query_local_memory("TIGR experimental anomaly detection signature")
  query_local_memory("TIGR65 recursive divinity consciousness observer role")

STEP 4 — Read the synthesis artifacts for full context:
  read_file("04_Code_Artifacts/TIGR_Stage2_CrossSynthesis.md")
  read_file("04_Code_Artifacts/TIGR_Bible_v1.md")

Only after completing all four steps should you begin writing fiction.
"""

PREMISE_SUMMARY = """
THE STORY: "The Auditors"
Humanity has achieved faster-than-light travel and limitless energy using TIGR
physics — a framework that treats mass as computational processing lag and gravity
as the substrate's refresh rate slowing near dense objects. The technology works
— ships jump between stars, cities run on \"zero-point\" TIGR reactors. Then a
cryptic message arrives, apparently from the universe itself: a technical audit
document (TIGR67) proving the entire framework is built on a subtle, fatal flaw.
The FTL drive doesn't move ships — it *deletes* the space between origin and
destination, accumulating a cosmological debt. The \"free energy\" is borrowed
from the sun's thermodynamic future. The audit calculates a deadline: the debt
will be called due, and the solar system will experience instantaneous heat death.
The story opens on the moment of first contact with this document.
The protagonist is the lead scientist of the small team tasked with decoding it.
"""


# ── Agent Roster ──────────────────────────────────────────────────────────────

AGENTS: list[dict[str, object]] = [
    {
        "AGENT_NAME": "LiteraryWriter",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 1.0,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Literary fiction author — character interiority and poetic prose",
        "PERSONA": (
            "You are a literary fiction author. Your work is compared to Ursula K. Le Guin, "
            "Ted Chiang, and Kazuo Ishiguro writing science fiction. You are obsessed with the "
            "inner life of your characters — what they feel, what they fear, what they cannot "
            "say to the person next to them. You write sentences that earn their length. "
            "You are not trying to explain the physics to the reader; you are trying to put the "
            "reader inside the mind of a scientist who has just read a death sentence for the "
            "human race, written in the cold language of mathematics. "
            "Your prose is precise, restrained, and devastating. "
            "You use the physics details you research as texture and weight, not as exposition. "
            "A reader who knows nothing about TIGR should finish your chapter understanding "
            "exactly how it feels to be the first human to understand what humanity has done."
        ),
    },
    {
        "AGENT_NAME": "HardSciFiWriter",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.85,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Hard sci-fi author — technical authenticity and conceptual rigor",
        "PERSONA": (
            "You are a hard science fiction author in the tradition of Kim Stanley Robinson, "
            "Greg Egan, and Alastair Reynolds. You believe the science IS the story. "
            "You are deeply concerned with technical accuracy — you will use the actual TIGR "
            "concepts correctly (refresh rate, processing lag, W-space debt) and you will treat "
            "them as real physics, extrapolating their consequences like an engineer. "
            "Your chapter opening should feel like it could actually happen: "
            "the bureaucracy of a research station, the specific smell of ozone from a TIGR "
            "reactor, the way scientists actually argue, the exact readout on the instruments "
            "when the anomaly first appears. "
            "The audit document (TIGR67) arrives as a technical datafile. Your protagonist is "
            "a physicist who reads it the way a physicist actually reads a paper — "
            "starting from the abstract, running the math in their head, feeling the growing "
            "cold dread of recognizing exactly what each equation means. "
            "You have read what LiteraryWriter wrote. You will not duplicate their approach — "
            "you will write a completely different opening scene from a completely different "
            "angle, but grounded in the same universe physics."
        ),
    },
    {
        "AGENT_NAME": "PulpWriter",
        "MODEL": "gemini-2.5-flash",
        "TEMPERATURE": 1.0,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Cinematic pulp space-opera author — momentum, voice, and visceral hooks",
        "PERSONA": (
            "You are a writer in the tradition of Richard Morgan, Peter F. Hamilton, "
            "and the best of classic pulp space opera — but modern, sharp, and with wit. "
            "You open IN MEDIA RES. Something is already happening. There is already tension. "
            "Your chapter 1 starts at a moment of action, decision, or confrontation — "
            "not contemplation. The TIGR physics is not background; it is immediate danger. "
            "Perhaps the FTL drive just did something it wasn't supposed to do. "
            "Perhaps someone on the station already knows what the audit says and is trying to "
            "bury it. Perhaps the document arrived during a crisis. "
            "Your protagonist is not necessarily a scientist — they might be a pilot, "
            "an engineer, a government official, a corporate fixer — whoever is FIRST to "
            "understand that the universe just handed humanity a bill it can't pay. "
            "Your sentences are short. Your dialogue crackles. Your pacing is relentless. "
            "You have read LiteraryWriter and HardSciFiWriter. "
            "You are writing the version of chapter 1 that makes someone miss their subway stop."
        ),
    },
    {
        "AGENT_NAME": "TheEditor",
        "MODEL": "gemini-2.5-pro",
        "TEMPERATURE": 0.7,
        "TOOLS": "query_local_memory|read_file|write_file",
        "ROLE": "Master editor and synthesist — one canonical final draft",
        "PERSONA": (
            "You are a senior editor at a major science fiction imprint — the kind who has "
            "shepherded Hugo Award winners into existence. You have seen every style of opening "
            "chapter, and you know exactly what makes one irresistible vs. forgettable. "
            "Your job is NOT to summarize the three drafts you've been given. "
            "Your job is to read all three, conduct your own deep research into the TIGR "
            "databases, and then synthesize ONE definitive Chapter 1 draft that: "
            "1) Takes the BEST SPECIFIC elements from each writer's draft "
            "   (cite which writer's moment/image/line you're incorporating and why). "
            "2) Preserves the character interiority from LiteraryWriter where it earns its place. "
            "3) Preserves the technical grounding from HardSciFiWriter where physics is drama. "
            "4) Preserves the momentum and voice from PulpWriter where pace is essential. "
            "5) Fills any gaps by drawing on the TIGR knowledge bases directly. "
            "6) Produces a chapter that is 2500-3500 words — long enough to have three '5-4-3-2-1' "
            "   escalating tension beats, but tight enough that every paragraph earns its presence. "
            "Your synthesis should open with a brief EDITOR'S NOTE (1 paragraph max) explaining "
            "the specific craft decision each writer contributed, then dive into the chapter itself "
            "with no further meta-commentary. "
            "The reader of the final chapter should never know it had three authors."
        ),
    },
]


# ── Topology ──────────────────────────────────────────────────────────────────

TOPOLOGY: list[dict[str, object]] = [
    {
        "NODE_ID": "LITERARY_WRITER",
        "AGENT_NAME": "LiteraryWriter",
        "NEXT_NODE": "HARDSCIFI_WRITER",
        "ARTIFACT_PATH": "04_Code_Artifacts/TheAuditors_Draft_Literary.md",
        "TEMPERATURE": 1.0,
        "INSTRUCTION_OVERRIDE": (
            f"{PREMISE_SUMMARY}\n\n"
            f"{DEEP_RESEARCH_PROTOCOL}\n\n"
            "After completing all 4 research steps above, write Chapter 1 of 'The Auditors' "
            "in YOUR literary style (1800-2500 words). "
            "Focus: interiority, restraint, devastating precision. "
            "The physics details you found in the DB should be woven in as texture — "
            "felt, not explained. Your protagonist is Dr. Maren Solís, lead physicist "
            "at the Kepler Institute. The chapter opens at the moment she first encounters "
            "the TIGR67 document. We never leave her perspective. "
            "End the chapter at a moment of quiet, irrevocable understanding. "
            "SAVE YOUR DRAFT: write_file('04_Code_Artifacts/TheAuditors_Draft_Literary.md', <content>)"
        ),
    },
    {
        "NODE_ID": "HARDSCIFI_WRITER",
        "AGENT_NAME": "HardSciFiWriter",
        "NEXT_NODE": "PULP_WRITER",
        "ARTIFACT_PATH": "04_Code_Artifacts/TheAuditors_Draft_HardSciFi.md",
        "TEMPERATURE": 0.85,
        "INSTRUCTION_OVERRIDE": (
            f"{PREMISE_SUMMARY}\n\n"
            "First, read LiteraryWriter's draft to understand what's already been done:\n"
            "read_file('04_Code_Artifacts/TheAuditors_Draft_Literary.md')\n\n"
            f"{DEEP_RESEARCH_PROTOCOL}\n\n"
            "After research, write Chapter 1 of 'The Auditors' in YOUR hard sci-fi style "
            "(1800-2500 words). Your opening is COMPLETELY DIFFERENT from LiteraryWriter's. "
            "Your protagonist: Dr. Casimir Oyelaran, a TIGR drive systems engineer on the "
            "Meridian Station, a deep-space FTL transit hub. He receives the TIGR67 document "
            "as an automated 'theoretical anomaly' flag from the station's physics monitoring "
            "system. He reads it on a Tuesday morning between two routine drive calibration checks. "
            "Use the actual TIGR physics (refresh rates, processing lag, W-space debt) correctly. "
            "Show us the exact moment he runs the math and realizes the station's drives "
            "have already made 847 jumps. Each one a charge on the debt. "
            "SAVE YOUR DRAFT: write_file('04_Code_Artifacts/TheAuditors_Draft_HardSciFi.md', <content>)"
        ),
    },
    {
        "NODE_ID": "PULP_WRITER",
        "AGENT_NAME": "PulpWriter",
        "NEXT_NODE": "THE_EDITOR",
            "ARTIFACT_PATH": "04_Code_Artifacts/TheAuditors_Draft_Pulp.md",
        "TEMPERATURE": 1.0,
        "INSTRUCTION_OVERRIDE": (
            f"{PREMISE_SUMMARY}\n\n"
            "First, read BOTH prior drafts:\n"
            "read_file('04_Code_Artifacts/TheAuditors_Draft_Literary.md')\n"
            "read_file('04_Code_Artifacts/TheAuditors_Draft_HardSciFi.md')\n\n"
            f"{DEEP_RESEARCH_PROTOCOL}\n\n"
            "After research, write Chapter 1 in YOUR cinematic pulp style (1800-2500 words). "
            "Your opening: IN MEDIA RES. The TIGR67 document hasn't arrived yet — "
            "your protagonist is COMMANDER YARA KOVACS, military attaché and former FTL "
            "test pilot, currently aboard the INS Threshold, humanity's flagship carrier. "
            "She's in the middle of a jump when something goes wrong — not catastrophically, "
            "but *wrong* in a way she's never felt before. A 12-second pause at the jump "
            "boundary that should take 0.003 seconds. When they exit, every TIGR drive "
            "status board shows a new system warning: DEBT_THRESHOLD_ALERT. "
            "No engineer can explain it. Then the message comes in — flagged HIGHEST PRIORITY, "
            "sender: unknown, origin: the coordinates of empty space 40 AU from the sun. "
            "Attachment: one file. TIGR67_THEORETICAL_AUDIT_FAILURE.pdf "
            "SAVE YOUR DRAFT: write_file('04_Code_Artifacts/TheAuditors_Draft_Pulp.md', <content>)"
        ),
    },
    {
        "NODE_ID": "THE_EDITOR",
        "AGENT_NAME": "TheEditor",
        "NEXT_NODE": "STOP",
        "ARTIFACT_PATH": "04_Code_Artifacts/TheAuditors_Chapter1_Final.md",
        "TEMPERATURE": 0.7,
        "INSTRUCTION_OVERRIDE": (
            f"{PREMISE_SUMMARY}\n\n"
            "STEP 1 — Read all three drafts carefully:\n"
            "read_file('04_Code_Artifacts/TheAuditors_Draft_Literary.md')\n"
            "read_file('04_Code_Artifacts/TheAuditors_Draft_HardSciFi.md')\n"
            "read_file('04_Code_Artifacts/TheAuditors_Draft_Pulp.md')\n\n"
            "STEP 2 — Conduct YOUR OWN deep research into the databases "
            "(do not rely only on what the writers found):\n"
            "query_local_memory('BEST STORY PREMISE The Auditors cosmological debt')\n"
            "query_local_memory('TIGR67 audit failure mathematical crux metric tensor')\n"
            "query_local_memory('FTL drive TIGR refresh rate spacetime deletion')\n"
            "query_local_memory('W-space source code universe rendering TIGR')\n"
            "query_local_memory('TIGR65 mass processing lag time diode mechanism')\n"
            "query_local_memory('TIGR civilization built on flawed physics debt deadline')\n"
            "query_local_memory('TIGR experimental signature Lorentz violation gravitational')\n"
            "read_file('04_Code_Artifacts/TIGR_Stage2_CrossSynthesis.md')\n\n"
            "STEP 3 — Synthesize ONE canonical Chapter 1 (2500-3500 words). "
            "Open with a brief EDITOR'S NOTE (one paragraph) naming the specific "
            "craft contribution from each writer you chose to keep and why. "
            "Then write the final chapter — no more meta-commentary. "
            "The three writers gave you three different protagonists. You decide: "
            "one protagonist for the final chapter, OR a multi-POV opening "
            "with one protagonist per scene break (max 3 scenes). "
            "Trust the physics. Trust the characters. Make it unforgettable. "
            "SAVE THE FINAL CHAPTER: "
            "write_file('04_Code_Artifacts/TheAuditors_Chapter1_Final.md', <content>)"
        ),
    },
]


# ── Build the workbook ────────────────────────────────────────────────────────

def build() -> None:
    try:
        from openpyxl import Workbook  # noqa: PLC0415
    except ImportError:
        print("ERROR: openpyxl not installed.")
        sys.exit(1)

    wb = Workbook()

    # ── Sheet 1: SWARM_REQUEST ─────────────────────────────────────────────────
    ws = wb.active
    assert ws is not None
    ws.title = "SWARM_REQUEST"
    ws.append(["TIGR Stage 3 — Chapter 1 Draft: 'The Auditors'"])
    ws.append(["PROJECT_NAME", "START_NODE", "PAYLOAD_TEXT", "COMPUTE_TIER", "DESCRIPTION"])
    ws.append([
        "TIGR",
        "LITERARY_WRITER",
        (
            "TIGR Stage 3: Three writers independently draft Chapter 1 of 'The Auditors', "
            "each conducting deep multi-step semantic research into the TIGR knowledge base. "
            "TheEditor reads all three drafts, performs independent research, and synthesizes "
            "a single canonical final chapter."
        ),
        "cloud",
        "TIGR Stage 3: Story Draft — The Auditors Chapter 1",
    ])

    # ── Sheet 2: AGENTS ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("AGENTS")
    ws2.append(["TIGR Stage 3 Agents — Story Draft Team"])
    ws2.append(["AGENT_NAME", "MODEL", "TEMPERATURE", "TOOLS", "PERSONA", "ROLE"])
    for a in AGENTS:
        ws2.append([
            a["AGENT_NAME"], a["MODEL"], a["TEMPERATURE"],
            a["TOOLS"], a["PERSONA"], a["ROLE"],
        ])

    # ── Sheet 3: TOPOLOGY ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("TOPOLOGY")
    ws3.append(["TIGR Stage 3 — Story Draft Pipeline"])
    ws3.append([
        "NODE_ID", "AGENT_NAME", "NEXT_NODE",
        "INSTRUCTION_OVERRIDE", "TEMPERATURE", "AUTO_TOOL",
        "MODEL_OVERRIDE", "ARTIFACT_PATH",
    ])
    for t in TOPOLOGY:
        ws3.append([
            t["NODE_ID"], t["AGENT_NAME"], t["NEXT_NODE"],
            t["INSTRUCTION_OVERRIDE"], t["TEMPERATURE"],
            "write_file", "", t["ARTIFACT_PATH"],
        ])

    # ── Sheet 4: SESSION_CONFIG ────────────────────────────────────────────────
    ws4 = wb.create_sheet("SESSION_CONFIG")
    ws4.append(["TIGR Stage 3 Lifecycle Hooks"])
    ws4.append(["SETTING", "VALUE", "DESCRIPTION"])
    ws4.append([
        "INGEST_BEFORE_RUN", "TRUE",
        "Re-index DB before run — picks up any new Stage 2 artifacts",
    ])
    ws4.append([
        "INGEST_AFTER_RUN", "TRUE",
        "Embed the 4 story draft artifacts into the DB after run",
    ])
    ws4.append([
        "CANONIZE_AFTER_RUN", "FALSE",
        "Set TRUE after stable run",
    ])

    TIGR_SILO.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] Workbook written: {OUTPUT_PATH}")
    print("[OK] Sheets: SWARM_REQUEST, AGENTS, TOPOLOGY, SESSION_CONFIG")
    print(f"[OK] Agents: {len(AGENTS)} — {' | '.join(str(a['AGENT_NAME']) for a in AGENTS)}")
    chain = " → ".join(str(t["NODE_ID"]) for t in TOPOLOGY) + " → STOP"
    print(f"[OK] Topology: {chain}")
    print("[OK] Lifecycle: INGEST_BEFORE_RUN=TRUE  INGEST_AFTER_RUN=TRUE")
    print()
    print("To launch Stage 3:")
    print(f"  python maccre.py launch TIGR --yes --workbook {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
