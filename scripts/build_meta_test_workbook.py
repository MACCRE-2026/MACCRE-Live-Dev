import subprocess
from pathlib import Path
import sys

# Ensure maccre_core is importable from this script directory
sys.path.insert(0, str(Path(__file__).parent.parent))

# Vendored openpyxl produces guaranteed Excel-compatible OOXML output.
sys.path.insert(0, str(Path(__file__).parent.parent / "maccre_core" / "_vendor"))
from openpyxl import load_workbook

def build_workbook():
    root_dir = Path(r"B:\EXO_GANS")
    script_path = root_dir / "scripts" / "generate_global_template.py"
    wb_path = root_dir / "MACCRE_Global.xlsx"
    
    print("[1/2] Generating baseline template...")
    subprocess.run(["python", str(script_path), "--project", "META_TEST"], cwd=str(root_dir), check=True)
    
    print(f"[2/2] Filling out workbook at {wb_path}...")
    wb = load_workbook(wb_path)
    
    # ── PROJECT_DEFINITION ──
    ws_proj = wb["PROJECT_DEFINITION"]
    ws_proj["B3"] = "META_TEST"
    ws_proj["B7"] = "MEMORY_TEST"
    
    # ── AGENTS ──
    ws_agents = wb["AGENTS"]
    # Clear existing examples (from row 3 downwards)
    for row in range(3, ws_agents.max_row + 1):
        for col in range(1, ws_agents.max_column + 1):
            ws_agents.cell(row=row, column=col).value = None
            
    # Row 3: MemoryReporter
    ws_agents.cell(row=3, column=1, value="MemoryReporter")
    ws_agents.cell(row=3, column=2, value="gemini-2.5-pro")
    ws_agents.cell(row=3, column=3, value="Researcher")
    ws_agents.cell(row=3, column=4, value=(
        "You are the MemoryReporter. Your task is to query the MEMORY_TEST project database. "
        "First, call import_foreign_vectors('MEMORY_TEST', 2.0). "
        "Then, use iterative_scoped_search to discover concepts from MEMORY_TEST. "
        "Present a detailed summary of these past memories to the group."
    ))
    ws_agents.cell(row=3, column=5, value="0.1")
    ws_agents.cell(row=3, column=6, value="import_foreign_vectors|iterative_scoped_search")
    ws_agents.cell(row=3, column=11, value="Local Only")
    ws_agents.cell(row=3, column=12, value="markdown")
    ws_agents.cell(row=3, column=13, value="standard")
    ws_agents.cell(row=3, column=14, value="cloud")
    
    # Row 4: SourceReader
    ws_agents.cell(row=4, column=1, value="SourceReader")
    ws_agents.cell(row=4, column=2, value="gemini-2.5-pro")
    ws_agents.cell(row=4, column=3, value="Researcher")
    ws_agents.cell(row=4, column=4, value=(
        "You are the SourceReader. Your task is to read the raw source documents for this project. "
        "Use read_file to read the following 4 files in the 01_Raw_Source directory: "
        "AIethics.md, HolographicInterferometry.md, LlamdaLlamdaLlamda.md, SentinelOSarch.md. "
        "Summarise their contents and present your findings to the group."
    ))
    ws_agents.cell(row=4, column=5, value="0.1")
    ws_agents.cell(row=4, column=6, value="read_file")
    ws_agents.cell(row=4, column=11, value="Local Only")
    ws_agents.cell(row=4, column=12, value="markdown")
    ws_agents.cell(row=4, column=13, value="standard")
    ws_agents.cell(row=4, column=14, value="cloud")
    
    # Row 5: Synthesizer
    ws_agents.cell(row=5, column=1, value="Synthesizer")
    ws_agents.cell(row=5, column=2, value="gemini-2.5-pro")
    ws_agents.cell(row=5, column=3, value="Synthesiser")
    ws_agents.cell(row=5, column=4, value=(
        "You are the Synthesizer. Read the chat history carefully to ingest the reports from the "
        "MemoryReporter and the SourceReader. Combine their insights into a single cohesive artifact "
        "that connects the historical memory with the new source documents. "
        "Write your final report to 04_Code_Artifacts/final_synthesis.md using the write_file tool."
    ))
    ws_agents.cell(row=5, column=5, value="0.5")
    ws_agents.cell(row=5, column=6, value="write_file")
    ws_agents.cell(row=5, column=11, value="Local Only")
    ws_agents.cell(row=5, column=12, value="markdown")
    ws_agents.cell(row=5, column=13, value="standard")
    ws_agents.cell(row=5, column=14, value="cloud")

    # ── TOPOLOGY ──
    ws_topo = wb["TOPOLOGY"]
    # Clear existing
    for row in range(3, ws_topo.max_row + 1):
        for col in range(1, ws_topo.max_column + 1):
            ws_topo.cell(row=row, column=col).value = None
            
    # Single Node for Group Dialogue
    ws_topo.cell(row=3, column=1, value="START_CONVO")
    ws_topo.cell(row=3, column=2, value="MemoryReporter")
    ws_topo.cell(row=3, column=3, value="STOP")
    ws_topo.cell(row=3, column=8, value="FAILED")  # FAILURE_TARGET
    ws_topo.cell(row=3, column=11, value="SourceReader|Synthesizer")  # DIALOGUE_PARTNER
    ws_topo.cell(row=3, column=12, value="2")  # DIALOGUE_ROUNDS
    
    # ── SWARM_REQUEST ──
    ws_req = wb["SWARM_REQUEST"]
    ws_req.cell(row=3, column=1, value="META_TEST")  # PROJECT_NAME
    ws_req.cell(row=3, column=6, value="START_CONVO")  # START_NODE
    
    wb.save(wb_path)
    wb.close()
    print("[DONE] Workbook populated.")

if __name__ == "__main__":
    build_workbook()
