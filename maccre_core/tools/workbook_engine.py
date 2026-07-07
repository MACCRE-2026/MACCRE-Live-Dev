# ┌─────────────────────────────────────────────────────────────────────────────┐
# │  MACCREv2 ENGINEERING DOCTRINE                             Law Rev: 19.0   │
# ├─────────────────────────────────────────────────────────────────────────────┤
# │  I.   TYPING      All signatures: explicit Python 3.11+ type hints.        │
# │  II.  LINTING     Zero unused imports. No wildcards. 120-char line max.    │
# │  III. PATHS       Never hardcode absolute paths. Use get_maccre_root().     │
# │                   Default params: def f(p:str='') -> None: p=p or root/x   │
# │  IV.  DATACENTER  5-Tier: 01_Raw_Source · 02_Dynamic_Context               │
# │                           03_Agent_Ledgers · 04_Code_Artifacts             │
# │                           05_Rendered_Media                                 │
# │  V.   DIAMOND     Gen: temp=1.0  ·  Critic: temp=0.1 + dataclass schema   │
# │  VI.  ABSTRACTION All I/O behind abc.ABC before any concrete driver.       │
# │  VII. TEARDOWN    try/finally on all handles (omni clean compliance).      │
# │  VIII.TELEMETRY   No bare print(). logger only. JSON → 03_Agent_Ledgers.  │
# └─────────────────────────────────────────────────────────────────────────────┘
"""
maccre_core/tools/workbook_engine.py
======================================
Workbook Completeness Engine — Phases 5 & 6.

Provides:
  - check_workbook_completeness(): analyse an xlsx for section readiness,
    compute per-section FinOps estimates, print the EXECUTION_PLAN table.
  - execute_workbook_sections(): run only the operator-approved sections.
  - get_pricing_table(): pull live prices from maccre_router for workbook reference.

Design:
  The completeness engine reads the workbook read-only, scores each section,
  and returns a structured plan. The caller (maccre.py global / launch) decides
  whether to prompt for [y/N] or proceed automatically.
"""
from __future__ import annotations

import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# openpyxl lives in _vendor when not installed into the active venv
try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ModuleNotFoundError:
    _vendor = Path(__file__).parent.parent / "_vendor"
    sys.path.insert(0, str(_vendor))
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # noqa: PLC0415

logger = logging.getLogger("maccre_core")

# ── Section readiness thresholds ──────────────────────────────────────────────
# (required_field_names, min_rows_for_partial)
_SECTION_RULES: dict[str, dict[str, Any]] = {
    "PROJECT_DEFINITION": {
        "required": ["PROJECT_NAME"],
        "optional": ["DESCRIPTION", "SESSION_LABEL"],
        "type": "kv",   # key-value layout
    },
    "SWARM_REQUEST": {
        "required": ["PROJECT_NAME", "START_NODE"],
        "optional": ["PAYLOAD_TEXT", "PAYLOAD_PATH", "COMPUTE_TIER"],
        "type": "row",
    },
    "AGENTS": {
        "required": ["AGENT_NAME", "MODEL"],
        "optional": ["TEMPERATURE", "TOOLS", "PERSONA"],
        "type": "table",
        "min_rows": 1,
    },
    "TOPOLOGY": {
        "required": ["NODE_ID", "AGENT_NAME", "NEXT_NODE"],
        "optional": ["AUTO_TOOL", "TEMPERATURE", "INSTRUCTION_OVERRIDE"],
        "type": "table",
        "min_rows": 1,
    },
    "SESSION_CONFIG": {
        "required": ["SESSION_LABEL"],
        "optional": ["INGEST_BEFORE_RUN", "CANONIZE_AFTER_RUN", "OUTPUT_FORMATS"],
        "type": "kv",
    },
}

# Estimated average output tokens per agent node (used for rough FinOps)
_AVG_NODE_TOKENS: int = 20000


# ── FinOps pricing reference ───────────────────────────────────────────────────


def get_pricing_table() -> dict[str, dict[str, float]]:
    """Return the live model pricing table from maccre_router.

    Falls back to a hardcoded minimal table if the router is unavailable.

    Returns:
        Dict keyed by model_id: {"input_mtok": float, "output_mtok": float}
    """
    try:
        from maccre_core.maccre_router import UniversalRouter  # noqa: PLC0415
        router = UniversalRouter()
        raw: dict[str, Any] = getattr(router, "_PRICING_TABLE", {})
        if raw:
            return {k: {"input_mtok": v.get("in", 0.0), "output_mtok": v.get("out", 0.0)}
                    for k, v in raw.items()}
    except Exception:  # noqa: BLE001
        pass
    # Fallback minimal table
    return {
        "gemini-2.5-flash":   {"input_mtok": 0.15,  "output_mtok": 0.60},
        "gemini-2.5-pro":     {"input_mtok": 1.25,  "output_mtok": 5.00},
        "gemini-2.5-flash-8b":{"input_mtok": 0.075, "output_mtok": 0.30},
        "gemma3:9b":          {"input_mtok": 0.0,   "output_mtok": 0.0},
    }


def _estimate_node_cost(model: str, pricing: dict[str, dict[str, float]]) -> float:
    """Estimate API cost for one topology node using average token count."""
    row = pricing.get(model) or pricing.get("gemini-2.5-flash", {})
    per_tok = row.get("output_mtok", 0.0) / 1_000_000
    return round(per_tok * _AVG_NODE_TOKENS, 6)


# ── Parsing helpers ────────────────────────────────────────────────────────────


def _norm(val: Any) -> str:
    return str(val or "").lstrip("★* ").strip().upper().replace(" ", "_")


def _s(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def _header_map(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[2], start=1):
        k = _norm(cell.value)
        if k:
            result[k] = col_idx
    return result


def _get_cell(ws: Any, row_idx: int, col_name: str, hmap: dict[str, int]) -> str:
    col = hmap.get(col_name)
    if col is None:
        return ""
    cell = ws.cell(row=row_idx, column=col)
    return _s(cell.value)


# ── Dataclasses ────────────────────────────────────────────────────────────────


@dataclass
class SectionResult:
    """Readiness assessment for one workbook section."""
    name: str
    status: str          # "READY" | "PARTIAL" | "INCOMPLETE" | "MISSING"
    filled: int = 0      # required fields satisfied
    required: int = 0
    execute: bool = True # operator toggle (from EXECUTION_PLAN sheet or default)
    est_cost_usd: float = 0.0
    notes: list[str] = field(default_factory=list)


@dataclass
class WorkbookPlan:
    """Complete completeness assessment for an xlsx workbook."""
    workbook_path: Path
    wb_type: str                        # "global" or "session"
    sections: list[SectionResult]
    project_name: str = ""
    session_label: str = ""
    total_est_cost: float = 0.0

    @property
    def actionable_sections(self) -> list[SectionResult]:
        return [s for s in self.sections if s.execute and s.status in ("READY", "PARTIAL")]

    @property
    def can_run(self) -> bool:
        return len(self.actionable_sections) > 0


# ── Completeness Engine ────────────────────────────────────────────────────────


def check_workbook_completeness(path: Path, wb_type: str = "session") -> WorkbookPlan:
    """Analyse an xlsx workbook and compute a readiness plan for each section.

    Args:
        path:    Absolute path to the workbook.
        wb_type: 'global' or 'session' — determines which sheets to evaluate.

    Returns:
        WorkbookPlan with per-section SectionResult entries.
    """
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_names: list[str] = wb.sheetnames
    pricing = get_pricing_table()

    # Determine which sections to check for this workbook type
    if wb_type == "global":
        candidate_sections = ["PROJECT_DEFINITION", "AGENTS", "TOPOLOGY", "SWARM_REQUEST"]
    else:
        candidate_sections = ["SESSION_CONFIG", "SWARM_REQUEST", "AGENTS", "TOPOLOGY"]

    results: list[SectionResult] = []
    project_name = ""
    session_label = ""

    # ── Read EXECUTION_PLAN execute toggles if present ────────────────────────
    exec_toggles: dict[str, bool] = {}
    if "EXECUTION_PLAN" in sheet_names:
        ws_ep = wb["EXECUTION_PLAN"]
        ep_hmap = _header_map(ws_ep)
        for row in ws_ep.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            section = _norm(row[0])
            # EXECUTE? column — bool checkbox or TRUE/FALSE string
            col = ep_hmap.get("EXECUTE")
            if col and col <= len(row):
                raw = row[col - 1]
                exec_toggles[section] = str(raw or "").strip().upper() not in ("FALSE", "NO", "0", "")
            else:
                exec_toggles[section] = True

    # ── PROJECT_DEFINITION (kv sheet) ─────────────────────────────────────────
    if "PROJECT_DEFINITION" in sheet_names and "PROJECT_DEFINITION" in candidate_sections:
        ws = wb["PROJECT_DEFINITION"]
        kv: dict[str, str] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            k = _norm(row[0])
            v = _s(row[1]) if len(row) > 1 else ""
            if k:
                kv[k] = v
        project_name = kv.get("PROJECT_NAME", "")
        session_label = kv.get("SESSION_LABEL", "")
        notes = []
        filled = 1 if project_name else 0
        if not project_name:
            notes.append("PROJECT_NAME is required")
        status = "READY" if project_name else "INCOMPLETE"
        results.append(SectionResult(
            name="PROJECT_DEFINITION",
            status=status,
            filled=filled,
            required=1,
            execute=exec_toggles.get("PROJECT_DEFINITION", True),
            est_cost_usd=0.0,
            notes=notes,
        ))

    elif "PROJECT_DEFINITION" in candidate_sections:
        results.append(SectionResult("PROJECT_DEFINITION", "MISSING", notes=["Sheet not found"]))

    # ── SESSION_CONFIG (kv sheet) ─────────────────────────────────────────────
    if "SESSION_CONFIG" in sheet_names and "SESSION_CONFIG" in candidate_sections:
        ws = wb["SESSION_CONFIG"]
        kv = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            k = _norm(row[0])
            v = _s(row[1]) if len(row) > 1 else ""
            if k:
                kv[k] = v
        session_label = kv.get("SESSION_LABEL", "")
        # PROJECT_NAME may also come from session config
        if not project_name:
            project_name = kv.get("PROJECT_NAME", "")
        notes: list[str] = []
        filled = 1 if session_label else 0
        if not session_label:
            notes.append("SESSION_LABEL recommended (will auto-generate if empty)")
            filled = 0
        results.append(SectionResult(
            name="SESSION_CONFIG",
            status="READY",  # session config is always optional
            filled=filled,
            required=1,
            execute=exec_toggles.get("SESSION_CONFIG", True),
            notes=notes,
        ))

    elif "SESSION_CONFIG" in candidate_sections:
        results.append(SectionResult("SESSION_CONFIG", "MISSING", notes=["Sheet not found"]))

    # ── AGENTS (table sheet) ──────────────────────────────────────────────────
    if "AGENTS" in sheet_names and "AGENTS" in candidate_sections:
        ws = wb["AGENTS"]
        hmap = _header_map(ws)
        rows_data: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            name_col = hmap.get("AGENT_NAME")
            if not name_col or not row or len(row) < name_col or not row[name_col - 1]:
                continue
            rows_data.append({
                "name":  _s(row[hmap.get("AGENT_NAME", 1) - 1] if hmap.get("AGENT_NAME") else None),
                "model": _s(row[hmap.get("MODEL", 2) - 1] if hmap.get("MODEL") else None),
            })
        notes = []
        agent_cost = 0.0
        if rows_data:
            for ag in rows_data:
                m = ag.get("model", "gemini-2.5-flash") or "gemini-2.5-flash"
                agent_cost += _estimate_node_cost(m, pricing)
            status = "READY"
            filled = len(rows_data)
        else:
            status = "INCOMPLETE"
            filled = 0
            notes.append("No agents defined")
        results.append(SectionResult(
            name="AGENTS",
            status=status,
            filled=filled,
            required=1,
            execute=exec_toggles.get("AGENTS", True),
            est_cost_usd=round(agent_cost, 6),
            notes=notes,
        ))

    elif "AGENTS" in candidate_sections:
        results.append(SectionResult("AGENTS", "MISSING", notes=["Sheet not found"]))

    # ── TOPOLOGY (table sheet) ────────────────────────────────────────────────
    if "TOPOLOGY" in sheet_names and "TOPOLOGY" in candidate_sections:
        ws = wb["TOPOLOGY"]
        hmap = _header_map(ws)
        node_models: list[str] = []
        nodes_missing_agent: list[str] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            nid_col = hmap.get("NODE_ID")
            if not nid_col or not row or len(row) < nid_col or not row[nid_col - 1]:
                continue
            node_id = _s(row[nid_col - 1])
            agent_col = hmap.get("AGENT_NAME")
            agent = _s(row[agent_col - 1]) if agent_col and len(row) >= agent_col else ""
            model_col = hmap.get("MODEL_OVERRIDE")
            model = _s(row[model_col - 1]) if model_col and len(row) >= model_col else ""
            node_models.append(model or "gemini-2.5-flash")
            if not agent:
                nodes_missing_agent.append(node_id)
        notes = []
        topo_cost = sum(_estimate_node_cost(m, pricing) for m in node_models)
        if node_models and not nodes_missing_agent:
            status = "READY"
            filled = len(node_models)
        elif node_models and nodes_missing_agent:
            status = "PARTIAL"
            filled = len(node_models) - len(nodes_missing_agent)
            notes.append(f"Nodes missing AGENT_NAME: {', '.join(nodes_missing_agent)}")
        else:
            status = "INCOMPLETE"
            filled = 0
            notes.append("No nodes defined")
        results.append(SectionResult(
            name="TOPOLOGY",
            status=status,
            filled=filled,
            required=max(len(node_models), 1),
            execute=exec_toggles.get("TOPOLOGY", True),
            est_cost_usd=round(topo_cost, 6),
            notes=notes,
        ))

    elif "TOPOLOGY" in candidate_sections:
        results.append(SectionResult("TOPOLOGY", "MISSING", notes=["Sheet not found"]))

    # ── SWARM_REQUEST (single row) ────────────────────────────────────────────
    if "SWARM_REQUEST" in sheet_names and "SWARM_REQUEST" in candidate_sections:
        ws = wb["SWARM_REQUEST"]
        hmap = _header_map(ws)
        req_rows = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))
        req: tuple[Any, ...] = req_rows[0] if req_rows else ()

        def _get_req(key: str) -> str:
            col = hmap.get(key)
            if col is None or col > len(req):
                return ""
            return _s(req[col - 1])

        pname     = _get_req("PROJECT_NAME")
        start     = _get_req("START_NODE")
        payload_t = _get_req("PAYLOAD_TEXT")
        payload_p = _get_req("PAYLOAD_PATH")
        has_payload = bool(payload_t.strip() or payload_p.strip())
        if not project_name and pname:
            project_name = pname
        notes = []
        filled = 0
        if pname:
            filled += 1
        else:
            notes.append("PROJECT_NAME missing")
        if start:
            filled += 1
        else:
            notes.append("START_NODE missing")
        if not has_payload:
            notes.append("No payload (PAYLOAD_TEXT or PAYLOAD_PATH) — swarm will not run")
        if filled == 2 and has_payload:
            status = "READY"
        elif filled == 2 and not has_payload:
            status = "PARTIAL"
        elif filled >= 1:
            status = "PARTIAL"
        else:
            status = "INCOMPLETE"
        results.append(SectionResult(
            name="SWARM_REQUEST",
            status=status,
            filled=filled,
            required=2,
            execute=exec_toggles.get("SWARM_REQUEST", True) and has_payload,
            notes=notes,
        ))

    elif "SWARM_REQUEST" in candidate_sections:
        results.append(SectionResult("SWARM_REQUEST", "MISSING", notes=["Sheet not found"]))

    wb.close()

    total_cost = round(sum(s.est_cost_usd for s in results), 6)
    return WorkbookPlan(
        workbook_path=path,
        wb_type=wb_type,
        sections=results,
        project_name=project_name,
        session_label=session_label,
        total_est_cost=total_cost,
    )


# ── Terminal Renderer ──────────────────────────────────────────────────────────


def render_execution_plan(plan: WorkbookPlan) -> str:
    """Render the EXECUTION_PLAN summary table for the operator terminal.

    Args:
        plan: WorkbookPlan from check_workbook_completeness().

    Returns:
        Formatted multi-line string for printing.
    """
    STATUS_ICON = {
        "READY":      "✅ Ready    ",
        "PARTIAL":    "⚠  Partial  ",
        "INCOMPLETE": "❌ Incomplete",
        "MISSING":    "❌ Missing  ",
    }
    lines: list[str] = [
        "",
        f"┌ EXECUTION_PLAN ─── {plan.workbook_path.name} {'─' * 30}",
        f"│ {'Section':<20} {'Status':<14} {'Execute?':<9} {'Est. Cost':>10}  Notes",
        "├" + "─" * 74,
    ]
    for s in plan.sections:
        icon    = STATUS_ICON.get(s.status, s.status)
        toggle  = "[x]" if s.execute else "[ ]"
        cost    = f"${s.est_cost_usd:.4f}" if s.est_cost_usd else "--      "
        note    = s.notes[0][:28] if s.notes else ""
        lines.append(f"│ {s.name:<20} {icon:<14} {toggle:<9} {cost:>10}  {note}")
    lines += [
        "└" + "─" * 74,
        f"  Estimated session cost: ${plan.total_est_cost:.4f}",
    ]

    actionable = plan.actionable_sections
    if actionable:
        names = " + ".join(s.name for s in actionable)
        lines.append(f"  {len(actionable)} of {len(plan.sections)} sections actionable: {names}")
    else:
        lines.append("  ⚠  No sections actionable — fill in required fields and re-run.")

    lines.append("")
    return "\n".join(lines)
