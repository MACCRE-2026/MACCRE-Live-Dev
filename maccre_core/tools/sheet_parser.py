# ┌─────────────────────────────────────────────────────────────────────────────┐
# │  MACCREv2 ENGINEERING DOCTRINE                             Law Rev: 19.0   │
# ├─────────────────────────────────────────────────────────────────────────────┤
# │  I.   TYPING      All signatures: explicit Python 3.11+ type hints.        │
# │  II.  LINTING     Zero unused imports. No wildcards. 120-char line max.    │
# │  III. PATHS       Never hardcode absolute paths. Use get_maccre_root().     │
# │                   Default params: def f(p:str='') -> None: p=p or root/x   │
# │  IV.  DATACENTER  5-Tier: 01_Raw_Source · 02_Dynamic_Context               │
# │                           03_Agent_Ledgers · 04_Code_Artifacts             │
# │                           05_Rendered_Media                                 │
# │  V.   DIAMOND     Gen: temp=1.0  ·  Critic: temp=0.1 + dataclass schema   │
# │  VI.  ABSTRACTION All I/O behind abc.ABC before any concrete driver.       │
# │  VII. TEARDOWN    try/finally on all handles (omni clean compliance).      │
# │  VIII.TELEMETRY   No bare print(). logger only. JSON → 03_Agent_Ledgers.  │
# └─────────────────────────────────────────────────────────────────────────────┘
"""
maccre_core/tools/sheet_parser.py
===================================
Parses MACCRE_Swarm_Request.xlsx into typed internal structures.

This module is the bridge between the Excel intake surface and the core
materialisation engine. It replaces agent_roster.csv, topology.csv, and
project_schema.json with a single portable workbook.

Machine-readability contract (matches generate_template.py):
  - Row 1: decorative title — NEVER parsed.
  - Row 2: column headers (may be prefixed with "* " for required fields).
  - Row 3+: data rows — empty rows are silently skipped.
  - Column order is irrelevant — parser resolves by normalised header name.
  - PIPELINE_CONFIG / MEMORY_CONFIG / VAULT_KEYS use SETTING | VALUE layout.
"""
from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import sys

# openpyxl lives in _vendor when not installed into the active venv
try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ModuleNotFoundError:
    _vendor = Path(__file__).parent.parent / "_vendor"
    sys.path.insert(0, str(_vendor))
    from openpyxl import load_workbook  # type: ignore[import-untyped]  # noqa: PLC0415

from maccre_core.tools.design_tools import AgentDesign, NodeDesign, SwarmDesign

logger = logging.getLogger("maccre_core")


# ── Types ──────────────────────────────────────────────────────────────────────


@dataclass
class AgentExtra:
    """AI Studio parameters not yet wired to the router — captured for future use."""

    agent_name: str
    top_p: float | None = None
    top_k: int | None = None
    max_output_tokens: int | None = None
    thinking_budget: int | None = None        # -1=auto, 0=off, N=budget
    search_grounding: bool = False
    brave_search: bool = False
    url_context: list[str] = field(default_factory=list)
    response_format: str = "text"              # text | json | markdown
    safety_level: str = "standard"             # minimal | standard | strict
    compute_tier: str = "cloud"                # cloud | local | hybrid


@dataclass
class ParsedWorkbook:
    """Complete, typed representation of MACCRE_Swarm_Request.xlsx."""

    # ── SWARM_REQUEST ──────────────────────────────────────────────────────────
    project_name: str
    description: str
    compute_tier: str
    payload_text: str
    payload_path: str
    start_node: str
    output_folder: str
    notify_webhook: str

    # ── AGENTS / TOPOLOGY ─────────────────────────────────────────────────────
    agents: list[AgentDesign]               # core fields — used by materialiser
    agent_extras: dict[str, AgentExtra]     # extended AI Studio params — future router
    topology: list[NodeDesign]

    # ── CONFIG SHEETS ─────────────────────────────────────────────────────────
    pipeline_config: dict[str, str]
    memory_config: dict[str, str]
    vault_refs: dict[str, str]
    linked_projects: list[str]


# ── Helpers ────────────────────────────────────────────────────────────────────


def _norm(raw: Any) -> str:
    """Normalise a cell value to a clean uppercase key."""
    return str(raw or "").lstrip("\u2605* ").strip().upper().replace(" ", "_")


def _s(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def _f(val: Any, default: float = 0.7) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _i(val: Any, default: int = 0) -> int:
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _b(val: Any) -> bool:
    return str(val or "").strip().upper() in ("TRUE", "YES", "1")


def _header_map(ws: Any) -> dict[str, int]:
    """Map normalised column name → 1-based col index from row 2."""
    result: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[2], start=1):
        key = _norm(cell.value)
        if key:
            result[key] = col_idx
    return result


def _get(row: tuple[Any, ...], hmap: dict[str, int], key: str, default: str = "") -> str:
    col = hmap.get(key)
    if col is None or col > len(row):
        return default
    return _s(row[col - 1]) or default


def _kv_sheet(ws: Any, start_row: int = 3) -> dict[str, str]:
    """Read SETTING | VALUE sheets (PIPELINE_CONFIG, MEMORY_CONFIG, VAULT_KEYS)."""
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if not row or not row[0]:
            continue
        k = _norm(row[0])
        v = _s(row[1]) if len(row) > 1 else ""
        if k:
            result[k] = v
    return result


# ── Parser ─────────────────────────────────────────────────────────────────────


def parse_workbook(path: Path) -> ParsedWorkbook:
    """Parse a MACCRE_Swarm_Request.xlsx file into a ParsedWorkbook.

    Args:
        path: Absolute path to the .xlsx file.

    Returns:
        ParsedWorkbook with all typed fields populated.

    Raises:
        ValueError: If required fields are missing or sheets are absent.
        FileNotFoundError: If the path does not exist.
    """
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    logger.info("[SheetParser] Loading workbook: %s", path)
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)

    sheet_names: list[str] = wb.sheetnames

    # ── SWARM_REQUEST ──────────────────────────────────────────────────────────
    if "SWARM_REQUEST" not in sheet_names:
        raise ValueError("Workbook is missing the 'SWARM_REQUEST' sheet.")

    ws_req = wb["SWARM_REQUEST"]
    req_hmap = _header_map(ws_req)
    req_rows = list(ws_req.iter_rows(min_row=3, max_row=3, values_only=True))
    req: tuple[Any, ...] = req_rows[0] if req_rows else ()

    project_name   = _get(req, req_hmap, "PROJECT_NAME")
    description    = _get(req, req_hmap, "DESCRIPTION")
    compute_tier   = _get(req, req_hmap, "COMPUTE_TIER", "cloud")
    payload_text   = _get(req, req_hmap, "PAYLOAD_TEXT")
    payload_path   = _get(req, req_hmap, "PAYLOAD_PATH")
    start_node     = _get(req, req_hmap, "START_NODE")
    output_folder  = _get(req, req_hmap, "OUTPUT_FOLDER")
    notify_webhook = _get(req, req_hmap, "NOTIFY_WEBHOOK")

    if not project_name:
        raise ValueError("SWARM_REQUEST!PROJECT_NAME is required.")
    if not start_node:
        raise ValueError("SWARM_REQUEST!START_NODE is required.")

    # Reject the placeholder example text
    _PLACEHOLDER = "paste your story"
    if _PLACEHOLDER in payload_text.lower():
        payload_text = ""

    logger.info("[SheetParser] Project=%s  start_node=%s", project_name, start_node)

    # ── AGENTS ────────────────────────────────────────────────────────────────
    if "AGENTS" not in sheet_names:
        raise ValueError("Workbook is missing the 'AGENTS' sheet.")

    ws_ag = wb["AGENTS"]
    ag_hmap = _header_map(ws_ag)
    agents: list[AgentDesign] = []
    agent_extras: dict[str, AgentExtra] = {}

    for row in ws_ag.iter_rows(min_row=3, values_only=True):
        name = _get(row, ag_hmap, "AGENT_NAME")  # type: ignore[arg-type]
        if not name:
            continue

        model       = _get(row, ag_hmap, "MODEL", "gemini-3-flash-preview")  # type: ignore[arg-type]
        temperature = _f(_get(row, ag_hmap, "TEMPERATURE", "0.7"))  # type: ignore[arg-type]
        tools       = _get(row, ag_hmap, "TOOLS", "write_file")  # type: ignore[arg-type]
        persona     = _get(row, ag_hmap, "PERSONA", f"You are {name}, a MACCREv2 swarm agent.")  # type: ignore[arg-type]
        role        = _get(row, ag_hmap, "ROLE", "Swarm agent")  # type: ignore[arg-type]

        agents.append(AgentDesign(
            name=name,
            model=model,
            role=role,
            persona_instructions=persona,
            tools=tools,
            temperature=temperature,
        ))

        # Extended AI Studio params — stored for future router use
        top_p_raw   = _get(row, ag_hmap, "TOP_P")  # type: ignore[arg-type]
        top_k_raw   = _get(row, ag_hmap, "TOP_K")  # type: ignore[arg-type]
        max_tok_raw = _get(row, ag_hmap, "MAX_OUTPUT_TOKENS")  # type: ignore[arg-type]
        think_raw   = _get(row, ag_hmap, "THINKING_BUDGET")  # type: ignore[arg-type]
        url_raw     = _get(row, ag_hmap, "URL_CONTEXT")  # type: ignore[arg-type]

        agent_extras[name] = AgentExtra(
            agent_name=name,
            top_p=float(top_p_raw) if top_p_raw else None,
            top_k=int(float(top_k_raw)) if top_k_raw else None,
            max_output_tokens=int(float(max_tok_raw)) if max_tok_raw else None,
            thinking_budget=int(float(think_raw)) if think_raw else None,
            search_grounding=_b(_get(row, ag_hmap, "SEARCH_GROUNDING")),  # type: ignore[arg-type]
            brave_search=_b(_get(row, ag_hmap, "BRAVE_SEARCH")),  # type: ignore[arg-type]
            url_context=[u.strip() for u in url_raw.split(",") if u.strip()] if url_raw else [],
            response_format=_get(row, ag_hmap, "RESPONSE_FORMAT", "text"),  # type: ignore[arg-type]
            safety_level=_get(row, ag_hmap, "SAFETY_LEVEL", "standard"),  # type: ignore[arg-type]
            compute_tier=_get(row, ag_hmap, "COMPUTE_TIER", compute_tier),  # type: ignore[arg-type]
        )
        logger.debug("[SheetParser] Agent loaded: %s (%s, T=%.1f)", name, model, temperature)

    if not agents:
        raise ValueError("AGENTS sheet: at least one populated agent row is required.")

    # ── TOPOLOGY ──────────────────────────────────────────────────────────────
    if "TOPOLOGY" not in sheet_names:
        raise ValueError("Workbook is missing the 'TOPOLOGY' sheet.")

    ws_tp = wb["TOPOLOGY"]
    tp_hmap = _header_map(ws_tp)
    topology: list[NodeDesign] = []

    for row in ws_tp.iter_rows(min_row=3, values_only=True):
        node_id = _get(row, tp_hmap, "NODE_ID")  # type: ignore[arg-type]
        if not node_id:
            continue

        topology.append(NodeDesign(
            node_id=node_id,
            agent_name=_get(row, tp_hmap, "AGENT_NAME"),  # type: ignore[arg-type]
            next_node=_get(row, tp_hmap, "NEXT_NODE", "STOP"),  # type: ignore[arg-type]
            instruction_override=_get(row, tp_hmap, "INSTRUCTION_OVERRIDE"),  # type: ignore[arg-type]
            model_override=_get(row, tp_hmap, "MODEL_OVERRIDE", ""),  # type: ignore[arg-type]
            temperature=_f(_get(row, tp_hmap, "TEMPERATURE", "0.7")),  # type: ignore[arg-type]
            max_recursion=int(float(_get(row, tp_hmap, "MAX_RECURSION", "3"))),  # type: ignore[(arg-type, call-overload)]
            wait_for=_get(row, tp_hmap, "WAIT_FOR", "none"),  # type: ignore[arg-type]
            failure_target=_get(row, tp_hmap, "FAILURE_TARGET", "FAILED"),  # type: ignore[arg-type]
            artifact_path=_get(row, tp_hmap, "ARTIFACT_PATH", ""),  # type: ignore[arg-type]
            live_profile=_get(row, tp_hmap, "LIVE_PROFILE", ""),  # type: ignore[arg-type]
            dialogue_partner=_get(row, tp_hmap, "DIALOGUE_PARTNER", ""),  # type: ignore[arg-type]
            dialogue_rounds=int(float(_get(row, tp_hmap, "DIALOGUE_ROUNDS", "0") or "0")),  # type: ignore[arg-type]
        ))
        logger.debug("[SheetParser] Node loaded: %s -> %s", node_id, topology[-1].next_node)


    if not topology:
        raise ValueError("TOPOLOGY sheet: at least one populated node row is required.")

    # ── Config sheets (optional) ───────────────────────────────────────────────
    pipeline_config = _kv_sheet(wb["PIPELINE_CONFIG"]) if "PIPELINE_CONFIG" in sheet_names else {}
    memory_config   = _kv_sheet(wb["MEMORY_CONFIG"])   if "MEMORY_CONFIG" in sheet_names else {}
    vault_refs      = _kv_sheet(wb["VAULT_KEYS"])       if "VAULT_KEYS" in sheet_names else {}
    
    linked_projects: list[str] = []
    if "PROJECT_DEFINITION" in sheet_names:
        proj_def = _kv_sheet(wb["PROJECT_DEFINITION"])
        lp_raw = proj_def.get("LINKED_PROJECTS", "")
        linked_projects = [lp.strip() for lp in lp_raw.split(",") if lp.strip()]

    wb.close()

    logger.info(
        "[SheetParser] Parsed OK — %d agents, %d nodes.",
        len(agents), len(topology),
    )
    return ParsedWorkbook(
        project_name=project_name,
        description=description,
        compute_tier=compute_tier or "cloud",
        payload_text=payload_text,
        payload_path=payload_path,
        start_node=start_node,
        output_folder=output_folder,
        notify_webhook=notify_webhook,
        agents=agents,
        agent_extras=agent_extras,
        topology=topology,
        pipeline_config=pipeline_config,
        memory_config=memory_config,
        vault_refs=vault_refs,
        linked_projects=linked_projects,
    )


# ── Materialiser ───────────────────────────────────────────────────────────────


def materialise_from_sheet(path: Path) -> str:
    """Parse an xlsx workbook and fully materialise a MACCRE swarm project.

    Steps:
    1. Parse the workbook (sheet_parser.parse_workbook).
    2. Build a SwarmDesign and call _materialise_swarm to create the workspace,
       agent roster, persona cards, and topology.
    3. Write agent_extras.json and pipeline_config.json to the project silo
       for future router enhancement.
    4. If PAYLOAD_TEXT is populated, write it to 01_Raw_Source/input.md.

    Returns:
        The [SWARM_READY] status string from _materialise_swarm, or a
        [SHEET_FAULT] string describing the error.
    """
    import os  # noqa: PLC0415

    try:
        parsed = parse_workbook(path)
    except (ValueError, FileNotFoundError) as exc:
        return f"[SHEET_FAULT] Parse error: {exc}"

    # Switch active project context so materialiser writes to the right silo
    os.environ["MACCRE_ACTIVE_PROJECT"] = parsed.project_name

    design = SwarmDesign(
        project_name=parsed.project_name,
        agents=parsed.agents,
        topology=parsed.topology,
        payload_description=parsed.description,
        missing_requirements=[],
        design_narrative=parsed.description,
    )

    from maccre_core.tools.design_tools import _materialise_swarm  # noqa: PLC0415

    try:
        result = _materialise_swarm(design)
    except Exception as exc:  # noqa: BLE001
        return f"[SHEET_FAULT] Materialisation error: {exc}"

    if "[DESIGN_FAULT]" in result:
        return result.replace("[DESIGN_FAULT]", "[SHEET_FAULT]")

    # ── Write extended config files to the project silo ───────────────────────
    try:
        from maccre_core.utils.path_resolver import get_datacenter_path  # noqa: PLC0415

        ctx_dir = get_datacenter_path("02_Dynamic_Context")

        # agent_extras.json — extended AI Studio params (captured, not yet routed)
        extras_payload = {
            name: {
                "top_p":            extra.top_p,
                "top_k":            extra.top_k,
                "max_output_tokens": extra.max_output_tokens,
                "thinking_budget":  extra.thinking_budget,
                "search_grounding": extra.search_grounding,
                "brave_search":     extra.brave_search,
                "url_context":      extra.url_context,
                "response_format":  extra.response_format,
                "safety_level":     extra.safety_level,
                "compute_tier":     extra.compute_tier,
            }
            for name, extra in parsed.agent_extras.items()
        }
        (ctx_dir / "agent_extras.json").write_text(
            json.dumps(extras_payload, indent=2), encoding="utf-8"
        )

        # pipeline_config.json
        if parsed.pipeline_config:
            (ctx_dir / "pipeline_config.json").write_text(
                json.dumps(parsed.pipeline_config, indent=2), encoding="utf-8"
            )

        # memory_config.json
        if parsed.memory_config:
            (ctx_dir / "memory_config.json").write_text(
                json.dumps(parsed.memory_config, indent=2), encoding="utf-8"
            )
        # project_schema.json
        if parsed.linked_projects:
            schema_payload = {"linked_projects": parsed.linked_projects}
            (ctx_dir / "project_schema.json").write_text(
                json.dumps(schema_payload, indent=2), encoding="utf-8"
            )
            
    except Exception as exc:  # noqa: BLE001
        logger.warning("[SheetParser] Failed to write config sidecars: %s", exc)

    # ── Write inline payload if present ───────────────────────────────────────
    if parsed.payload_text.strip():
        try:
            import re as _re  # noqa: PLC0415
            from maccre_core.tools.storage_tools import write_file  # noqa: PLC0415
            _safe_proj = _re.sub(r"[^\w\-]", "_", parsed.project_name)
            _payload_fname = f"01_Raw_Source/{_safe_proj}_payload.md"
            write_file(_payload_fname, parsed.payload_text)
            logger.info("[SheetParser] Inline payload written to %s", _payload_fname)
        except Exception as exc:  # noqa: BLE001
            logger.warning("[SheetParser] Could not write inline payload: %s", exc)

    # ── Auto-copy payload from PAYLOAD_PATH if specified ────────────────────────
    # Text files (.md .txt .json .csv .py) are read and written as the initial
    # payload, making them discoverable by the first swarm node.
    # Media files (.mp3 .wav .mp4 .jpg etc.) are wrapped in a [[MEDIA:]] token
    # so the multi-modal router can inject them as inlineData automatically.
    if parsed.payload_path.strip():
        try:
            import re as _re  # noqa: PLC0415
            from pathlib import Path as _Path  # noqa: PLC0415
            from maccre_core.tools.storage_tools import write_file  # noqa: PLC0415

            _src = _Path(parsed.payload_path.strip())
            _safe_proj = _re.sub(r"[^\w\-]", "_", parsed.project_name)
            _TEXT_EXTS = {".md", ".txt", ".json", ".csv", ".py", ".yaml", ".yml", ".xml", ".html"}
            _MEDIA_EXTS_PP = {
                ".mp3", ".wav", ".m4a", ".aac", ".flac", ".ogg",
                ".mp4", ".webm", ".mov", ".avi", ".mkv",
                ".jpg", ".jpeg", ".png", ".gif", ".webp", ".pdf",
            }
            if not _src.exists():
                logger.warning("[SheetParser] PAYLOAD_PATH not found: %s", _src)
            elif _src.suffix.lower() in _TEXT_EXTS:
                _content = _src.read_text(encoding="utf-8", errors="replace")
                _dst_fname = f"01_Raw_Source/{_safe_proj}_from_path.md"
                write_file(_dst_fname, _content)
                logger.info("[SheetParser] PAYLOAD_PATH copied to %s (%d chars)", _dst_fname, len(_content))
            elif _src.suffix.lower() in _MEDIA_EXTS_PP:
                # Wrap in [[MEDIA:]] token so the router injects it as inlineData
                _media_payload = f"[[MEDIA: {_src.absolute()}]]\n"
                _dst_fname = f"01_Raw_Source/{_safe_proj}_from_path.md"
                write_file(_dst_fname, _media_payload)
                logger.info("[SheetParser] PAYLOAD_PATH media token written to %s", _dst_fname)
            else:
                logger.warning("[SheetParser] PAYLOAD_PATH extension '%s' unrecognised — skipped.", _src.suffix)
        except Exception as exc:  # noqa: BLE001
            logger.warning("[SheetParser] PAYLOAD_PATH auto-copy failed: %s", exc)


    return result
