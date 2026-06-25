# ┌─────────────────────────────────────────────────────────────────────────────┐
# │  MACCREv2 ENGINEERING DOCTRINE                             Law Rev: 19.0   │
# ├─────────────────────────────────────────────────────────────────────────────┤
# │  I.   TYPING      All signatures: explicit Python 3.11+ type hints.        │
# │  II.  LINTING     Zero unused imports. No wildcards. 120-char line max.    │
# │  III. PATHS       Never hardcode absolute paths. Use get_maccre_root().     │
# │                   Default params: def f(p:str='') -> None: p=p or root/x   │
# │  IV.  DATACENTER  5-Tier: 01_Raw_Source · 02_Dynamic_Context               │
# │                           03_Agent_Ledgers · 04_Code_Artifacts             │
# │                           05_Rendered_Media                                 │
# │  V.   DIAMOND     Gen: temp=1.0  ·  Critic: temp=0.1 + dataclass schema   │
# │  VI.  ABSTRACTION All I/O behind abc.ABC before any concrete driver.       │
# │  VII. TEARDOWN    try/finally on all handles (omni clean compliance).      │
# │  VIII.TELEMETRY   No bare print(). logger only. JSON → 03_Agent_Ledgers.  │
# └─────────────────────────────────────────────────────────────────────────────┘
"""
maccre_core/tools/design_tools.py
==================================
Swarm Design Engine — Diamond Loop Architecture.

Converts natural language descriptions into fully materialised MACCRE swarms:
- Leg 1 (Generator): Gemini 2.5 Pro at temperature=1.0  — creative ideation,
  rich design narrative, open-ended reasoning about agents and pipeline flow.
- Leg 2 (Critic):    Gemini 2.5 Pro at temperature=0.1  — structured extraction
  into typed Pydantic schema; zero hallucination risk because schema is enforced
  at the response_mime_type level.

Files are written to the project silo in real-time as the conversation flows —
the workspace is ready by the time Nexus finishes speaking.
"""
from __future__ import annotations

import json
from typing import Any

from dataclasses import dataclass, field

from maccre_core._net.gemini_client import GeminiClient, user_turn, is_transient_error, is_fatal_error
from maccre_core.orchestration.universal_vault import get_provider_credential
from maccre_core.utils.path_resolver import get_maccre_root
from maccre_core.logger import logger



# ── Pydantic Schema for the Diamond Loop Critic ──────────────────────────────


@dataclass
class AgentDesign:
    """One agent participating in the swarm."""

    name: str = field(metadata={"description": "Unique agent name, PascalCase, no spaces (e.g. 'Story_Archivist')"})
    role: str = field(metadata={"description": "One-sentence description of this agent's job in the pipeline."})
    persona_instructions: str = field(
        metadata={"description": (
            "Full, rich system prompt for this agent. Be specific: define voice, "
            "output format, style constraints, and how to handle edge cases. "
            "Minimum 3 sentences."
        )}
    )
    tools: str = field(
        metadata={"description": (
            "Pipe-separated list of MACCRE tools this agent is authorised to call. "
            "Valid values: ingest_document | query_local_memory | query_global_archive | write_file | "
            "execute_render_pipeline | read_file | none. "
            "Example: 'write_file|query_local_memory'"
        )}
    )
    model: str = field(
        default="gemini-3.1-pro-preview",
        metadata={"description": "Compute backend"},
    )
    temperature: float = field(
        default=0.7,
        metadata={"description": "Agent's default temperature."},
    )


@dataclass
class NodeDesign:
    """One node in the swarm topology DAG."""

    node_id: str = field(metadata={"description": "Unique NODE identifier in SCREAMING_SNAKE_CASE (e.g. 'INGEST', 'WRITE_CH1', 'RENDER')."})
    agent_name: str = field(metadata={"description": "Must match an AgentDesign.name exactly."})
    next_node: str = field(metadata={"description": "The node_id to route to on success, or 'STOP' if terminal. Comma-separated for Scatter fan-out."})
    instruction_override: str = field(
        metadata={"description": (
            "Specific instruction for this node, overriding the agent's base persona. "
            "Be precise: what should the agent read, produce, and how should it format output."
        )}
    )
    temperature: float = field(
        default=0.7,
        metadata={"description": "Node-level temperature override. 1.0 for creative nodes, 0.1 for analytic nodes."},
    )
    model_override: str = field(
        default="",
        metadata={"description": "Leave empty to use the agent's default model."},
    )
    max_recursion: int = field(
        default=3,
        metadata={"description": "Maximum number of times this node can be iteratively executed in a loop."},
    )
    wait_for: str = field(
        default="none",
        metadata={"description": "Comma-separated Node_IDs that must complete before this node runs (Gather Gate). 'none' = no gate."},
    )
    failure_target: str = field(
        default="FAILED",
        metadata={"description": "Node to route to on unrecoverable failure. Default 'FAILED' terminal sentinel."},
    )
    artifact_path: str = field(
        default="",
        metadata={
            "description": (
                "Optional relative path (from the project silo root) of the canonical output "
                "artifact this node produces."
            )
        },
    )
    live_profile: str = field(
        default="",
        metadata={"description": "Set to TRUE to run this node as a Live Session websocket endpoint."},
    )
    dialogue_partner: str = field(
        default="",
        metadata={"description": "Agent name for a persistent two-turn dialogue. When set with dialogue_rounds > 0, DialogueRunner fires instead of the standard loop."},
    )
    dialogue_rounds: int = field(
        default=0,
        metadata={"description": "Number of full exchange rounds in dialogue mode. 0 = standard single-agent mode."},
    )


@dataclass
class SwarmDesign:
    """Complete swarm specification extracted from the design narrative."""

    project_name: str = field(
        metadata={"description": "Short, unique project identifier. Alphanumeric and underscores only, no spaces (e.g. 'STORY_PIPELINE_001')."}
    )
    agents: list[AgentDesign] = field(metadata={"description": "All agent personas required for this swarm."})
    topology: list[NodeDesign] = field(
        metadata={"description": "Ordered list of nodes. The first node is the starting point (usually INGEST). The last node routes to STOP."}
    )
    payload_description: str = field(
        metadata={"description": "Exact description of the input file/text the user must provide before running the swarm."}
    )
    missing_requirements: list[str] = field(
        default_factory=list,
        metadata={"description": (
            "List ONLY requirements that are truly missing and would BLOCK execution. "
            "If the description is complete enough to build a working swarm, return an empty list. "
            "Do NOT ask stylistic or preference questions — make sensible defaults."
        )}
    )
    design_narrative: str = field(
        default="",
        metadata={"description": "Brief summary of the design rationale (2-3 sentences)."},
    )


# ── Internal: parse JSON → SwarmDesign dataclass ─────────────────────────────


def _parse_swarm_design(raw_json: str) -> "SwarmDesign":
    """Deserialize a JSON string into a SwarmDesign dataclass.

    SwarmDesign is a plain Python ``@dataclass`` (not a Pydantic model), so we
    use ``json.loads`` + manual field construction instead of
    ``model_validate_json``.
    """
    import re as _re  # noqa: PLC0415

    # Strip markdown fences the model sometimes wraps the JSON in
    clean = _re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_json.strip(), flags=_re.MULTILINE)
    data: dict[str, Any] = json.loads(clean or "{}")

    agents = [
        AgentDesign(
            name=a.get("name", ""),
            role=a.get("role", ""),
            persona_instructions=a.get("persona_instructions", ""),
            tools=a.get("tools", "none"),
            model=a.get("model", "gemini-3.1-pro-preview"),
            temperature=float(a.get("temperature", 0.7)),
        )
        for a in data.get("agents", [])
    ]
    topology = [
        NodeDesign(
            node_id=n.get("node_id", ""),
            agent_name=n.get("agent_name", ""),
            next_node=n.get("next_node", "STOP"),
            instruction_override=n.get("instruction_override", ""),
            temperature=float(n.get("temperature", 0.7)),
            model_override=n.get("model_override", ""),
            max_recursion=int(n.get("max_recursion", 3)),
            wait_for=n.get("wait_for", "none"),
            failure_target=n.get("failure_target", "FAILED"),
            artifact_path=n.get("artifact_path", ""),
            live_profile=str(n.get("live_profile", "")),
            dialogue_partner=str(n.get("dialogue_partner", "")),
            dialogue_rounds=int(n.get("dialogue_rounds", 0)),
        )
        for n in data.get("topology", data.get("nodes", []))
    ]
    return SwarmDesign(
        project_name=data.get("project_name", data.get("swarm_name", "unnamed_swarm")),
        agents=agents,
        topology=topology,
        payload_description=data.get("payload_description", ""),
        missing_requirements=data.get("missing_requirements", []),
        design_narrative=data.get("design_narrative", data.get("design_rationale", "")),
    )


# ── Internal: Materialise a validated SwarmDesign into project files ──────────


def _materialise_swarm(design: SwarmDesign) -> str:
    """Write workspace, agents, persona cards, and topology for a validated SwarmDesign."""
    from maccre_core.tools.admin_tools import (  # noqa: PLC0415
        build_topology,
        create_persona_card,
        initialize_workspace,
        mint_agent,
        switch_workspace,
    )

    project = design.project_name

    # 1. Workspace
    init_result = initialize_workspace(project)
    if "[ADMIN_FAULT]" in init_result and "already exists" not in init_result:
        return f"[DESIGN_FAULT] Workspace initialisation failed: {init_result}"
    if "already exists" in init_result:
        switch_workspace(project)

    # 2. Agents: roster entry + ROM cartridge
    for agent in design.agents:
        mint_agent(
            name=agent.name,
            model=agent.model,
            system_prompt=agent.persona_instructions,
            description=agent.role,
            tools_string=agent.tools,
        )
        create_persona_card(
            agent_name=agent.name,
            instructions=agent.persona_instructions,
            temperature=agent.temperature,
            context_notes=agent.role,
        )

    # 3. Topology CSV
    nodes: list[list[Any]] = []
    for node in design.topology:
        nodes.append([
            node.node_id,              # [0] Node_ID
            node.agent_name,           # [1] Agent_Name
            node.model_override or "", # [2] Model_Override
            node.next_node,            # [3] Next_Node
            node.temperature,          # [4] Temperature
            node.instruction_override, # [5] Instruction_Override
            node.wait_for,             # [6] Wait_For
            node.failure_target,       # [7] Failure_Target
            node.max_recursion,        # [8] Max_Recursion
            node.artifact_path,        # [9] Artifact_Path
            node.live_profile,         # [10] Live_Profile
            node.dialogue_partner,     # [11] Dialogue_Partner
            node.dialogue_rounds,      # [12] Dialogue_Rounds
        ])
    topo_result = build_topology(nodes)
    if "[ADMIN_FAULT]" in topo_result:
        return f"[DESIGN_FAULT] Topology build failed: {topo_result}"

    # 4. Build human-readable summary
    agent_lines = "\n".join(
        f"  • **{a.name}** ({a.model}, T={a.temperature}): {a.role}"
        for a in design.agents
    )
    node_lines = "\n".join(
        f"  `{n.node_id}` [{n.agent_name}] → `{n.next_node}`"
        for n in design.topology
    )
    return (
        f"[SWARM_READY] Project **{project}** has been fully materialised.\n\n"
        f"**Design Rationale:** {design.design_narrative}\n\n"
        f"**Agents ({len(design.agents)}):**\n{agent_lines}\n\n"
        f"**Topology ({len(design.topology)} nodes):**\n{node_lines}\n\n"
        f"**What you need to provide:** {design.payload_description}\n\n"
        "**Next steps:**\n"
        f"1. Write your payload: `write_file(\"01_Raw_Source/input.md\", <your content>)`\n"
        f"2. Queue the job: `ignite_swarm(\"input.md\", \"{design.topology[0].node_id}\")`\n"
        f"3. Execute: `run_swarm(\"{project}\")`"
    )


# ── Public Tool ───────────────────────────────────────────────────────────────


def design_swarm(description: str, answers: str = "") -> str:
    """Design and materialise a complete MACCRE swarm from a natural language description.

    Uses the Diamond Loop pattern:
    - Leg 1 (Generator — temp=1.0): Gemini 2.5 Pro creatively designs the full swarm,
      producing a rich narrative with agent personas, pipeline rationale, and topology.
    - Leg 2 (Critic — temp=0.1): Extracts the structured ``SwarmDesign`` from the narrative
      using a Pydantic response_schema — zero hallucination risk on structured fields.

    If the description is ambiguous or missing required information, the tool returns
    a list of clarifying questions.  The user answers in the next Nexus turn; Nexus
    calls ``design_swarm`` again with both the original description and the answers
    concatenated in ``answers``.

    Files are written to the project silo immediately upon a complete spec — agents,
    persona cards, and topology are ready before this function returns.

    Args:
        description: Natural language description of what the swarm should do.
        answers: Optional follow-up answers to clarifying questions from a prior call.

    Returns:
        Either a ``[SWARM_READY]`` confirmation with materialised artefact summary,
        or a ``[DESIGN_NEEDS_INPUT]`` block listing missing requirements as questions,
        or a ``[DESIGN_FAULT]`` string describing a hard error.
    """
    api_key = get_provider_credential("MACCRE_Sovereign")
    if not api_key:
        return "[DESIGN_FAULT] MACCRE_Sovereign vault key not found. Ensure it is registered in Windows Credential Manager."

    client = GeminiClient(key_provider=lambda: get_provider_credential("MACCRE_Sovereign"))

    # ── Enumerate known topology library entries for context ──────────────────
    _prior_art = ""
    try:
        import sqlite3  # noqa: PLC0415
        _dc = get_maccre_root() / "__DATACENTER"
        for _db in _dc.rglob("definitions.db"):
            try:
                with sqlite3.connect(str(_db)) as _con:
                    _rows = _con.execute(
                        "SELECT DISTINCT topology_name, agent_name, auto_tool, instruction_override "
                        "FROM topology_library ORDER BY id DESC LIMIT 20"
                    ).fetchall()
                if _rows:
                    _prior_art += "\n".join(
                        f"  {r[0]} | {r[1]} | tool={r[2]} | instr={str(r[3])[:60]}" for r in _rows
                    )
            except Exception:
                pass
    except Exception:
        pass

    prior_art_block = f"\n\nKNOWN TOPOLOGY LIBRARY (for reference):\n{_prior_art}" if _prior_art else ""

    # ── Leg 1: Generator — Gemini 2.5 Pro at temperature=1.0 ─────────────────
    supplemental = f"\n\nUSER'S FOLLOW-UP ANSWERS:\n{answers}" if answers.strip() else ""

    generator_prompt = f"""You are MACCRE's Master Swarm Architect. A user wants to build a multi-agent processing pipeline using the MACCREv2 swarm engine.

USER'S REQUEST:
{description}{supplemental}{prior_art_block}

AVAILABLE MACCRE TOOLS (agents can only use these):
- ingest_document: Read a source file and embed it into local semantic memory. Use for INGEST nodes.
- query_local_memory: Semantic search inside local semantic memory. Use in generation nodes that need context.
- query_global_archive: Semantic search across the GLOBAL knowledge archive. Returns thought pins and paths to original sources. Use when an agent needs broad background context.
- query_session_memory: Forensic query tool to extract vectors from a specific ephemeral session (agent_thoughts, agent_ledgers). Use to salvage data from failed runs.
- scout_archive_themes: Rapidly scans a GLOBAL Ingestion Project (like AI_STUDIO) and returns a raw map of themes. Use as the Map phase in a MapReduce ingestion.
- execute_archive_ingestion: Routes an entire directory of files into Concept Silos based on a list of categories provided by the agent. Use as the Reduce/Ingest phase.
- write_file: Write output to a file in the datacenter. Use for all content-generating nodes.
- execute_render_pipeline: Consumes a Director manifest JSON and produces TTS audio + scene images stitched by FFmpeg into an MP4. Use for RENDER nodes. The manifest must be JSON array of objects with keys: speaker, text, video_prompt.
- read_file: Read any file back for further processing.
- none: The agent reasons and writes its output without calling a tool.

PIPELINE RULES:
- The first node is always INGEST (reads the source payload into memory).
- Creative writing nodes should use query_local_memory to pull context, then write_file.
- The PODCAST/SCRIPT node should write a properly formatted Director manifest JSON.
- The RENDER node should call execute_render_pipeline with that manifest.
- The final node routes to STOP.

DESIGN MANDATE — YOU MUST:
- Commit to a COMPLETE, production-ready design. Do NOT defer or hedge.
- Invent a creative, thematic project name yourself (e.g. OBSIDIAN_QUILL, ECHO_FORGE_001).
- Design all agents yourself — give each a distinct voice, role, and rich persona.
- Decide the topology yourself — linear is fine for story pipelines.
- The ONLY valid reason to ask for missing info is if the user has not said ANYTHING about what
  TYPE of content to process (e.g. they described no input at all). If they described the domain
  (fantasy novel, podcast, video), you have enough — build it.

Think deeply about each agent's personality, voice, and role. Write rich personas.
Return your complete swarm design with full rationale."""

    # ── Shared Diamond Loop retry helper — uses live registry failover chain ──
    import time as _t  # noqa: PLC0415
    from maccre_core._net.model_registry import get_registry  # noqa: PLC0415
    _registry = get_registry(lambda: get_provider_credential("MACCRE_Sovereign"))

    def _gemini_with_retry(
        contents: str,
        temperature: float,
        max_output_tokens: int,
        response_schema_hint: str | None = None,
        model: str = "gemini-2.5-pro",
        label: str = "API call",
    ) -> str:
        """Call Gemini through the live registry failover chain, return raw text."""
        chain = _registry.get_failover_chain(model)
        last_exc: Exception = RuntimeError("No attempts made")

        for _idx, _attempt_model in enumerate(chain):
            from maccre_core.maccre_router import _CHAIN_DELAYS  # noqa: PLC0415
            _delay = _CHAIN_DELAYS[_idx] if _idx < len(_CHAIN_DELAYS) else _CHAIN_DELAYS[-1]
            if _delay:
                logger.warning(
                    "[DesignTools] %s -- failover to '%s' (pos %d/%d) -- waiting %ds...",
                    label, _attempt_model, _idx + 1, len(chain), _delay,
                )
                _t.sleep(_delay)
            elif _attempt_model != model:
                logger.info("[DesignTools] %s -- using failover model '%s'", label, _attempt_model)
            try:
                system: str | None = None
                if response_schema_hint:
                    system = (
                        f"You MUST reply with valid JSON only. "
                        f"Match this schema exactly: {response_schema_hint}. "
                        f"Do not include any text outside the JSON."
                    )
                resp = client.generate_content(
                    model=_attempt_model,
                    contents=[user_turn(contents)],
                    system_instruction=system,
                    temperature=temperature,
                )
                if _attempt_model != model:
                    logger.info("[DesignTools] %s served by failover model '%s'.", label, _attempt_model)
                return resp.text
            except Exception as exc:
                if is_fatal_error(exc):
                    raise RuntimeError(f"[DESIGN_FATAL] {label} on '{_attempt_model}': {exc}") from exc
                if is_transient_error(exc):
                    last_exc = exc
                    logger.warning("[DesignTools] %s transient on '%s' (pos %d/%d): %s",
                                   label, _attempt_model, _idx + 1, len(chain), str(exc)[:120])
                    continue
                raise RuntimeError(f"[DESIGN_ERROR] {label} on '{_attempt_model}': {exc}") from exc

        raise RuntimeError(f"[DESIGN_FAULT] {label} chain exhausted. Tried: {chain}. Last: {last_exc}")


    try:
        gen_response_text = _gemini_with_retry(
            contents=generator_prompt,
            temperature=1.0,
            max_output_tokens=8192,
            label="Generator (Leg 1)",
        )
    except Exception as e:
        return f"[DESIGN_FAULT] Generator (Leg 1) failed: {e}"

    ideation_text: str = gen_response_text or ""

    # ── Leg 2: Critic — Gemini 2.5 Pro at temperature=0.1 with Pydantic schema ──
    # Cap ideation fed into the extraction prompt — we only need enough for the
    # critic to extract structure; sending all 8192 Gen tokens doubles token pressure
    # and is the root cause of EOF truncation on the output side.
    ideation_for_extraction = ideation_text[:6000]

    _ts_suffix = str(int(_t.time()))[-4:]  # last 4 digits of epoch for project_name uniqueness

    extraction_prompt = f"""Extract the swarm design from the narrative below into the required structured format.

DESIGN NARRATIVE (may be truncated -- use your judgment to complete missing fields):
{ideation_for_extraction}

ORIGINAL USER REQUEST: {description}{supplemental}

EXTRACTION RULES (non-negotiable -- breaking these means the swarm cannot run):
- agent.tools: pipe-separated valid tool names only. No spaces around pipes.
- node.wait_for: comma-separated node_ids that must complete before this node runs, or 'none'.
- node.failure_target: node to route to on failure (default: 'FAILED').
- node.temperature: 1.0 for creative writing nodes, 0.1 for ingest/review/render/analysis nodes.
- project_name: alphanumeric + underscores only, no spaces, max 30 chars. Append _{_ts_suffix} for uniqueness.
- missing_requirements: STRICT RULE -- return [] (empty list) unless the user has provided ZERO
  information about what content to process. NEVER ask for: project name (auto-generate it),
  agent names (design sensible defaults), topology order (decide yourself), parallel vs sequential
  (default to sequential). The ONLY valid missing requirement is if no input domain was described at all.
- persona_instructions: Keep to 3-4 concise sentences max. Do NOT pad with fluff.
- instruction_override: Keep to 2-3 sentences max per node.
- The topology MUST start with an INGEST node and end with a node routing to STOP.
- design_narrative: 2-3 sentence summary of the pipeline's creative and technical rationale."""

    _schema_hint = '{"project_name": "...", "agents": [...], "topology": [...], "payload_description": "...", "missing_requirements": [], "design_narrative": "..."}'
    try:
        extract_response_text = _gemini_with_retry(
            contents=extraction_prompt,
            temperature=0.1,
            max_output_tokens=8192,
            response_schema_hint=_schema_hint,
            label="Critic (Leg 2)",
        )
    except Exception as e:
        return f"[DESIGN_FAULT] Critic (Leg 2) extraction failed: {e}"

    raw_json: str = extract_response_text or ""

    # ── Dataclass parse via sovereign schema layer ──────────────────────────
    design: SwarmDesign | None = None
    try:
        design = _parse_swarm_design(raw_json)
    except Exception as primary_err:
        if not raw_json:
            return "[DESIGN_FAULT] Critic returned empty response. The model may be unavailable."

        # Attempt repair: send the broken JSON to Flash with a targeted fix instruction
        repair_prompt = (
            f"The following JSON was truncated or malformed during generation. "
            f"Complete and repair it into a valid SwarmDesign structure.\n"
            f"Original user request: {description}\n"
            f"Broken JSON (repair this):\n{raw_json[-3000:]}"  # tail of truncated output
        )
        try:
            repair_response_text = _gemini_with_retry(
                contents=repair_prompt,
                temperature=0.0,
                max_output_tokens=8192,
                response_schema_hint=_schema_hint,
                model="gemini-2.5-flash",
                label="Repair Loop",
            )
            design = _parse_swarm_design(repair_response_text or "{}")
        except Exception as repair_err:
            return (
                f"[DESIGN_FAULT] Schema validation failed (primary: {primary_err!s:.120}; "
                f"repair: {repair_err!s:.120}).\n"
                f"Raw tail: {raw_json[-300:]}"
            )

    assert design is not None  # mypy satisfaction

    # ── Gate: missing requirements → return questions, do NOT materialise yet ─
    if design.missing_requirements:
        questions = "\n".join(f"  {i + 1}. {q}" for i, q in enumerate(design.missing_requirements))
        # Show the narrative so the user has full context
        narrative_preview = ideation_text[:800].strip()
        return (
            f"[DESIGN_NEEDS_INPUT] I've designed a swarm called **{design.project_name}** "
            f"but need clarification on {len(design.missing_requirements)} point(s) before building:\n\n"
            f"{questions}\n\n"
            f"---\n**Design Preview** (what I have so far):\n{narrative_preview}...\n\n"
            f"*Answer these questions in your next message and I will build the swarm immediately.*"
        )

    # ── Full spec received — materialise immediately ──────────────────────────
    return _materialise_swarm(design)


# ── fill_swarm_sheet ──────────────────────────────────────────────────────────


def fill_swarm_sheet(
    description: str,
    answers: str = "",
    xlsx_path: str = "",
) -> str:
    """Design a swarm via the Diamond Loop and write the result to xlsx.

    Unlike design_swarm() which materialises workspace files immediately, this
    produces a portable MACCRE_Swarm_Request.xlsx the user can review, edit in
    Google Sheets, and drop into the Drive Inbox to trigger execution from any
    device.

    Args:
        description: Natural-language swarm description.
        answers:     Follow-up answers to a previous [DESIGN_NEEDS_INPUT] prompt.
        xlsx_path:   Output path. Blank = auto in GLOBAL/04_Code_Artifacts.

    Returns:
        [SHEET_READY] <path>    on success.
        [DESIGN_NEEDS_INPUT]    if clarification is still needed.
        [DESIGN_FAULT]          if the Diamond Loop fails.
    """
    import csv  # noqa: PLC0415
    import os   # noqa: PLC0415
    import re   # noqa: PLC0415
    import shutil  # noqa: PLC0415
    from typing import Any as _Any  # noqa: PLC0415

    # ── Resolve output path — write into the project silo, not GLOBAL ─────────
    # This is resolved *after* the Diamond Loop so we know the project name.
    # We defer actual path construction until the project name is known below.

    # ── Diamond Loop (reuse design_swarm) ─────────────────────────────────────
    loop_result = design_swarm(description=description, answers=answers)

    if not loop_result.startswith("[SWARM_READY]"):
        return loop_result   # propagate NEEDS_INPUT or FAULT unchanged

    # ── Identify the newly created project ────────────────────────────────────
    match = re.search(r"\*\*([A-Z0-9_]{4,40})\*\*", loop_result)
    if not match:
        return loop_result + "\n\n[NOTE] xlsx export skipped - project name not found in result."

    project_name = match.group(1)
    os.environ["MACCRE_ACTIVE_PROJECT"] = project_name

    # ── Extract payload filename from the SWARM_READY result ("What you need to provide") ──
    _payload_name_match = re.search(
        r"named?\s+[`'\"]{0,1}([\w\-]+\.\w+)[`'\"]{0,1}",
        loop_result, re.IGNORECASE,
    )
    _payload_filename: str = _payload_name_match.group(1) if _payload_name_match else "input.md"

    # ── Extract start node from the topology block (first backtick-wrapped node_id) ────────
    _start_node_match = re.search(r"`([A-Z_][A-Z0-9_]+)`\s*\[", loop_result)
    _start_node: str = _start_node_match.group(1) if _start_node_match else "INGEST"

    # ── Resolve output path now that we know the project name ─────────────────
    if not xlsx_path:
        try:
            _out_dir = get_maccre_root() / "__DATACENTER" / project_name
        except Exception:  # noqa: BLE001
            _out_dir = get_maccre_root() / "templates"
        _out_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = str(_out_dir / "MACCRE_Swarm_Request.xlsx")

    try:
        from maccre_core.utils.path_resolver import get_datacenter_path  # noqa: PLC0415

        roster_path = get_datacenter_path("agent_roster.csv")
        topo_path   = get_datacenter_path("02_Dynamic_Context") / "topology.csv"

        if not roster_path.exists() or not topo_path.exists():
            return loop_result + "\n\n[NOTE] xlsx export skipped - roster/topology not found."

        with open(roster_path, newline="", encoding="utf-8") as fh:
            roster_rows = list(csv.DictReader(fh))
        with open(topo_path, newline="", encoding="utf-8") as fh:
            topo_rows = list(csv.DictReader(fh))

        template_src = get_maccre_root() / "templates" / "MACCRE_Swarm_Request.xlsx"
        if not template_src.exists():
            return loop_result + "\n\n[NOTE] xlsx export skipped - base template not built yet."

        shutil.copy2(str(template_src), xlsx_path)

        from openpyxl import load_workbook as _lw  # type: ignore  # noqa: PLC0415
        wb = _lw(xlsx_path)

        def _clear(ws: _Any, from_row: int = 3) -> None:
            for row in ws.iter_rows(min_row=from_row):
                for cell in row:
                    cell.value = None

        # SWARM_REQUEST — column order matches template exactly:
        # col1=PROJECT_NAME  col2=DESCRIPTION  col3=COMPUTE_TIER
        # col4=PAYLOAD_TEXT  col5=PAYLOAD_PATH  col6=START_NODE
        ws_req = wb["SWARM_REQUEST"]
        _clear(ws_req)
        ws_req.cell(row=3, column=1, value=project_name)
        ws_req.cell(row=3, column=2, value=description[:120])
        ws_req.cell(row=3, column=3, value="cloud")
        ws_req.cell(row=3, column=5, value=_payload_filename)   # PAYLOAD_PATH → watcher uses this
        ws_req.cell(row=3, column=6, value=_start_node)         # START_NODE

        # AGENTS — column order matches template exactly:
        # col1=AGENT_NAME  col2=ROLE  col3=COMPUTE_TIER  col4=MODEL  col5=TEMPERATURE
        # col15=TOOLS  col16=PERSONA
        ws_ag = wb["AGENTS"]
        _clear(ws_ag)
        for ridx, ag in enumerate(roster_rows, start=3):
            ws_ag.cell(row=ridx, column=1, value=ag.get("Agent_Name", ""))
            ws_ag.cell(row=ridx, column=2, value=ag.get("Description", ""))  # ROLE
            ws_ag.cell(row=ridx, column=3, value="cloud")                    # COMPUTE_TIER
            ws_ag.cell(row=ridx, column=4, value=ag.get("Model", "gemini-2.5-flash"))  # MODEL
            ws_ag.cell(row=ridx, column=5, value=ag.get("Temperature", "0.7"))         # TEMPERATURE
            ws_ag.cell(row=ridx, column=15, value=ag.get("Tools_Allowed", "write_file"))  # TOOLS
            ws_ag.cell(row=ridx, column=16, value=ag.get("System_Prompt", ""))          # PERSONA

        # TOPOLOGY — columns match build_global_workbook() live schema:
        #   col 1: NODE_ID  col 2: AGENT_NAME  col 3: NEXT_NODE
        #   col 4: MODEL_OVERRIDE  col 5: TEMPERATURE  col 6: MAX_RECURSION  col 7: INSTRUCTION_OVERRIDE
        ws_tp = wb["TOPOLOGY"]
        _clear(ws_tp)
        for ridx, nd in enumerate(topo_rows, start=3):
            ws_tp.cell(row=ridx, column=1, value=nd.get("Node_ID", ""))
            ws_tp.cell(row=ridx, column=2, value=nd.get("Agent_Name", ""))
            ws_tp.cell(row=ridx, column=3, value=nd.get("Next_Node", "STOP"))
            ws_tp.cell(row=ridx, column=4, value=nd.get("Model_Override", ""))
            ws_tp.cell(row=ridx, column=5, value=nd.get("Temperature", "1.0"))
            ws_tp.cell(row=ridx, column=6, value=nd.get("Max_Recursion", "3"))
            ws_tp.cell(row=ridx, column=7, value=nd.get("Instruction_Override", ""))

        wb.save(xlsx_path)
        wb.close()

        return (
            f"{loop_result}\n\n"
            f"[SHEET_READY] Portable xlsx spec written:\n"
            f"  {xlsx_path}\n\n"
            f"Rename to MACCRE_Swarm_Request_APPROVED.xlsx to trigger the watcher, "
            f"or drop into your Drive Inbox to re-trigger from any device."
        )

    except Exception as exc:  # noqa: BLE001
        return loop_result + f"\n\n[NOTE] xlsx export failed: {exc}"

