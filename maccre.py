# ┌─────────────────────────────────────────────────────────────────────────────┐
# │  MACCREv2 ENGINEERING DOCTRINE                             Law Rev: 19.0   │
# ├─────────────────────────────────────────────────────────────────────────────┤
# │  I.   TYPING      All signatures: explicit Python 3.11+ type hints.        │
# │  II.  LINTING     Zero unused imports. No wildcards. 120-char line max.    │
# │  III. PATHS       Never hardcode absolute paths. Use get_maccre_root().     │
# │                   Default params: def f(p:str='') -> None: p=p or root/x   │
# │  IV.  DATACENTER  5-Tier: 01_Raw_Source · 02_Dynamic_Context               │
# │                           03_Agent_Ledgers · 04_Code_Artifacts             │
# │                           05_Rendered_Media                                 │
# │  V.   DIAMOND     Gen: temp=1.0  ·  Critic: temp=0.1 + dataclass schema   │
# │  VI.  ABSTRACTION All I/O behind abc.ABC before any concrete driver.       │
# │  VII. TEARDOWN    try/finally on all handles (omni clean compliance).      │
# │  VIII.TELEMETRY   No bare print(). logger only. JSON → 03_Agent_Ledgers.  │
# └─────────────────────────────────────────────────────────────────────────────┘
"""
maccre.py
==========
MACCREv2 Headless CLI — The Sovereign Master Interface.
Pure, zero-dependency entry point for the Multi-Agent Conversational
Concept Refinement Engine.

Usage:
  python maccre.py ignite <payload_path> [--node <NODE>]
  python maccre.py status
  python maccre.py canonize --project <PROJECT_ID> --session <SESSION_ID>
  python maccre.py chat [--agent <NAME>] [--model <MODEL_ID>]
"""

import argparse
import os
import platform
import shutil
import sqlite3
import subprocess
import sys

from pathlib import Path

# ── C-Level Stdout Guard ──────────────────────────────────────────────────────
# Redirect C-level fd 1 (stdout) to fd 2 (stderr) so that rogue C-extension
# writes (FFmpeg, ChromaDB telemetry, etc.) never contaminate the JSON pipeline.
# System encoding is left natively handled to prevent execution halt.
import io as _io
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
# ─────────────────────────────────────────────────────────────────────────────

from maccre_core.logger import logger  # noqa: E402
from maccre_core.orchestration.local_broker import LocalMessageBroker  # noqa: E402
from maccre_core.tools.rag_tools import merge_session_to_project  # noqa: E402
from maccre_core.tools.telemetry_tools import export_and_purge_thoughts  # noqa: E402
from maccre_core.utils.path_resolver import get_datacenter_path  # noqa: E402

_DB_PATH = str(get_datacenter_path("swarm_queue.db"))


# ── Pre-Flight Smoke Gate ─────────────────────────────────────────────────────

def _run_preflight(skip: bool = False) -> None:
    """Run the zero-cost smoke test as a hard gate before any swarm launch.

    Bypassed by ``skip=True`` (--skip-preflight flag) or the
    ``MACCRE_SKIP_PREFLIGHT`` environment variable.  A failed smoke test
    calls ``sys.exit(1)`` so no downstream code runs.
    """
    if skip or os.environ.get("MACCRE_SKIP_PREFLIGHT", ""):
        print("[PRE-FLIGHT] Bypassed (--skip-preflight / MACCRE_SKIP_PREFLIGHT).")
        return
    print("[PRE-FLIGHT] Running smoke test... (--skip-preflight to bypass)")
    from maccre_core.tests.smoke_test import run_smoke_test  # noqa: PLC0415
    ok = run_smoke_test()
    if not ok:
        print("[PRE-FLIGHT] \u2717 FAILED \u2014 fix errors above before launching.")
        sys.exit(1)
    print("[PRE-FLIGHT] \u2713 Passed.\n")


# ── OS-Agnostic Terminal Multiplexer ─────────────────────────────────────────

def spawn_pane(cmd_args: list[str], title: str = "MACCREv2 Sub-Pane") -> None:
    """Spawn a detached native terminal window running the given command.

    Detection priority:
        1. Termux / Samsung DeX (Android + TMUX)
        2. Windows (cmd /k)
        3. macOS (osascript Terminal)
        4. Linux desktop (gnome-terminal → konsole → xterm)

    Args:
        cmd_args: The command and arguments to run in the new pane.
        title:    Window/tab title for identification.
    """
    system = platform.system().lower()

    # 1. Termux / Samsung DeX detection
    if "com.termux" in os.environ.get("PREFIX", ""):
        if "TMUX" in os.environ:
            subprocess.Popen(["tmux", "split-window", "-v", *cmd_args])
            logger.info("[MULTIPLEXER] tmux split-window spawned: %s", cmd_args)
        else:
            print("[!] Termux detected outside TMUX. Run `tmux` first for multiplexing.")
        return

    # 2. Windows native
    if system == "windows":
        subprocess.Popen(
            ["start", "cmd", "/k", "title", title, "&", *cmd_args],
            shell=True,
        )
        logger.info("[MULTIPLEXER] Windows cmd pane spawned: %s", cmd_args)
        return

    # 3. macOS native
    if system == "darwin":
        escaped = " ".join(cmd_args).replace('"', '\\"')
        script = f'tell application "Terminal" to do script "{escaped}"'
        subprocess.Popen(["osascript", "-e", script])
        logger.info("[MULTIPLEXER] macOS Terminal pane spawned: %s", cmd_args)
        return

    # 4. Linux desktop (X11/Wayland)
    if system == "linux":
        terminal_candidates: list[list[str]] = [
            ["gnome-terminal", "--"],
            ["konsole", "-e"],
            ["xterm", "-e"],
        ]
        for term in terminal_candidates:
            if shutil.which(term[0]):
                subprocess.Popen([*term, *cmd_args])
                logger.info("[MULTIPLEXER] Linux terminal pane spawned via %s: %s", term[0], cmd_args)
                return
        print("[!] No supported Linux terminal emulator found in PATH.")
        logger.warning("[MULTIPLEXER] No Linux terminal emulator found. Candidates: %s", terminal_candidates)


# ── Interactive Master CLI Shell ──────────────────────────────────────────────


# ── Core CLI Commands ─────────────────────────────────────────────────────────

def _log_launch_failure(job_id: str, payload_path: str, starting_node: str, error: str) -> None:
    """Persist a launch failure record to 03_Agent_Ledgers for post-mortem review."""
    try:
        from maccre_core.utils.path_resolver import get_datacenter_path  # noqa: PLC0415
        ledger_dir = get_datacenter_path("03_Agent_Ledgers")
        ledger_dir.mkdir(parents=True, exist_ok=True)
        record = {
            "timestamp": __import__("datetime").datetime.utcnow().isoformat() + "Z",
            "event": "LAUNCH_FAILURE",
            "job_id": job_id,
            "payload_path": payload_path,
            "starting_node": starting_node,
            "error": error,
        }
        with open(ledger_dir / "launch_failures.jsonl", "a", encoding="utf-8") as fh:
            fh.write(__import__("json").dumps(record) + "\n")
    except Exception:  # noqa: BLE001
        pass  # Never let failure logging crash the process


def _launch_watcher(job_id: str, project: str) -> None:
    """Open the SwarmWatcher in a new console window (Windows only)."""
    if sys.platform != "win32":
        return
    try:
        from maccre_core.utils.path_resolver import get_maccre_root as _gr  # noqa: PLC0415
        watcher = _gr() / "maccre_core" / "tools" / "swarm_watcher.py"
        if not watcher.exists():
            logger.warning("[WATCHER] swarm_watcher.py not found — skipping watcher launch")
            return
        title = f"MACCRE Watcher [{job_id}]"
        cmd = (
            f'start "{title}" {sys.executable} '
            f'"{watcher}" --job-id {job_id} --project {project}'
        )
        import subprocess  # noqa: PLC0415
        subprocess.Popen(cmd, shell=True, cwd=str(watcher.parent.parent.parent))
        logger.info("[WATCHER] Dashboard launched for job %s", job_id)
    except Exception as exc:  # noqa: BLE001
        logger.warning("[WATCHER] Could not open watcher window: %s", exc)


def ignite_swarm(payload_path: str, starting_node: str = "OSINT") -> None:
    """Injects a payload into the SQLite WAL and ignites the pipeline.

    job_id format: ``job_{YYYYMMDD-HHMMSS-{4rand}}`` — human-readable,
    timestamped, and registered to project_registry.db so every launch
    path (not just ``global``) has a full audit record.
    """
    from maccre_core.utils.session_manager import (  # noqa: PLC0415
        generate_session_id,
        register_project,
        register_session,
    )
    project = os.environ.get("MACCRE_ACTIVE_PROJECT", "GLOBAL")
    job_id = f"job_{generate_session_id()}"  # e.g. job_20260607-161715-a3k9

    # Register every launch to the audit trail — previously only the
    # ``global`` command path did this.
    register_project(project, description="")
    register_session(
        project_name=project,
        session_id=job_id,
        workbook_type="launch",
    )

    try:
        if payload_path and payload_path.lower() != "none" and not os.path.exists(payload_path):
            msg = f"Payload not found: {payload_path}"
            logger.error("[FATAL] %s", msg)
            _log_launch_failure(job_id, payload_path, starting_node, msg)
            sys.exit(1)

        broker = LocalMessageBroker(_DB_PATH)
        broker.inject_task(job_id=job_id, payload_path=payload_path if payload_path.lower() != "none" else "none", starting_node=starting_node)

        logger.info("[IGNITION] Payload injected. Job ID: %s", job_id)
        logger.info("[IGNITION] Starting node : %s", starting_node)
        logger.info("[IGNITION] Run 'python -m maccre_core.orchestration.swarm_worker' to process.")
        print(f"[IGNITION] Job {job_id} queued at node '{starting_node}'. Worker ready.")

        # ── Launch watcher dashboard ────────────────────────────────────────
        _launch_watcher(job_id, project)

    except SystemExit:
        raise
    except Exception as exc:  # noqa: BLE001
        import traceback  # noqa: PLC0415
        tb = traceback.format_exc()
        logger.error("[IGNITION] Launch failed for job %s: %s\n%s", job_id, exc, tb)
        _log_launch_failure(job_id, payload_path, starting_node, f"{exc}\n{tb}")
        raise


def check_status() -> None:
    """Reads the SQLite WAL directly to report swarm status."""
    if not os.path.exists(_DB_PATH):
        logger.info("[STATUS] Swarm queue is empty or uninitialized.")
        print("[STATUS] Swarm queue is empty or uninitialized.")
        return

    with sqlite3.connect(_DB_PATH) as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        cursor = conn.execute(
            "SELECT id, job_id, current_node, lock_status, actual_cost "
            "FROM task_queue ORDER BY id DESC LIMIT 15"
        )
        rows = cursor.fetchall()

    if not rows:
        logger.info("[STATUS] Queue is empty.")
        print("[STATUS] Queue is empty.")
        return

    print(f"{'ID':<5} | {'JOB_ID':<12} | {'NODE':<20} | {'STATUS':<15} | {'COST'}")
    print("-" * 72)
    for r in rows:
        print(f"{r[0]:<5} | {r[1]:<12} | {r[2]:<20} | {r[3]:<15} | ${float(r[4] or 0):.6f}")


def canonize_session(project: str, session: str) -> None:
    """Executes the complete session canonization."""
    logger.info("[CANONIZE] Promoting Session '%s' → Project '%s'...", session, project)
    print(f"[CANONIZE] Promoting Session '{session}' → Project '{project}'...")
    try:
        from maccre_core.tools.rag_tools import canonize_session as rt_canonize_session  # noqa: PLC0415
        merge_res = rt_canonize_session(session, project)
        logger.info("[CANONIZE] Complete. Result:\n%s", merge_res)
        print(f"{merge_res}")
        print("[CANONIZE] Complete.")
    except Exception as exc:  # noqa: BLE001
        logger.error("[CANONIZE] Canonization failed: %s", exc)
        print(f" ⚠  Canonization failed: {exc}")


def canonize_project(project: str) -> None:
    """Executes the L2 → L3 memory promotion."""
    logger.info("[CANONIZE] Promoting Project '%s' → GLOBAL...", project)
    print(f"[CANONIZE] Promoting Project '{project}' → GLOBAL...")
    try:
        from maccre_core.tools.rag_tools import canonize_project_to_global  # noqa: PLC0415
        merge_res = canonize_project_to_global(project_name=project)
        logger.info("[CANONIZE] Complete. Result: %s", merge_res)
        print(f" ✓ {merge_res}")
        print("[CANONIZE] Complete.")
    except Exception as exc:  # noqa: BLE001
        logger.error("[CANONIZE] Project promotion failed: %s", exc)
        print(f" ⚠  Project promotion failed: {exc}")


def global_command(workbook_path: str = "", yes: bool = False, skip_preflight: bool = False) -> None:
    """Process MACCRE_Global.xlsx — create project, materialise swarm, optionally ignite."""
    from maccre_core.tools.workbook_engine import check_workbook_completeness, render_execution_plan  # noqa: PLC0415
    from maccre_core.tools.sheet_parser import materialise_from_sheet  # noqa: PLC0415
    from maccre_core.tools.admin_tools import run_swarm  # noqa: PLC0415
    from maccre_core.utils.session_manager import (  # noqa: PLC0415
        generate_session_id,
        register_project,
        register_session,
        complete_session,
    )
    from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415
    import sqlite3 as _sq  # noqa: PLC0415

    wb_path = Path(workbook_path) if workbook_path else get_maccre_root() / "MACCRE_Global.xlsx"

    if not wb_path.exists():
        print(f"[GLOBAL_FAULT] Workbook not found: {wb_path}")
        sys.exit(1)

    plan = check_workbook_completeness(wb_path, wb_type="global")
    print(render_execution_plan(plan), flush=True)

    if not plan.can_run:
        print("[GLOBAL_FAULT] Nothing actionable in workbook. Fill required fields and retry.")
        sys.exit(1)

    if not yes:
        names = " + ".join(s.name for s in plan.actionable_sections)
        try:
            answer = input(f"Proceed with {names}? [y/N]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            sys.exit(0)
        if answer not in ("y", "yes"):
            sys.exit(0)

    session_id = generate_session_id(plan.session_label)
    project    = plan.project_name or "UNNAMED"
    running_sections = [s.name for s in plan.actionable_sections]

    print(f"\n[GLOBAL] Session: {session_id}  Project: {project}", flush=True)

    # Register project + session
    register_project(project, description="")
    register_session(
        project_name=project,
        session_id=session_id,
        label=plan.session_label,
        workbook_type="global",
        sections_run=running_sections,
        est_cost_usd=plan.total_est_cost,
    )

    import shutil
    dc_archive = get_maccre_root() / "__DATACENTER" / project / "CompletedSessions"
    dc_archive.mkdir(parents=True, exist_ok=True)
    archive_path = dc_archive / f"{session_id}_Swarm_Request.xlsx"
    shutil.copy2(wb_path, archive_path)

    # Materialise swarm (creates workspace, agents, topology)
    _run_preflight(skip=skip_preflight)

    mat_sections = {"PROJECT_DEFINITION", "AGENTS", "TOPOLOGY", "SWARM_REQUEST"}
    if any(s.name in mat_sections for s in plan.actionable_sections):
        print("[GLOBAL] Materialising swarm from workbook...", flush=True)
        result = materialise_from_sheet(wb_path)
        print(f"  {result}", flush=True)

    # Run swarm if SWARM_REQUEST is actionable
    swarm_ready = any(
        s.name == "SWARM_REQUEST" and s.execute for s in plan.actionable_sections
    )
    actual_cost = 0.0
    if swarm_ready:
        from maccre_core.tools.sheet_parser import parse_workbook as _parse  # noqa: PLC0415
        from maccre_core.orchestration.local_broker import LocalMessageBroker  # noqa: PLC0415

        _parsed = _parse(wb_path)
        _start_node = _parsed.start_node or "START"
        _payload_file = _parsed.payload_path.strip() if _parsed.payload_path.strip() else "none"

        # job_id IS the session_id — same format generator, no bridge needed.
        # session_id was already registered above via register_session().
        _job_id = f"job_{session_id}"  # session_id already YYYYMMDD-HHMMSS-{4rand}
        _broker = LocalMessageBroker(str(get_datacenter_path("swarm_queue.db")))
        _broker.inject_task(
            job_id=_job_id,
            payload_path=_payload_file,
            starting_node=_start_node,
        )

        print("[GLOBAL] Running swarm...", flush=True)
        run_result = run_swarm(project_name=project, max_cycles=500, timeout_seconds=3600)
        print(f"\n{run_result}", flush=True)
        try:
            db = str(get_datacenter_path("swarm_queue.db"))
            with _sq.connect(db) as conn:
                row = conn.execute(
                    "SELECT COALESCE(SUM(actual_cost), 0.0) FROM task_queue"
                ).fetchone()
                actual_cost = float(row[0]) if row else 0.0
        except Exception:
            pass

    complete_session(session_id, actual_cost)
    print(f"\n[GLOBAL] ✓ Session {session_id} complete.  Actual cost: ${actual_cost:.6f}", flush=True)


def launch_command(
    project_name: str,
    yes: bool = False,
    resume: bool = False,
    from_node: str = "",
    workbook_path: str = "",
    skip_preflight: bool = False,
) -> None:
    """Process the Session Workbook for a project and run the swarm.

    --workbook PATH:  Use a specific workbook instead of MACCRE_Session.xlsx.
    --resume:         Skip materialise/ignite — run against existing pending queue rows.
    --from-node NODE: Insert a fresh pending row at NODE then resume.
                      Both flags allow targeted checkpoint restarts without re-running
                      upstream nodes in the pipeline.
    """
    from maccre_core.tools.workbook_engine import check_workbook_completeness, render_execution_plan  # noqa: PLC0415
    from maccre_core.tools.sheet_parser import materialise_from_sheet  # noqa: PLC0415
    from maccre_core.tools.admin_tools import run_swarm  # noqa: PLC0415
    from maccre_core.tools.rag_tools import ingest_project  # noqa: PLC0415
    from maccre_core.utils.session_manager import (  # noqa: PLC0415
        generate_session_id,
        register_session,
        complete_session,
    )
    from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415
    import sqlite3 as _sq  # noqa: PLC0415
    from openpyxl import load_workbook as _lwb  # noqa: PLC0415

    os.environ["MACCRE_ACTIVE_PROJECT"] = project_name

    # ── RESUME / FROM-NODE: lightweight checkpoint restart, bypasses full launch ──
    if resume or from_node:
        from maccre_core.utils.path_resolver import get_datacenter_path as _gdcp  # noqa: PLC0415
        if from_node:
            from maccre_core.orchestration.local_broker import LocalMessageBroker  # noqa: PLC0415
            from maccre_core.utils.session_manager import generate_session_id  # noqa: PLC0415
            _payload = str(_gdcp("01_Raw_Source/input.md"))
            _retry_id = f"job_{generate_session_id('resume')}"  # e.g. job_20260607-162035-a3k9-resume
            broker = LocalMessageBroker()
            broker.inject_task(
                job_id=_retry_id,
                payload_path=_payload,
                starting_node=from_node.upper(),
            )
            print(f"[RESUME] Checkpoint row inserted: job={_retry_id} node={from_node.upper()}")
        print(f"[RESUME] Running pending queue for '{project_name}' (materialise skipped)...")
        run_result = run_swarm(project_name=project_name, max_cycles=100, timeout_seconds=1800)
        print(f"\n{run_result}")
        return


    # Look for workbook: explicit --workbook flag > MACCRE_Session.xlsx > legacy name
    base = get_maccre_root() / "__DATACENTER" / project_name
    if workbook_path:
        wb_path = Path(workbook_path)
    else:
        wb_path = base / "MACCRE_Session.xlsx"
        if not wb_path.exists():
            wb_path = base / "MACCRE_Swarm_Request.xlsx"
    if not wb_path.exists():
        print(f"[LAUNCH_FAULT] No workbook found: {wb_path}")
        print("  Expected: MACCRE_Session.xlsx  (or use --workbook PATH)")
        sys.exit(1)

    plan = check_workbook_completeness(wb_path, wb_type="session")
    print(render_execution_plan(plan))

    if not plan.can_run:
        print("[LAUNCH_FAULT] Nothing actionable. Fill in required fields and retry.")
        sys.exit(1)

    # Read SESSION_CONFIG hooks
    ingest_before = False
    ingest_after  = False
    canonize_after = False
    try:
        _wb = _lwb(filename=str(wb_path), read_only=True, data_only=True)
        if "SESSION_CONFIG" in _wb.sheetnames:
            ws_sc = _wb["SESSION_CONFIG"]
            for row in ws_sc.iter_rows(min_row=3, values_only=True):
                if not row or not row[0]:
                    continue
                k = str(row[0]).strip().upper().replace(" ", "_")
                v = str(row[1] or "").strip().upper() if len(row) > 1 else ""
                if k == "INGEST_BEFORE_RUN":
                    ingest_before = v in ("TRUE", "YES", "1")
                if k == "INGEST_AFTER_RUN":
                    ingest_after = v in ("TRUE", "YES", "1")
                if k == "CANONIZE_AFTER_RUN":
                    canonize_after = v in ("TRUE", "YES", "1")
        _wb.close()
    except Exception:  # noqa: BLE001
        pass

    if not yes:
        names = " + ".join(s.name for s in plan.actionable_sections)
        try:
            answer = input(f"Launch '{project_name}' — run {names}? [y/N]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\nAborted.")
            sys.exit(0)
        if answer not in ("y", "yes"):
            print("Aborted.")
            sys.exit(0)

    session_label = plan.session_label or ""
    session_id = generate_session_id(session_label)
    register_session(
        project_name=project_name,
        session_id=session_id,
        label=session_label,
        workbook_type="session",
        sections_run=[s.name for s in plan.actionable_sections],
        est_cost_usd=plan.total_est_cost,
    )
    print(f"\n[LAUNCH] Session: {session_id}  Project: {project_name}")

    import shutil
    dc_archive = get_maccre_root() / "__DATACENTER" / project_name / "CompletedSessions"
    dc_archive.mkdir(parents=True, exist_ok=True)
    archive_path = dc_archive / f"{session_id}_Session.xlsx"
    shutil.copy2(wb_path, archive_path)
    print(f"[LAUNCH] Session workbook frozen to {archive_path.name}")

    if ingest_before:
        print("[LAUNCH] INGEST_BEFORE_RUN=TRUE — ingesting 01_Raw_Source...")
        ingest_project(project_name, session_id=session_id)


    # ── Pre-Flight Smoke Gate ────────────────────────────────────────────────
    _run_preflight(skip=skip_preflight)
    # ─────────────────────────────────────────────────────────────────────────

    # Materialise + run
    print("[LAUNCH] Materialising from workbook...")
    mat_result = materialise_from_sheet(wb_path)
    print(f"  {mat_result}")

    # ── Pre-Flight Topology Validation ───────────────────────────────────────
    # Run AFTER materialise (which writes topology.csv) so the validator reads
    # the exact file that swarm_worker will consume. Zero API cost.
    print("[LAUNCH] Running pre-flight topology validation...")
    try:
        from maccre_core.orchestration.topology_engine import TopologyEngine  # noqa: PLC0415
        _topo = TopologyEngine()
        _report = _topo.validate()
        print(_report.render_table(), end="")
        if not _report.is_ok and not _report.skipped:
            print(
                "[LAUNCH] ✗ Pre-flight FAILED — fix the errors above before re-running.\n"
                "  (Set MACCRE_SKIP_VALIDATE=1 to bypass for dynamic topologies.)"
            )
            sys.exit(1)
    except Exception as _val_err:
        print(f"[LAUNCH] ⚠  Validation skipped (could not load topology): {_val_err}")
    # ─────────────────────────────────────────────────────────────────────────



    if any(s.name == "SWARM_REQUEST" and s.execute for s in plan.actionable_sections):
        from maccre_core.tools.admin_tools import ignite_swarm  # noqa: PLC0415
        # Read START_NODE from the SWARM_REQUEST sheet
        try:
            _wb2 = _lwb(filename=str(wb_path), read_only=True, data_only=True)
            _ws_req = _wb2["SWARM_REQUEST"]
            _hmap = {
                str(c.value).strip().upper().lstrip("\u2605* ").strip(): c.column - 1
                for c in next(_ws_req.iter_rows(min_row=2, max_row=2))  # type: ignore[call-overload]
                if c.value is not None
            }
            _start_node = "ANCHOR"  # sensible default
            _payload_path = "input.md"
            for _row in _ws_req.iter_rows(min_row=3, values_only=True):
                if _row and _row[0]:
                    sn_idx = _hmap.get("START_NODE")
                    if sn_idx is not None and sn_idx < len(_row) and _row[sn_idx]:
                        _start_node = str(_row[sn_idx]).strip()
                        
                    pl_idx = _hmap.get("PAYLOAD_PATH")
                    if pl_idx is not None and pl_idx < len(_row) and _row[pl_idx]:
                        _payload_path = str(_row[pl_idx]).strip()
                    break
            _wb2.close()
        except Exception:  # noqa: BLE001
            _start_node = "ANCHOR"
            _payload_path = "input.md"
            
        print(f"[LAUNCH] Igniting swarm at node: {_start_node}")
        ignite_result = ignite_swarm(_payload_path, _start_node)
        print(f"  {ignite_result}")
        if "[ADMIN_SUCCESS]" not in ignite_result:
            print("[LAUNCH] Ignition failed — aborting run. Check payload path.")
        else:
            print("[LAUNCH] Running swarm...")
            run_result = run_swarm(project_name=project_name, max_cycles=100, timeout_seconds=1800)
            print(f"\n{run_result}")

    actual_cost = 0.0
    try:
        db = str(get_datacenter_path("swarm_queue.db"))
        with _sq.connect(db) as conn:
            row = conn.execute(
                "SELECT COALESCE(SUM(actual_cost), 0.0) FROM task_queue"
            ).fetchone()
            actual_cost = float(row[0]) if row else 0.0
    except Exception:  # noqa: BLE001
        pass

    if ingest_after:
        print("[LAUNCH] INGEST_AFTER_RUN=TRUE — ingesting 04_Code_Artifacts into knowledge store...")
        ingest_project(project_name, session_id=session_id)

    complete_session(session_id, actual_cost)

    if canonize_after:
        print("[LAUNCH] CANONIZE_AFTER_RUN=TRUE — canonizing session...")
        canonize_session(project_name, session_id)

    print(f"\n[LAUNCH] ✓ Session {session_id} complete.  Actual cost: ${actual_cost:.6f}")


def ingest_command(project_name: str) -> None:
    """Bulk-ingest 01_Raw_Source for a project using SHA-256 hash manifest."""
    from maccre_core.tools.rag_tools import ingest_project  # noqa: PLC0415
    from maccre_core.utils.session_manager import generate_session_id  # noqa: PLC0415
    session_id = generate_session_id("ingest")
    ingest_project(project_name, session_id=session_id)


def smoke_command() -> None:
    """Run the pre-flight smoke test standalone and exit 0/1."""
    from maccre_core.tests.smoke_test import run_smoke_test  # noqa: PLC0415
    sys.exit(0 if run_smoke_test() else 1)


def run_command(
    project: str,
    payload: str,
    node: str = "OSINT",
    yes: bool = False,
    skip_preflight: bool = False,
    visuals: bool = False,
) -> None:
    """Direct swarm launcher — no workbook required.

    Payload resolution order:
      1. ``@/path/to/file.md``  — explicit file reference (@ prefix)
      2. Existing filesystem path  — reads the file
      3. Bare text string  — written directly to 01_Raw_Source/input.md

    Examples (human or agentic):
      python maccre.py run NewsNexus "Tell me about Iran" --node OSINT --yes
      python maccre.py run NewsNexus @brief.md --skip-preflight
      python maccre.py run NewsNexus /abs/path/payload.md --yes
    """
    from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415
    from maccre_core.tools.admin_tools import run_swarm  # noqa: PLC0415
    from maccre_core.orchestration.local_broker import LocalMessageBroker as _Broker  # noqa: PLC0415
    from maccre_core.orchestration.topology_engine import TopologyEngine  # noqa: PLC0415

    os.environ["MACCRE_ACTIVE_PROJECT"] = project
    os.environ["MACCRE_SKIP_AUTH"] = "1"
    if visuals:
        os.environ["MACCRE_VISUALS"] = "1"

    # ── Pre-flight gate ───────────────────────────────────────────────────────
    _run_preflight(skip=skip_preflight)

    # ── Resolve payload ───────────────────────────────────────────────────────
    payload_text: str
    if payload.startswith("@"):
        p = Path(payload[1:])
        if not p.exists():
            print(f"[RUN] Payload file not found: {p}")
            sys.exit(1)
        payload_text = p.read_text(encoding="utf-8")
    elif Path(payload).exists():
        payload_text = Path(payload).read_text(encoding="utf-8")
    else:
        payload_text = payload

    dc_root = get_maccre_root() / "__DATACENTER" / project
    input_path = dc_root / "01_Raw_Source" / "input.md"
    input_path.parent.mkdir(parents=True, exist_ok=True)
    input_path.write_text(payload_text, encoding="utf-8")
    print(f"[RUN] Payload ({len(payload_text)} chars) → {input_path}")

    # ── Topology pre-flight validate ──────────────────────────────────────────
    print("[RUN] Validating topology...")
    try:
        _topo = TopologyEngine()
        _report = _topo.validate()
        print(_report.render_table(), end="")
        if not _report.is_ok and not _report.skipped:
            print("[RUN] \u2717 Topology validation failed \u2014 fix errors above.")
            sys.exit(1)
    except Exception as _ve:  # noqa: BLE001
        print(f"[RUN] \u26a0  Topology validation skipped: {_ve}")

    # ── Confirm ───────────────────────────────────────────────────────────────
    if not yes:
        try:
            ans = input(f"Launch '{project}' from node '{node.upper()}'? [y/N]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            sys.exit(0)
        if ans not in ("y", "yes"):
            sys.exit(0)

    # ── Inject + run ──────────────────────────────────────────────────────────
    from maccre_core.utils.session_manager import generate_session_id  # noqa: PLC0415
    job_id = f"job_{generate_session_id('resume')}"  # e.g. job_20260607-162035-a3k9-resume
    _db = str(dc_root / "swarm_queue.db")
    _broker = _Broker(_db)
    _broker.inject_task(
        job_id=job_id,
        payload_path=str(input_path),
        starting_node=node.upper(),
    )
    print(f"[RUN] Job {job_id} queued at node '{node.upper()}'")
    run_result = run_swarm(project_name=project, max_cycles=500, timeout_seconds=3600)
    print(f"\n{run_result}")


def audit_command(
    project: str,
    job_id: str = "",
    node: str = "",
    tail: int = 0,
) -> None:
    """Display forensic tool audit sidecars from 03_Agent_Ledgers.

    Audits are written by the swarm_worker whenever the agentic tool loop fires.
    Each file captures every tool call + result verbatim, timestamped per node.

    Examples:
      python maccre.py audit NewsNexus
      python maccre.py audit NewsNexus --job job_resume_abc123
      python maccre.py audit NewsNexus --node OSINT --tail 80
    """
    from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415

    ledger_root = get_maccre_root() / "__DATACENTER" / project / "03_Agent_Ledgers"
    if not ledger_root.exists():
        print(f"[AUDIT] No ledgers found for project '{project}'.")
        return

    node_prefix = node.upper() + "_" if node else ""
    pattern = f"tool_audit_{node_prefix}*.md"

    files: list[Path]
    if job_id:
        job_dir = ledger_root / job_id
        files = sorted(
            job_dir.glob(pattern) if job_dir.exists() else [],
            key=lambda p: p.stat().st_mtime,
        )
    else:
        files = sorted(
            ledger_root.rglob(pattern),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )

    if not files:
        print(f"[AUDIT] No tool audit files found (pattern: {pattern}).")
        print("  Audits are only written when tools fire. Check that OSINT ran a tool loop.")
        return

    for audit_file in files:
        rel = audit_file.relative_to(ledger_root)
        print(f"\n{'='*72}\n  {rel}\n{'='*72}")
        content = audit_file.read_text(encoding="utf-8")
        if tail:
            lines = content.splitlines()
            content = "\n".join(lines[-tail:])
        print(content)


def new_command(project_name: str) -> None:
    """Provision a new DATACENTER project silo with 5-tier directory tree and fresh DBs."""
    from maccre_core.tools.admin_tools import initialize_workspace  # noqa: PLC0415
    from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415
    import subprocess as _sp  # noqa: PLC0415

    print(f"[NEW] Provisioning project silo: '{project_name}'...")
    result = initialize_workspace(project_name)
    print(result)

    if "[ADMIN_SUCCESS]" in result:
        # Auto-generate a fresh MACCRE_Session.xlsx into the new silo
        root = get_maccre_root()
        gen_script = root / "scripts" / "generate_global_template.py"
        if gen_script.exists():
            _sp.run(
                [sys.executable, str(gen_script), "--session", "--project", project_name],
                cwd=str(root), check=False, capture_output=True,
            )
            print(f"[NEW] MACCRE_Session.xlsx written to __DATACENTER/{project_name}/")
        print(f"[NEW] Ready. Next step:\n"
              f"  1. Drop source files into __DATACENTER/{project_name}/01_Raw_Source/\n"
              f"  2. python maccre.py ingest {project_name}\n"
              f"  3. Fill __DATACENTER/{project_name}/MACCRE_Session.xlsx\n"
              f"  4. python maccre.py launch {project_name}")


def sessions_command(action: str) -> None:
    """List or kill tracked MACCRE session processes to prevent stale WAL locks."""
    import json as _json  # noqa: PLC0415
    from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415

    registry_path = get_maccre_root() / ".session_pids.json"

    if action == "list":
        if not registry_path.exists():
            print("[SESSIONS] No active session registry found.")
            return
        with open(registry_path, encoding="utf-8") as f:
            entries: list[dict[str, object]] = _json.load(f)
        if not entries:
            print("[SESSIONS] Registry is empty.")
            return
        print(f"{'PID':<8} {'PROJECT':<20} {'STARTED':<22} {'DB'}")
        print("-" * 80)
        for e in entries:
            pid = e.get("pid", "?")
            proj = e.get("project", "?")
            ts   = e.get("started", "?")
            db   = e.get("db_path", "?")
            # Mark dead pids
            alive = _is_pid_alive(int(pid)) if isinstance(pid, int) else False
            tag = "" if alive else "[ZOMBIE]"
            print(f"{pid:<8} {proj:<20} {str(ts):<22} {db} {tag}")

    elif action == "kill":
        if not registry_path.exists():
            print("[SESSIONS] No registry to purge.")
            return
        with open(registry_path, encoding="utf-8") as f:
            entries = _json.load(f)
        killed = 0
        live: list[dict[str, object]] = []
        for e in entries:
            pid = e.get("pid")
            if isinstance(pid, int) and _is_pid_alive(pid):
                try:
                    import signal  # noqa: PLC0415
                    import os as _os  # noqa: PLC0415
                    _os.kill(pid, signal.SIGTERM)
                    print(f"[SESSIONS] Sent SIGTERM to PID {pid} ({e.get('project', '?')})")
                    killed += 1
                except Exception as exc:  # noqa: BLE001
                    print(f"[SESSIONS] Could not kill PID {pid}: {exc}")
            else:
                live.append(e)
        with open(registry_path, "w", encoding="utf-8") as f:
            _json.dump(live, f, indent=2)
        print(f"[SESSIONS] Killed {killed} session(s). Registry cleaned.")


def _is_pid_alive(pid: int) -> bool:
    """Return True if a process with the given PID is currently running."""
    import os as _os  # noqa: PLC0415
    try:
        _os.kill(pid, 0)   # signal 0 = existence check, no actual signal sent
        return True
    except OSError:
        return False




def main() -> None:
    parser = argparse.ArgumentParser(
        description="MACCREv2 — Multi-Agent Conversational Concept Refinement Engine",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Workbook Sovereign Commands:\n"
            "  python maccre.py new <name>                  → provision new project silo\n"
            "  python maccre.py global                      → process MACCRE_Global.xlsx\n"
            "  python maccre.py launch <project>            → run Session Workbook\n"
            "  python maccre.py ingest <project>            → hash-aware document ingest\n"
            "  python maccre.py status [project]            → queue + session status\n"
            "  python maccre.py canonize --project P --session S\n"
            "  python maccre.py sessions list               → show tracked swarm PIDs\n"
            "  python maccre.py sessions kill               → kill stale swarm PIDs\n"
            "  python maccre.py mcp                         → start MCP server\n"
            "  python maccre.py sync --project <name>       → Drive sync\n"
            "  python maccre.py logs clear --project P      → erase telemetry\n"
            "\n"
            "Direct Swarm Commands (no workbook required):\n"
            "  python maccre.py smoke                       → pre-flight check ($0.00)\n"
            "  python maccre.py run <project> \"<payload>\"  → direct swarm launch\n"
            "  python maccre.py run NewsNexus \"Tell me about Iran\" --node OSINT --yes\n"
            "  python maccre.py run NewsNexus @brief.md --skip-preflight\n"
            "  python maccre.py audit <project>             → forensic tool audit logs\n"
            "  python maccre.py audit NewsNexus --node OSINT --tail 80\n"
        ),
    )
    parser.add_argument("--toglog", action="store_true", help="Toggle OFF verbose JSON Bug-logs")
    subparsers = parser.add_subparsers(dest="command", required=True)

    # new — provision a fresh project silo
    new_parser = subparsers.add_parser(
        "new", help="Provision a new project silo with 5-tier DATACENTER and fresh databases"
    )
    new_parser.add_argument("project_name", help="Alphanumeric project name (e.g. NewsNexus)")

    # sessions — PID registry management to prevent stale WAL locks
    sessions_parser = subparsers.add_parser(
        "sessions", help="List or kill tracked swarm session PIDs (fixes SQLite WAL locks)"
    )
    sessions_parser.add_argument(
        "action", choices=["list", "kill"],
        help="list = show all tracked PIDs | kill = SIGTERM all live PIDs and clean registry"
    )

    # global
    global_parser = subparsers.add_parser(
        "global", help="Process MACCRE_Global.xlsx (create project + materialise swarm)"
    )
    global_parser.add_argument(
        "--workbook", default="", help="Override path to MACCRE_Global.xlsx"
    )
    global_parser.add_argument(
        "--yes", "-y", action="store_true", help="Skip confirmation prompt"
    )
    global_parser.add_argument(
        "--skip-preflight", action="store_true", dest="skip_preflight",
        help="Bypass the pre-flight smoke test (or set MACCRE_SKIP_PREFLIGHT=1)"
    )

    # launch
    launch_parser = subparsers.add_parser(
        "launch", help="Process a project's Session Workbook and run the swarm"
    )
    launch_parser.add_argument("project", help="Project silo name")
    launch_parser.add_argument(
        "--yes", "-y", action="store_true", help="Skip confirmation prompt"
    )
    launch_parser.add_argument(
        "--resume", action="store_true",
        help="Skip materialise/ignite — run against existing pending queue rows only"
    )
    launch_parser.add_argument(
        "--workbook", default="", dest="workbook_path",
        help="Path to a specific .xlsx workbook (default: MACCRE_Session.xlsx in project silo)"
    )
    launch_parser.add_argument(
        "--from-node", dest="from_node", default="",
        help="Insert a fresh pending row at NODE then run (implies --resume). "
             "Example: --from-node RENDER"
    )
    launch_parser.add_argument(
        "--skip-preflight", action="store_true", dest="skip_preflight",
        help="Bypass the pre-flight smoke test (or set MACCRE_SKIP_PREFLIGHT=1)"
    )

    # ingest
    ingest_parser = subparsers.add_parser(
        "ingest", help="Hash-aware bulk ingest of 01_Raw_Source for a project"
    )
    ingest_parser.add_argument("project", help="Project silo name")

    # smoke — standalone pre-flight gate
    subparsers.add_parser(
        "smoke", help="Run the pre-flight smoke test ($0.00, Gemma API)"
    )

    # run — direct swarm launcher, no workbook required
    run_parser = subparsers.add_parser(
        "run",
        help="Direct swarm launch without a workbook (human + agentic use)",
    )
    run_parser.add_argument("project", help="Project silo name (e.g. NewsNexus)")
    run_parser.add_argument(
        "payload",
        help="Inline text, @file.md reference, or absolute file path",
    )
    run_parser.add_argument(
        "--node", default="OSINT",
        help="Entry node to start from (default: OSINT)",
    )
    run_parser.add_argument("--yes", "-y", action="store_true", help="Skip confirmation")
    run_parser.add_argument(
        "--skip-preflight", action="store_true", dest="skip_preflight",
        help="Bypass the pre-flight smoke test",
    )
    run_parser.add_argument(
        "--visuals", action="store_true",
        help="Enable visual generation (audio-only is default)",
    )

    # audit — read forensic tool audit sidecars
    audit_parser = subparsers.add_parser(
        "audit", help="Display forensic tool audit logs from agent ledgers"
    )
    audit_parser.add_argument("project", help="Project silo name")
    audit_parser.add_argument("--job", default="", dest="job_id", help="Filter by job ID")
    audit_parser.add_argument("--node", default="", help="Filter by node name (e.g. OSINT)")
    audit_parser.add_argument(
        "--tail", type=int, default=0,
        help="Show only the last N lines of each audit file",
    )

    # status
    status_parser = subparsers.add_parser("status", help="View queue and session status")
    status_parser.add_argument(
        "project", nargs="?", default="", help="Filter by project (optional)"
    )

    # canonize
    canonize_parser = subparsers.add_parser(
        "canonize", help="Merge L1 session memory into L2 project memory, or L2 to L3 global"
    )
    canonize_parser.add_argument("--project", required=True, help="Project ID")
    canonize_parser.add_argument("--session", default="", help="Session ID (leave empty to canonize project to global)")

    # mcp
    subparsers.add_parser("mcp", help="Start the MACCRE MCP server (agentic interface)")

    # brief — synchronous session re-contextualization
    brief_parser = subparsers.add_parser(
        "brief", help="Print a session brief (git log + cost + sentinel health). Zero cost."
    )
    brief_parser.add_argument(
        "--project", default="", help="Override active project (default: MACCRE_ACTIVE_PROJECT)"
    )

    # pattern — CLI surface for the pattern library
    pattern_parser = subparsers.add_parser("pattern", help="Submit or poll a swarm topology pattern")
    pattern_sub = pattern_parser.add_subparsers(dest="pattern_action", required=True)
    pattern_sub.add_parser("list", help="List registered patterns")
    p_submit = pattern_sub.add_parser("submit", help="Submit a pattern")
    p_submit.add_argument("name", help="Pattern name (e.g. simulation_swarm)")
    p_submit.add_argument("payload", help="Path to payload .md file or inline text")
    p_submit.add_argument("--project", default="", help="Project silo")
    p_submit.add_argument("--cost-limit", type=float, default=5.0, dest="cost_limit")
    p_poll = pattern_sub.add_parser("poll", help="Poll a running pattern job")
    p_poll.add_argument("job_id", help="job_id from submit")
    p_poll.add_argument("--silo", default="", help="Silo project name (optional)")

    # config
    config_parser = subparsers.add_parser(
        "config", help="Manage MACCRE configuration and sovereign keys"
    )
    config_parser.add_argument("action", choices=["set-key"], help="config action")
    config_parser.add_argument("payload_string", help="The raw API key string")
    config_parser.add_argument(
        "--vault-name", dest="vault_name", default="",
        help="Explicit vault target name (bypasses fingerprinting). "
             "Example: --vault-name BRAVE_SEARCH_API_KEY",
    )

    # sync
    sync_parser = subparsers.add_parser(
        "sync", help="Sync project database nuggets with Google Drive"
    )
    sync_parser.add_argument("--project", required=True, help="Project silo name")
    sync_parser.add_argument("--export", action="store_true", help="Export nuggets to Drive")
    sync_parser.add_argument(
        "--list", action="store_true", dest="list_nuggets", help="List available nuggets"
    )

    # ignite (internal — not on help surface but still accessible)
    ignite_parser = subparsers.add_parser(argparse.SUPPRESS)
    ignite_parser.add_argument("payload")
    ignite_parser.add_argument("--node", default="OSINT")

    # logs
    logs_parser = subparsers.add_parser("logs", help="Manage session and project logs")
    logs_parser.add_argument("action", choices=["clear"], help="Action to perform")
    logs_parser.add_argument("--project", required=True, help="Target project")
    logs_parser.add_argument("--session", default="all", help="Target session ID or 'all'")
    logs_parser.add_argument("--type", choices=["op", "bug", "all"], default="all", help="Log category")

    # intercept
    intercept_parser = subparsers.add_parser("intercept", help="Inject a live Hot-Mic override into a running session")
    intercept_parser.add_argument("session", help="Target session ID")
    intercept_parser.add_argument("message", help="The override text to inject")

    # workbook
    wb_parser = subparsers.add_parser("workbook", help="Generate or fire the MACCRE Global Workbook")
    wb_sub = wb_parser.add_subparsers(dest="wb_action", required=True)
    wb_refresh = wb_sub.add_parser("refresh", help="Regenerate MACCRE_Global.xlsx with live data")
    wb_refresh.add_argument("--project", default="", help="Pre-populate from project silo (default: GLOBAL)")
    wb_refresh.add_argument("--out", default="", help="Override output path")
    wb_fire = wb_sub.add_parser("fire", help="Read workbook SWARM_REQUEST and launch a swarm")
    wb_fire.add_argument("--workbook", default="", help="Path to xlsx (default: MACCRE_Global.xlsx at root)")
    wb_fire.add_argument("--yes", action="store_true", help="Skip confirmation prompt")

    # topology
    topo_parser = subparsers.add_parser("topology", help="Manage the named topology library")
    topo_sub = topo_parser.add_subparsers(dest="topo_action", required=True)
    topo_sub.add_parser("list", help="List all saved topologies").add_argument(
        "--project", default="", help="Filter to project library (default: GLOBAL)"
    )
    topo_save = topo_sub.add_parser("save", help="Save the active project topology to the library")
    topo_save.add_argument("name", help="Name for the topology")
    topo_save.add_argument("--project", required=True, help="Source project silo")
    topo_save.add_argument("--description", default="", help="Optional description")
    topo_load = topo_sub.add_parser("load", help="Load a named topology into a project")
    topo_load.add_argument("name", help="Topology name to load")
    topo_load.add_argument("--project", required=True, help="Target project silo")
    topo_load.add_argument("--yes", action="store_true", help="Skip confirmation")
    topo_del = topo_sub.add_parser("delete", help="Delete a named topology")
    topo_del.add_argument("name", help="Topology name to delete")
    topo_del.add_argument("--project", default="", help="Library to delete from (default: GLOBAL)")

    args = parser.parse_args()

    if getattr(args, "toglog", False):
        import maccre_core.logger as mlogger
        mlogger.ENABLE_DEBUG_LOGGING = False

    if args.command == "new":
        new_command(args.project_name)
    elif args.command == "sessions":
        sessions_command(args.action)
    elif args.command == "global":
        global_command(
            workbook_path=args.workbook,
            yes=args.yes,
            skip_preflight=args.skip_preflight,
        )
    elif args.command == "launch":
        launch_command(
            project_name=args.project,
            yes=args.yes,
            resume=args.resume,
            from_node=args.from_node,
            workbook_path=args.workbook_path,
            skip_preflight=args.skip_preflight,
        )
    elif args.command == "ingest":
        ingest_command(project_name=args.project)
    elif args.command == "smoke":
        smoke_command()
    elif args.command == "run":
        run_command(
            project=args.project,
            payload=args.payload,
            node=args.node,
            yes=args.yes,
            skip_preflight=args.skip_preflight,
            visuals=args.visuals,
        )
    elif args.command == "audit":
        audit_command(
            project=args.project,
            job_id=args.job_id,
            node=args.node,
            tail=args.tail,
        )
    elif args.command == "status":
        check_status()
    elif args.command == "canonize":
        if args.session:
            canonize_session(args.project, args.session)
        else:
            canonize_project(args.project)
    elif args.command == "chat":
        print("[DEPRECATED] The 'chat' command has been retired. Use 'maccre.py mcp' instead.")
    elif args.command == "mcp":
        from maccre_mcp import mcp  # noqa: PLC0415
        mcp.run()
    elif args.command == "config":
        if args.action == "set-key":
            from maccre_core.orchestration.key_ingestor import ingest_key  # noqa: PLC0415
            print("[INGESTOR] Analyzing key entropy...")
            result = ingest_key(args.payload_string, vault_name=getattr(args, "vault_name", ""))
            print(result)
    elif args.command == "sync":
        from maccre_core.tools.sync_tools import (  # noqa: PLC0415
            export_project_nugget,
            import_project_nuggets,
            list_project_nuggets,
        )
        if args.list_nuggets:
            print(list_project_nuggets(args.project))
        elif args.export:
            print(export_project_nugget(args.project))
        else:
            print(import_project_nuggets(args.project))
    elif args.command == "ignite":
        ignite_swarm(args.payload, args.node)
    elif args.command == "logs":
        if args.action == "clear":
            from maccre_core.logger import clear_session_logs  # noqa: PLC0415
            print(f"[LOGS] Purging [{args.type}] logs for [{args.project}] (Session: {args.session})...")
            if args.session == "all":
                # Clear all matching logs in directory
                from maccre_core.utils.path_resolver import get_maccre_root  # noqa: PLC0415
                path = get_maccre_root() / "__DATACENTER" / args.project
                count = 0
                for cat in (["Op-logs", "Bug-logs"] if args.type == "all" else [f"{args.type.title()}-logs"]):
                    target_dir = path / cat
                    if target_dir.exists():
                        import shutil
                        shutil.rmtree(target_dir)
                        count += 1
                print(f"[LOGS] Blew away {count} directories.")
            else:
                res = clear_session_logs(args.project, args.session, args.type)
                print(f"[LOGS] {res}")
    elif args.command == "intercept":
        broker = LocalMessageBroker()
        broker.inject_interrupt(args.session, args.message)
        print(f"\n[HOT-MIC] ⚡ Priority override deployed into Session [{args.session}]!\n> '{args.message}'\n")
    elif args.command == "brief":
        from maccre_core.tools.pattern_tools import get_session_brief  # noqa: PLC0415
        pid = args.project or os.environ.get("MACCRE_ACTIVE_PROJECT", "GLOBAL")
        print(get_session_brief(pid))
    elif args.command == "pattern":
        from maccre_core.tools.pattern_tools import (  # noqa: PLC0415
            list_patterns as _list_patterns,
            submit_pattern as _submit_pattern,
            poll_human_gate as _poll_gate,
        )
        if args.pattern_action == "list":
            print(_list_patterns())
        elif args.pattern_action == "submit":
            from pathlib import Path as _Path  # noqa: PLC0415
            payload_text = (
                _Path(args.payload).read_text(encoding="utf-8")
                if _Path(args.payload).exists()
                else args.payload
            )
            print(_submit_pattern(args.name, payload_text, args.project, args.cost_limit))
        elif args.pattern_action == "poll":
            print(_poll_gate(args.job_id, args.silo))
    elif args.command == "workbook":
        from pathlib import Path as _Path  # noqa: PLC0415
        from scripts.generate_global_template import build_global_workbook  # noqa: PLC0415
        from maccre_core.utils.path_resolver import get_maccre_root as _gmr_wb  # noqa: PLC0415
        _root = _gmr_wb()
        if args.wb_action == "refresh":
            _out = _Path(args.out) if args.out else _root / "MACCRE_Global.xlsx"
            build_global_workbook(_out, project_id=args.project)
        elif args.wb_action == "fire":
            _wb_path = _Path(args.workbook) if args.workbook else _root / "MACCRE_Global.xlsx"
            if not _wb_path.exists():
                print(f"[WORKBOOK] No workbook found at {_wb_path}. Run 'workbook refresh' first.")
            else:
                from maccre_core.tools.sheet_parser import parse_workbook as _parse  # noqa: PLC0415
                _parsed = _parse(str(_wb_path))
                _project = _parsed.project_name.strip()
                _start = _parsed.start_node or "START"
                if not _project:
                    print("[WORKBOOK] PROJECT_NAME is empty in the workbook. Fill it in and retry.")
                else:
                    if not args.yes:
                        _confirm = input(f"[WORKBOOK] Fire swarm for project '{_project}' at node '{_start}'? [y/N] ")
                        if _confirm.strip().lower() not in ("y", "yes"):
                            print("[WORKBOOK] Aborted.")
                            return
                    print(f"[WORKBOOK] Materialising and launching '{_project}'...")
                    from maccre_core.tools.sheet_parser import materialise_from_sheet  # noqa: PLC0415
                    print(materialise_from_sheet(str(_wb_path)))
                    ignite_swarm(payload_path=str(_root / "__DATACENTER" / _project / "01_Raw_Source" / "input.md"),
                                 starting_node=_start)
                    # Archive workbook
                    import shutil as _shutil  # noqa: PLC0415
                    _archive = _root / "__DATACENTER" / _project / "_workbook_archive"
                    _archive.mkdir(parents=True, exist_ok=True)
                    from maccre_core.utils.session_manager import generate_session_id  # noqa: PLC0415
                    _shutil.copy2(_wb_path, _archive / f"{generate_session_id()}_Global.xlsx")
                    print(f"[WORKBOOK] Archived to {_archive}")
    elif args.command == "topology":
        from maccre_core.topology_library import get_topology_store, save_topology_globally  # noqa: PLC0415
        from maccre_core.workbook_data import load_topology_csv, load_agent_roster_csv  # noqa: PLC0415
        import csv as _csv  # noqa: PLC0415
        if args.topo_action == "list":
            _store = get_topology_store(getattr(args, "project", ""))
            _rows = _store.list_all()
            if not _rows:
                print("[TOPOLOGY] No saved topologies found.")
            else:
                print(f"{'NAME':<30} {'NODES':<7} {'CREATED':<22} DESCRIPTION")
                print("-" * 80)
                for _r in _rows:
                    print(f"{_r['name']:<30} {_r['node_count']:<7} {_r['created_at'][:19]:<22} {_r['description']}")
        elif args.topo_action == "save":
            _topo = load_topology_csv(args.project)
            _roster = load_agent_roster_csv(args.project)
            save_topology_globally(args.name, _topo, _roster, args.description, args.project)
            print(f"[TOPOLOGY] Saved '{args.name}' ({len(_topo)} nodes) to {args.project} + GLOBAL libraries.")
        elif args.topo_action == "load":
            _store = get_topology_store("GLOBAL")
            try:
                _entry = _store.load(args.name)
            except KeyError:
                print(f"[TOPOLOGY] '{args.name}' not found in GLOBAL library.")
            else:
                if not args.yes:
                    _c = input(f"[TOPOLOGY] Overwrite topology in '{args.project}' with '{args.name}'? [y/N] ")
                    if _c.strip().lower() not in ("y", "yes"):
                        print("[TOPOLOGY] Aborted.")
                        return
                from maccre_core.utils.path_resolver import get_maccre_root as _gmr  # noqa: PLC0415
                _dc = _gmr() / "__DATACENTER" / args.project / "02_Dynamic_Context"
                _dc.mkdir(parents=True, exist_ok=True)
                _topo_path = _dc / "topology.csv"
                _topo_rows: list[dict[str, str]] = _entry["topology_rows"]
                if _topo_rows:
                    with _topo_path.open("w", newline="", encoding="utf-8") as _fh:
                        _w = _csv.DictWriter(_fh, fieldnames=list(_topo_rows[0].keys()))
                        _w.writeheader()
                        _w.writerows(_topo_rows)
                _roster_rows: list[dict[str, str]] = _entry.get("roster_rows") or []
                if _roster_rows:
                    _roster_path = _gmr() / "__DATACENTER" / args.project / "agent_roster.csv"
                    with _roster_path.open("w", newline="", encoding="utf-8") as _fh:
                        _w2 = _csv.DictWriter(_fh, fieldnames=list(_roster_rows[0].keys()))
                        _w2.writeheader()
                        _w2.writerows(_roster_rows)
                print(f"[TOPOLOGY] '{args.name}' loaded into '{args.project}' ({len(_topo_rows)} nodes).")
        elif args.topo_action == "delete":
            _store = get_topology_store(getattr(args, "project", ""))
            try:
                _store.delete(args.name)
                print(f"[TOPOLOGY] Deleted '{args.name}' from library.")
            except KeyError as _e:
                print(f"[TOPOLOGY] {_e}")

if __name__ == "__main__":
    main()
