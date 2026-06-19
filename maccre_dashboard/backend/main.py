import uvicorn
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import sys
import shutil
from pathlib import Path
import zmq
from pydantic import BaseModel
from typing import List, Dict, Any

# Ensure maccre_core is in path
maccre_root = Path(__file__).resolve().parent.parent.parent
if str(maccre_root) not in sys.path:
    sys.path.insert(0, str(maccre_root))

app = FastAPI(title="MACCREv2 Omni-Dashboard API")

# Allow Vite frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # For local dev
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize ZMQ PUB socket for Swarm Interrupts
ctx = zmq.Context()
pub_socket = ctx.socket(zmq.PUB)
pub_socket.bind("tcp://127.0.0.1:5557")

@app.get("/api/health")
async def health_check() -> dict[str, str]:
    return {"status": "online", "message": "Omni-Dashboard API is running", "root": str(maccre_root)}

@app.post("/api/control/pause")
async def pause_swarm() -> dict[str, str]:
    # Broadcast PAUSE flag to all listening universal nodes
    import json
    payload = json.dumps({"command": "PAUSE", "source": "OmniDashboard"}).encode("utf-8")
    pub_socket.send_multipart([b"MACCRE.INTERRUPT", payload])
    return {"status": "paused"}

@app.post("/api/control/resume")
async def resume_swarm() -> dict[str, str]:
    # Broadcast RESUME flag
    import json
    payload = json.dumps({"command": "RESUME", "source": "OmniDashboard"}).encode("utf-8")
    pub_socket.send_multipart([b"MACCRE.INTERRUPT", payload])
    return {"status": "resumed"}

class ReactNode(BaseModel):
    id: str
    type: str
    position: Dict[str, float]
    data: Dict[str, Any]

class ReactEdge(BaseModel):
    source: str
    target: str

class TopologyPayload(BaseModel):
    nodes: List[ReactNode]
    edges: List[ReactEdge]

@app.post("/api/control/compile")
async def compile_topology(payload: TopologyPayload) -> dict[str, str]:
    try:
        from maccre_core.utils.path_resolver import get_maccre_root
        import openpyxl
        
        root = get_maccre_root()
        template_src = root / "templates" / "MACCRE_Swarm_Request.xlsx"
        
        if not template_src.exists():
            raise FileNotFoundError(f"Template not found at {template_src}")
            
        out_dir = root / "__DATACENTER" / "GLOBAL"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / "MACCRE_LiveSession.xlsx"
        
        shutil.copy2(template_src, out_path)
        
        wb = openpyxl.load_workbook(out_path)
        ws_topo = wb["TOPOLOGY"]
        
        # Clear rows
        for row in ws_topo.iter_rows(min_row=3):
            for cell in row:
                cell.value = None
                
        next_map = {e.source: e.target for e in payload.edges}
        incoming_targets = {e.target for e in payload.edges}
        
        start_nodes = set()
        row_idx = 3
        
        for node in payload.nodes:
            label = node.data.get("label", "Unassigned")
            node_type = node.data.get("nodeType", "AGENT")
            
            # Explicit legacy START_NODE
            if label == "START_NODE":
                target = next_map.get(node.id)
                if target:
                    start_nodes.add(target)
                continue
                
            # Implicit START_NODE detection
            if node.id not in incoming_targets:
                start_nodes.add(node.id)
                
            node_id = node.id
            next_node = next_map.get(node.id, "END")
            instruction = node.data.get("instruction", "")
            
            if node_type == "WAIT/RETRY" or label == "WAIT/RETRY":
                recursion = str(node.data.get("recursion", "3"))
                is_live = "FALSE"
                label = "WAIT/RETRY"
            elif node_type == "HUMAN_GATE" or label == "HUMAN_GATE":
                recursion = "0"
                is_live = "FALSE"
                label = "HUMAN_GATE"
            else:
                recursion = "3"
                is_live = "TRUE" if node.data.get("isLive") else "FALSE"
                
            ws_topo.cell(row=row_idx, column=1, value=node_id)
            ws_topo.cell(row=row_idx, column=2, value=label)
            ws_topo.cell(row=row_idx, column=3, value=node.data.get("model", "")) # Model
            ws_topo.cell(row=row_idx, column=4, value=next_node)
            ws_topo.cell(row=row_idx, column=5, value=node.data.get("temp", "")) # Temp
            ws_topo.cell(row=row_idx, column=6, value=instruction)
            ws_topo.cell(row=row_idx, column=7, value=node.data.get("wait_for", "none")) # Wait_For
            ws_topo.cell(row=row_idx, column=8, value=node.data.get("failure_target", "FAILED")) # Failure_Target
            ws_topo.cell(row=row_idx, column=9, value=recursion)
            ws_topo.cell(row=row_idx, column=10, value=node.data.get("payload", "")) # Artifact
            ws_topo.cell(row=row_idx, column=11, value=is_live)
            
            row_idx += 1
            
        start_node_target = ",".join(start_nodes) if start_nodes else "END"
        ws_req = wb["SWARM_REQUEST"]
        ws_req.cell(row=3, column=6, value=start_node_target) # col6 is START_NODE
        
        wb.save(out_path)
        wb.close()
        
        # Broadcast the Resume interrupt or run sub-process to ignite swarm
        import json
        msg = json.dumps({"command": "RESUME", "source": "OmniDashboard"}).encode("utf-8")
        pub_socket.send_multipart([b"MACCRE.INTERRUPT", msg])
        
        return {"status": "compiled", "file": str(out_path)}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

class NudgeRequest(BaseModel):
    job_id: str
    target_node: str
    payload_text: str

@app.post("/api/control/nudge")
def nudge_swarm(req: NudgeRequest) -> dict[str, str]:
    from maccre_core.orchestration.local_broker import LocalMessageBroker
    import json
    try:
        broker = LocalMessageBroker()
        
        target_prefix = f"[@NUDGE_TARGET:{req.target_node}] " if req.target_node and req.target_node != "ALL" else ""
        full_text = target_prefix + req.payload_text
        
        broker.inject_interrupt(job_id=req.job_id, override_text=full_text)
        
        msg = json.dumps({
            "command": "NUDGE",
            "source": "OmniDashboard",
            "target": req.target_node,
            "text": req.payload_text
        }).encode("utf-8")
        pub_socket.send_multipart([b"MACCRE.INTERRUPT", msg])
        
        return {"status": "success", "message": f"Nudged {req.target_node}"}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/projects")
def get_projects() -> dict[str, Any]:
    from maccre_core.utils.path_resolver import get_maccre_root
    import os
    datacenter = get_maccre_root() / "__DATACENTER"
    projects = []
    
    reserved_folders = {
        "chroma_db", ".vault", "01_Raw_Source", "02_Dynamic_Context", 
        "03_Agent_Ledgers", "04_Code_Artifacts", "05_Rendered_Media", 
        "__GLOBAL_LEDGER", "telemetry"
    }
    
    if datacenter.exists():
        projects = [
            d.name for d in datacenter.iterdir() 
            if d.is_dir() and d.name not in reserved_folders and not d.name.startswith(".")
        ]
    active = os.environ.get("MACCRE_ACTIVE_PROJECT", "GLOBAL")
    return {"active": active, "projects": projects}

class ProjectRequest(BaseModel):
    project_name: str

@app.post("/api/projects/new")
def new_project(req: ProjectRequest) -> dict[str, str]:
    from maccre_core.tools.admin_tools import initialize_workspace
    result = initialize_workspace(req.project_name)
    if "SUCCESS" in result:
        return {"status": "success", "message": result}
    raise HTTPException(status_code=400, detail=result)

@app.post("/api/projects/switch")
def switch_project(req: ProjectRequest) -> dict[str, str]:
    from maccre_core.tools.admin_tools import switch_workspace
    result = switch_workspace(req.project_name)
    if "SUCCESS" in result:
        return {"status": "success", "message": result}
    raise HTTPException(status_code=400, detail=result)

@app.get("/api/picker")
def open_picker(type: str = "folder") -> dict[str, str]:
    import subprocess
    import sys
    
    script = f"""
import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)
if '{type}' == 'file':
    path = filedialog.askopenfilename(title="Select Payload Context File")
else:
    path = filedialog.askdirectory(title="Select Payload Context Directory")
if path:
    print(path, end="")
"""
    try:
        result = subprocess.run([sys.executable, "-c", script], capture_output=True, text=True)
        return {"path": result.stdout.strip()}
    except Exception as e:
        return {"path": "", "error": str(e)}

@app.get("/api/rosters")
def get_rosters() -> dict[str, List[dict]]:
    from maccre_core.workbook_data import load_full_agent_rosters
    rosters = load_full_agent_rosters()
    return rosters

class MintAgentRequest(BaseModel):
    agent_name: str
    model: str
    tools: str
    system_prompt: str
    description: str
    thinking_level: str = "Default"
    temperature: float = 1.0
    media_resolution: str = "Default"
    output_length: int = 8192
    top_p: float = 0.95
    safety_settings: bool = False
    # These become the agent_json blob

@app.post("/api/agents/mint")
def mint_agent(req: MintAgentRequest) -> dict[str, str]:
    from maccre_core.utils.path_resolver import get_maccre_root
    from maccre_core.agent_library import get_agent_store
    import json
    
    try:
        if not req.agent_name:
            raise HTTPException(status_code=400, detail="agent_name is required.")
            
        # 1. Save to SQLite (Global Library)
        store = get_agent_store("GLOBAL") # hardcoded global
        agent_json = json.dumps({
            "thinking_level": req.thinking_level,
            "media_resolution": req.media_resolution,
            "output_length": req.output_length,
            "top_p": req.top_p,
            "description": req.description,
            "safety_settings": req.safety_settings
        })
        
        agent_dict = {
            "agent_name": req.agent_name,
            "model": req.model,
            "tools_allowed": req.tools,
            "system_prompt": req.system_prompt,
            "temperature": req.temperature,
            "agent_json": agent_json
        }
        store.save(agent_dict, source_project="GLOBAL")

        # 2. Keep the CSV mirrored for workbook_data.py compatibility (if needed)
        import csv
        datacenter = get_maccre_root() / "__DATACENTER"
        silo = datacenter / "GLOBAL"
        if not silo.exists():
            silo.mkdir(parents=True, exist_ok=True)
            
        roster_path = silo / "agent_roster.csv"
        file_exists = roster_path.exists()

        rows: list[list[str]] = []
        if file_exists:
            with open(roster_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if row and row[0] != req.agent_name:
                        rows.append(row)

        rows.append([req.agent_name, req.model, req.tools, req.system_prompt, req.description])

        with open(roster_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Agent_Name", "Model", "Tools_Allowed", "System_Prompt", "Description"])
            writer.writerows(rows)

        return {"status": "success", "message": f"Agent '{req.agent_name}' saved globally."}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/models")
def get_models() -> List[str]:
    from maccre_core.orchestration.windows_vault import get_native_credential
    from maccre_core._net.model_registry import get_registry
    import traceback
    
    try:
        api_key = get_native_credential("MACCRE_Sovereign")
        if not api_key:
            return ["ERROR: Missing MACCRE_Sovereign in Vault"]
        
        reg = get_registry(api_key)
        # Use cached probe via _maybe_refresh() instead of forcing a 22-second sync delay every time.
        models = reg.all_models()
        if not models:
            return [
                "gemini-3.1-pro-preview",
                "gemini-3.1-flash-lite-preview",
                "gemini-3-pro-preview",
                "gemini-3-flash-preview",
                "gemini-2.5-pro",
                "gemini-2.5-flash",
                "deep-research-max-preview-04-2026",
                "gemma-4-31b-it"
            ]
        return models
    except Exception as e:
        traceback.print_exc()
        return [f"ERROR: {e}"]
@app.get("/api/telemetry")
async def get_telemetry() -> List[str]:
    from maccre_core.utils.path_resolver import get_maccre_root
    log_path = get_maccre_root() / "__DATACENTER" / "GLOBAL" / "build_pipeline.log"
    if not log_path.exists():
        return []
        
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            # Return last 100 lines
            return [line.strip() for line in lines[-100:]]
    except Exception:
        return []

if __name__ == "__main__":
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
